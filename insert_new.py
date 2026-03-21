#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
insert_new.py

Incremental update with ONE new digit:
- Reads digit_string.txt (199 digits, youngest-first)
- Inserts new digit at HEAD (youngest-first), drops the oldest (tail)
- Recomputes Markov order=2 and order=3
- Overwrites markov.csv and markov.xlsx (ws M_Result), and opens markov.xlsx (macOS)

Usage:
  python3 insert_new.py 7
"""

import sys
import math
import subprocess
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
except Exception as e:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    raise

DIGITS_TXT = Path("digit_string.txt")
MARKOV_CSV = Path("markov.csv")
MARKOV_XLSX = Path("markov.xlsx")
SHEET = "M_Result"
TICKER = "DXC"  # label only

STATE_LEN_2 = 2
STATE_LEN_3 = 3


def compute_markov(digits_youngest_first: str, order_k: int):
    digits_chrono = list(digits_youngest_first[::-1])  # oldest->youngest

    transitions = []
    for i in range(len(digits_chrono) - order_k):
        s = "".join(digits_chrono[i : i + order_k])
        s_next = "".join(digits_chrono[i + 1 : i + order_k + 1])
        transitions.append((s, s_next))

    states = []
    seen = set()
    for s, s_next in transitions:
        if s not in seen:
            seen.add(s)
            states.append(s)
        if s_next not in seen:
            seen.add(s_next)
            states.append(s_next)

    idx = {s: i for i, s in enumerate(states)}
    n = len(states)
    counts = np.zeros((n, n), dtype=float)
    for s, s_next in transitions:
        counts[idx[s], idx[s_next]] += 1.0

    P = counts.copy()
    row_sums = P.sum(axis=1)
    for i in range(n):
        if row_sums[i] > 0:
            P[i, :] /= row_sums[i]
        else:
            P[i, i] = 1.0

    pi = np.ones(n, dtype=float) / n
    for _ in range(5000):
        pi_next = pi @ P
        if np.linalg.norm(pi_next - pi, ord=1) < 1e-12:
            pi = pi_next
            break
        pi = pi_next
    pi = pi / pi.sum()

    current_state = "".join(digits_chrono[-order_k:])

    if current_state in idx:
        row = P[idx[current_state], :]
        digit_probs = {str(d): 0.0 for d in range(1, 10)}
        for j, s_next in enumerate(states):
            digit_probs[s_next[-1]] += row[j]
        next_digit_probs = pd.Series(digit_probs).sort_index()
    else:
        digit_probs = {str(d): 0.0 for d in range(1, 10)}
        for j, s in enumerate(states):
            digit_probs[s[-1]] += pi[j]
        next_digit_probs = pd.Series(digit_probs).sort_index()

    predicted_digit = str(next_digit_probs.idxmax())
    return states, P, pi, current_state, next_digit_probs, predicted_digit


def write_markov_csv(path: Path, blocks: list):
    lines = []
    lines.append(f"GeneratedUTC,{datetime.utcnow().isoformat()}Z")
    for b in blocks:
        title = b["title"]
        states = b["states"]
        P = b["P"]
        pi = b["pi"]
        current_state = b["current_state"]
        next_digit_probs = b["next_digit_probs"]
        predicted_digit = b["predicted_digit"]

        lines.append("")
        lines.append(f"[{title}]")
        lines.append(f"StateCount,{len(states)}")
        lines.append(f"CurrentState,{current_state}")
        lines.append(f"PredictedNextDigit,{predicted_digit}")

        lines.append("NextDigitProbs," + ",".join([f"d{d}" for d in range(1, 10)]))
        lines.append("," + ",".join([f"{next_digit_probs[str(d)]:.10f}" for d in range(1, 10)]))

        lines.append("StationaryVector")
        lines.append("State,Pi")
        for s, val in zip(states, pi):
            lines.append(f"{s},{val:.12f}")

        lines.append("TransitionMatrix")
        header = "From\\To," + ",".join(states)
        lines.append(header)
        for i, s in enumerate(states):
            row = ",".join([f"{P[i, j]:.12f}" for j in range(len(states))])
            lines.append(f"{s},{row}")

    path.write_text("\n".join(lines), encoding="utf-8")


def write_markov_xlsx(path: Path, blocks: list):
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    bold = Font(bold=True)
    ws["A1"] = "Ticker"
    ws["B1"] = TICKER
    ws["A2"] = "GeneratedUTC"
    ws["B2"] = datetime.utcnow().isoformat() + "Z"
    ws["A1"].font = bold
    ws["A2"].font = bold

    r = 4
    for b in blocks:
        title = b["title"]
        states = b["states"]
        P = b["P"]
        pi = b["pi"]
        current_state = b["current_state"]
        next_digit_probs = b["next_digit_probs"]
        predicted_digit = b["predicted_digit"]

        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=14)
        r += 2

        ws.cell(row=r, column=1, value="StateCount").font = bold
        ws.cell(row=r, column=2, value=len(states))
        r += 1

        ws.cell(row=r, column=1, value="CurrentState").font = bold
        ws.cell(row=r, column=2, value=current_state)
        r += 1

        ws.cell(row=r, column=1, value="PredictedNextDigit").font = bold
        ws.cell(row=r, column=2, value=predicted_digit)
        r += 2

        ws.cell(row=r, column=1, value="NextDigitProbs").font = bold
        r += 1
        ws.cell(row=r, column=1, value="Digit").font = bold
        ws.cell(row=r, column=2, value="Prob").font = bold
        r += 1
        for d in range(1, 10):
            ws.cell(row=r, column=1, value=str(d))
            ws.cell(row=r, column=2, value=float(next_digit_probs[str(d)]))
            r += 1
        r += 1

        ws.cell(row=r, column=1, value="StationaryVector").font = bold
        r += 1
        ws.cell(row=r, column=1, value="State").font = bold
        ws.cell(row=r, column=2, value="Pi").font = bold
        r += 1
        for s, val in zip(states, pi):
            ws.cell(row=r, column=1, value=s)
            ws.cell(row=r, column=2, value=float(val))
            r += 1
        r += 1

        ws.cell(row=r, column=1, value="TransitionMatrix").font = bold
        r += 1

        ws.cell(row=r, column=1, value="From\\To").font = bold
        for j, s_to in enumerate(states, start=2):
            c = ws.cell(row=r, column=j, value=s_to)
            c.font = bold
            c.alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
        r += 1

        for i, s_from in enumerate(states):
            ws.cell(row=r, column=1, value=s_from).font = bold
            for j in range(len(states)):
                ws.cell(row=r, column=j + 2, value=float(P[i, j]))
            r += 1

        r += 3

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    wb.save(path)


def open_file_macos(path: Path):
    if sys.platform == "darwin":
        try:
            subprocess.run(["open", str(path)], check=False)
        except Exception:
            pass


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 insert_new.py <new_digit 1..9>")
        sys.exit(2)

    new_digit = str(sys.argv[1]).strip()
    if new_digit not in [str(d) for d in range(1, 10)]:
        print("ERROR: new_digit must be a single digit 1..9")
        sys.exit(2)

    if not DIGITS_TXT.exists():
        print("ERROR: digit_string.txt not found. Run markov.py first.")
        sys.exit(1)

    s = DIGITS_TXT.read_text(encoding="utf-8").strip()
    if len(s) != 199 or any(ch not in "123456789" for ch in s):
        print("ERROR: digit_string.txt must contain exactly 199 digits (1..9), youngest-first.")
        sys.exit(1)

    # Insert new youngest at HEAD, drop oldest at TAIL
    s_new = new_digit + s[:198]
    DIGITS_TXT.write_text(s_new, encoding="utf-8")

    blocks = []
    for k in (STATE_LEN_2, STATE_LEN_3):
        states, P, pi, current_state, next_digit_probs, predicted_digit = compute_markov(s_new, k)
        blocks.append(
            dict(
                title=f"ORDER_{k}",
                states=states,
                P=P,
                pi=pi,
                current_state=current_state,
                next_digit_probs=next_digit_probs,
                predicted_digit=predicted_digit,
            )
        )

    write_markov_csv(MARKOV_CSV, blocks)
    write_markov_xlsx(MARKOV_XLSX, blocks)
    open_file_macos(MARKOV_XLSX)

    print("Updated.")
    print("  - digit_string.txt (new digit inserted at head)")
    print("  - markov.csv")
    print("  - markov.xlsx")


if __name__ == "__main__":
    main()
