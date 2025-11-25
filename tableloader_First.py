#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from tempfile import NamedTemporaryFile

# ============================================================
# CONFIGURATION (FIXED)
# ============================================================

INPUT_CSV = "/Users/kostayanev/NikyClean/Transactions.csv"
OUTPUT_XLSX = "/Users/kostayanev/NikyClean/TRADES.xlsx"

SHEET_NAME  = "Transactions"
TABLE_NAME  = "TabTrans"

# FIXED list of data-entry columns (names match table header EXACTLY)
DATA_COLUMNS = ["Date", "Prefix", "Ticker", "Number", "Price"]

# ============================================================
# HELPERS
# ============================================================

def excel_bounds(ref):
    start, end = ref.split(":")
    return coordinate_to_tuple(start), coordinate_to_tuple(end)


def get_table_info(ws, table_name):
    table = ws.tables[table_name]
    (min_row, min_col), (max_row, max_col) = excel_bounds(table.ref)

    header_row = min_row
    col_names = [
        ws.cell(row=header_row, column=c).value
        for c in range(min_col, max_col + 1)
    ]

    return header_row, min_col, max_col, col_names


# ============================================================
# SAFE TABLE UPDATE — DATA COLUMNS ONLY
# ============================================================

def safe_update_table(csv_file, xlsx_file, sheet_name, table_name, data_cols):
    df = pd.read_csv(csv_file)

    wb = load_workbook(xlsx_file)
    ws = wb[sheet_name]

    header_row, min_col, max_col, col_names = get_table_info(ws, table_name)

    col_index = {name: i for i, name in enumerate(col_names)}

    first_data_row = header_row + 1
    n_rows = len(df)

    # ----- UPDATE ONLY DATA COLUMNS -----
    for i in range(n_rows):
        for col_name in data_cols:
            excel_col = min_col + col_index[col_name]
            ws.cell(row=first_data_row + i, column=excel_col).value = df.at[i, col_name]

    # ----- CLEAR OLD ROWS BELOW NEW DATA -----
    last_old_row = ws.max_row
    last_new_row = first_data_row + n_rows - 1

    for r in range(last_new_row + 1, last_old_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = None

    # ----- RESIZE TABLE -----
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(max_col)
    new_ref = f"{get_column_letter(min_col)}{header_row}:{last_col_letter}{last_new_row}"

    ws.tables[table_name].ref = new_ref

    return wb


# ============================================================
# ATOMIC SAVE
# ============================================================

def atomic_save(wb, final_path):
    path = os.path.dirname(final_path)
    with NamedTemporaryFile(dir=path, suffix=".tmp", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    os.replace(tmp_path, final_path)
    print(f"✔ Atomic save → {final_path}")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    wb = safe_update_table(INPUT_CSV, OUTPUT_XLSX, SHEET_NAME, TABLE_NAME, DATA_COLUMNS)
    atomic_save(wb, OUTPUT_XLSX)
    print("✔ Table update completed successfully.")