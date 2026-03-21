#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
import math
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400
K = 3

THRESHOLD_A = 1.0   # use best from your previous search

# ============================================================
# UTILITIES
# ============================================================

def read_tickers():
    wb = load_workbook(XLSX, data_only=True)
    tickers = []
    dn = wb.defined_names.get(TICKERS_NAME)

    if dn:
        sheet_name, cell_range = next(dn.destinations)
        ws = wb[sheet_name]
        for row in ws[cell_range]:
            for cell in row:
                if cell.value:
                    tickers.append(str(cell.value).strip())
    else:
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if not v:
                break
            tickers.append(str(v).strip())
            r += 1

    return list(dict.fromkeys(tickers))

def download_closes(t):
    df = yf.download(t, period=PERIOD, interval="1d",
                     auto_adjust=True, progress=False)
    s = df["Close"].dropna()
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:,0]
    return s

def quantize_3(pct, a):
    if pct <= -a: return 1
    if pct >=  a: return 3
    return 2

def build_digits(closes, a):
    closes = closes.tail(N_DIGITS + 1)
    pct = closes.pct_change().iloc[1:] * 100.0
    digits = [str(quantize_3(float(x), a)) for x in pct]
    return "".join(digits[::-1])

def entropy(counts):
    tot = counts.sum()
    if tot == 0: return 0
    p = counts / tot
    p = p[p>0]
    return float(-(p*np.log2(p)).sum())

def mutual_info(digits, order):
    chrono = digits[::-1]
    n = len(chrono)

    state_map = {}
    joint = []
    next_counts = np.zeros(K)

    def idx(s):
        if s not in state_map:
            state_map[s] = len(joint)
            joint.append(np.zeros(K))
        return state_map[s]

    for i in range(order, n):
        s = chrono[i-order:i]
        x = int(chrono[i]) - 1
        j = idx(s)
        joint[j][x] += 1
        next_counts[x] += 1

    Hx = entropy(next_counts)

    HxS = 0
    total = next_counts.sum()
    for row in joint:
        rs = row.sum()
        if rs>0:
            HxS += (rs/total) * entropy(row)

    MI = Hx - HxS
    return MI, Hx

# ============================================================
# MAIN
# ============================================================

def main():

    tickers = read_tickers()
    closes_map = {}

    for t in tickers:
        try:
            closes_map[t] = download_closes(t)
        except:
            pass

    for ORDER in (2,3):

        MI_sum = 0
        Hx_sum = 0
        used = 0

        for t in closes_map:

            digits = build_digits(closes_map[t], THRESHOLD_A)
            MI, Hx = mutual_info(digits, ORDER)

            MI_sum += MI
            Hx_sum += Hx
            used += 1

        MI_avg = MI_sum / used
        Hx_avg = Hx_sum / used
        NMI = MI_avg / Hx_avg if Hx_avg>0 else 0

        print(f"\nORDER {ORDER}")
        print(f"MI_avg   = {MI_avg:.4f} bits")
        print(f"Hx_avg   = {Hx_avg:.4f} bits")
        print(f"NMI_avg  = {NMI:.4f}")

    print("\nGeneratedUTC:", datetime.now(UTC).isoformat())

if __name__ == "__main__":
    main()