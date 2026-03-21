#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import math
import subprocess
import re
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import numbers, Font

# ============================================================
# SETTINGS
# ============================================================

WORKBOOK = Path("markov.xlsx")
SHEET = "M_Results"

DAYS = 200
PCT_FMT = numbers.FORMAT_PERCENTAGE

# ============================================================
# UTILITIES
# ============================================================

def is_valid_ticker(x: str) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip().upper()
    if x in {"TICKER", "GENERATEDUTC"}:
        return False
    return bool(re.search(r"[A-Z]", x))

# ============================================================
# QUANTIZATION (3 STATES)
# ============================================================

def quantize(pct):
    if math.isnan(pct):
        return ""
    if pct <= -1.0:
        return "1"
    if pct >= 1.0:
        return "3"
    return "2"

# ============================================================
# MARKOV CORE
# ============================================================

def markov(digits, order):
    chrono = digits[::-1]  # oldest → youngest

    states = set()
    for i in range(len(chrono) - order):
        states.add("".join(chrono[i:i+order]))
        states.add("".join(chrono[i+1:i+order+1]))
    states = sorted(states)

    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))
    for i in range(len(chrono) - order):
        a = "".join(chrono[i:i+order])
        b = "".join(chrono[i+1:i+order+1])
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        if C[i].sum():
            P[i] = C[i] / C[i].sum()
        else:
            P[i, i] = 1.0

    # stationary distribution
    pi = np.ones(n) / n
    for _ in range(5000):
        pi2 = pi @ P
        if np.linalg.norm(pi2 - pi, 1) < 1e-12:
            break
        pi = pi2
    pi /= pi.sum()

    # prediction
    current = "".join(chrono[-order:])
    probs = {"1": 0.0, "2": 0.0, "3": 0.0}
    row = P[idx[current]]
    for j, s in enumerate(states):
        probs[s[-1]] += row[j]

    return current, probs, states, pi

# ============================================================
# HEADER WRITERS
# ============================================================

def write_headers(ws, row, order, states):
    bold = Font(bold=True)

    ws.cell(row, 1, "Ticker").font = bold
    ws.cell(row, 2, f"Youngest_O{order}").font = bold
    ws.cell(row, 3, "P1").font = bold
    ws.cell(row, 4, "P2").font = bold
    ws.cell(row, 5, "P3").font = bold

    col = 6
    for s in states:
        ws.cell(row, col, f"π{s}").font = bold
        col += 1

# ============================================================
# MAIN
# ============================================================

def main():
    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET]

    # ---- read tickers ----
    tickers = []
    r = 2
    while True:
        v = ws[f"A{r}"].value
        if not v:
            break
        if is_valid_ticker(v):
            tickers.append(v.strip())
        r += 1

    if not tickers:
        raise RuntimeError("No valid tickers found")

    ws.delete_rows(2, ws.max_row)

    # ---- download data ----
    data = {}
    for t in tickers:
        df = yf.download(
            t,
            period="400d",
            interval="1d",
            auto_adjust=False,
            progress=False
        )
        df = df.tail(DAYS)
        df["Pct"] = df["Close"].pct_change() * 100
        df["D"] = df["Pct"].apply(quantize)
        digits = df.loc[df["D"] != "", "D"].tolist()
        if len(digits) >= 50:
            data[t] = digits

    # ========================================================
    # ORDER 2
    # ========================================================

    order = 2
    sample_states = markov(next(iter(data.values())), order)[2]
    write_headers(ws, 1, order, sample_states)

    row = 2
    for t in tickers:
        if t not in data:
            continue
        cur, probs, states, pi = markov(data[t], order)

        ws.cell(row, 1, t)
        ws.cell(row, 2, cur)
        ws.cell(row, 3, probs["1"])
        ws.cell(row, 4, probs["2"])
        ws.cell(row, 5, probs["3"])

        col = 6
        for p in pi:
            ws.cell(row, col, p)
            col += 1

        row += 1

    # gap + header row index
    row += 1
    header_o3 = row

    # ========================================================
    # ORDER 3
    # ========================================================

    order = 3
    sample_states = markov(next(iter(data.values())), order)[2]
    write_headers(ws, header_o3, order, sample_states)

    row = header_o3 + 1
    for t in tickers:
        if t not in data:
            continue
        cur, probs, states, pi = markov(data[t], order)

        ws.cell(row, 1, t)
        ws.cell(row, 2, cur)
        ws.cell(row, 3, probs["1"])
        ws.cell(row, 4, probs["2"])
        ws.cell(row, 5, probs["3"])

        col = 6
        for p in pi:
            ws.cell(row, col, p)
            col += 1

        row += 1

    # ---- format percentages ----
    for r in ws.iter_rows(min_row=2, min_col=3):
        for c in r:
            if isinstance(c.value, (int, float)):
                c.number_format = PCT_FMT

    wb.save(WORKBOOK)
    subprocess.run(["open", WORKBOOK], check=False)
    print("Done.")

# ============================================================

if __name__ == "__main__":
    main()