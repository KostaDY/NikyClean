#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from datetime import datetime, UTC
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

from markov_core import read_tickers_from_excel, XLSX

STATE = Path("digits_400_state.csv")
RES_SHEET = "M_Results"
PCT_FMT = numbers.FORMAT_PERCENTAGE


def build_markov(chrono: str, order: int):
    transitions = []
    for i in range(len(chrono) - order):
        a = chrono[i:i+order]
        b = chrono[i+1:i+order+1]
        transitions.append((a, b))

    # full state space
    states = sorted(set(a for a, _ in transitions) | set(b for _, b in transitions))
    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))
    for a, b in transitions:
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        s = C[i].sum()
        if s:
            P[i] = C[i] / s
        else:
            P[i, i] = 1.0

    return states, idx, P


def eval_fullsample(chrono: str, order: int, states, idx, P):
    correct = 0
    total = 0
    brier_sum = 0.0

    for i in range(order, len(chrono)):
        current = chrono[i-order:i]
        factual = chrono[i]

        if current not in idx:
            continue

        row = P[idx[current]]
        probs = {"1": 0.0, "2": 0.0, "3": 0.0}
        for j, s in enumerate(states):
            probs[s[-1]] += row[j]

        pred = max(probs, key=probs.get)
        if pred == factual:
            correct += 1

        y = {"1": 0, "2": 0, "3": 0}
        y[factual] = 1
        brier_sum += sum((probs[d] - y[d]) ** 2 for d in "123")

        total += 1

    acc = correct / total if total else 0.0
    brier = brier_sum / total if total else 0.0
    return total, correct, acc, brier


def main():
    tickers = read_tickers_from_excel(XLSX)

    if not STATE.exists():
        raise RuntimeError("digits_400_state.csv not found. Run phase1 first.")

    df_state = pd.read_csv(STATE, dtype={"Ticker": str, "Digits": str})
    df_state = df_state.set_index("Ticker")

    wb = load_workbook(XLSX)
    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)
    ws.append(["GeneratedUTC", datetime.now(UTC).isoformat(), "Mode", "FullSample"])
    ws.append([])

    # Header row
    ws.append(["Ticker", "Order", "Observations", "Correct", "Accuracy", "BrierScore"])
    for c in range(1, 7):
        ws.cell(ws.max_row, c).font = bold

    for t in tickers:
        if t not in df_state.index:
            continue
        digits_youngest = str(df_state.at[t, "Digits"])
        if any(ch not in "123" for ch in digits_youngest):
            continue

        chrono = digits_youngest[::-1]  # oldest -> youngest

        for order in (2, 3):
            states, idx, P = build_markov(chrono, order)
            obs, cor, acc, brier = eval_fullsample(chrono, order, states, idx, P)
            ws.append([t, order, obs, cor, acc, brier])

    # Format %
    for row in ws.iter_rows(min_row=4, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (float, int)):
                cell.number_format = PCT_FMT

    wb.save(XLSX)
    print("OK: wrote results to markov.xlsx / M_Results")


if __name__ == "__main__":
    main()