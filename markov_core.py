#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Shared utilities for the Markov multiphase workflow.

- Reads tickers from markov.xlsx named range TICKERS (fallback to M_Config!A2:A...)
- Downloads closes via yfinance
- Builds digit sequences
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import List, Tuple

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook


# -----------------------------
# Configuration
# -----------------------------
XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"


# -----------------------------
# Quantization
# -----------------------------
def quantize(pct: float) -> str:
    """Map daily percent change to digit."""
    if pct <= -1.0:
        return "1"
    if pct >= 1.0:
        return "3"
    return "2"


# -----------------------------
# Excel tickers
# -----------------------------
def read_tickers_from_excel(
    xlsx_path: Path = XLSX,
    cfg_sheet: str = CFG_SHEET,
    named_range: str = TICKERS_NAME,
    fallback_col: str = "A",
    fallback_start_row: int = 2,
    fallback_max_rows: int = 2000,
) -> List[str]:
    """
    Read tickers from workbook-level named range TICKERS.
    If that fails for any reason, fall back to M_Config!A2:A... until first blank.
    """
    wb = load_workbook(xlsx_path, data_only=True)

    tickers: List[str] = []

    # ---- Try named range first
    dn = None
    try:
        dn = wb.defined_names.get(named_range)
    except Exception:
        dn = None

    if dn is not None:
        try:
            dests = list(dn.destinations)
            if dests:
                sheet_name, cell_range = dests[0]
                ws = wb[sheet_name]
                for row in ws[cell_range]:
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.strip():
                            tickers.append(v.strip())
        except Exception:
            tickers = []

    # ---- Fallback if named range missing/empty
    if not tickers:
        if cfg_sheet not in wb.sheetnames:
            raise RuntimeError(f"Config sheet '{cfg_sheet}' not found in {xlsx_path.name}")

        ws = wb[cfg_sheet]
        col_idx = ord(fallback_col.upper()) - ord("A") + 1

        for r in range(fallback_start_row, fallback_start_row + fallback_max_rows):
            v = ws.cell(r, col_idx).value
            if v is None or (isinstance(v, str) and not v.strip()):
                break
            if isinstance(v, str):
                tickers.append(v.strip())
            else:
                tickers.append(str(v).strip())

    if not tickers:
        raise RuntimeError("No tickers found (named range empty and fallback column empty).")

    # Deduplicate preserving order
    seen = set()
    out = []
    for t in tickers:
        if t not in seen:
            out.append(t)
            seen.add(t)
    return out


# -----------------------------
# Yahoo download
# -----------------------------
def download_closes(ticker: str, period: str = "900d") -> pd.Series:
    """
    Download daily close prices for a ticker. Returns a Series indexed by date.

    period must be long enough to safely include >= 401 trading closes.
    """
    df = yf.download(
        ticker,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
        threads=True,
    )

    if df is None or df.empty:
        raise RuntimeError(f"{ticker}: empty download")

    closes = df["Close"].dropna()
    if isinstance(closes, pd.DataFrame):
        closes = closes.iloc[:, 0].dropna()
    if closes.empty:
        raise RuntimeError(f"{ticker}: no Close data")
    return closes


def build_digits_from_closes(closes: pd.Series, n_digits: int = 400) -> Tuple[str, pd.DataFrame]:
    """
    Build a youngest-first digit string length n_digits from close series.

    Need n_digits + 1 closes (because digits are differences between consecutive closes).
    Uses last (n_digits + 1) trading closes.
    Returns:
      - digits_youngest_first (string length n_digits)
      - dataframe with Date, Close, PctChange, Digit (chronological)
    """
    need = n_digits + 1
    closes = closes.dropna()
    if len(closes) < need:
        raise RuntimeError(f"Not enough closes: need {need}, have {len(closes)}")

    closes_tail = closes.tail(need)

    df = pd.DataFrame({"Close": closes_tail})
    df["PctChange"] = df["Close"].pct_change() * 100.0
    df = df.iloc[1:].copy()  # drop first NaN pct-change
    df["Digit"] = df["PctChange"].apply(lambda x: quantize(float(x)))
    df["Date"] = pd.to_datetime(df.index).date.astype(str)

    # chronological digits are df["Digit"] from oldest->youngest
    digits_chrono = df["Digit"].tolist()  # length n_digits
    digits_youngest_first = "".join(digits_chrono[::-1])

    return digits_youngest_first, df[["Date", "Close", "PctChange", "Digit"]]