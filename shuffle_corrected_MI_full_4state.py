#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400
ORDER = 3
K = 4

# Optimized thresholds from your search
B1 = -1.0
B2 =  0.0
B3 =  1.25

N_SHUFFLES = 50   # increase to 100+ for more precision

# ============================================================
# UTILITIES
# ============================================================

def read_tickers():
    wb = load_workbook(XLSX, data_only=True)
    tickers = []
    dn = wb.defined_names.get(TICKERS_NAME)

    if dn:
        try:
            sheet_name, cell_range = next(dn.destinations)
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value:
                        tickers.append(str(cell.value).strip())
        except:
            pass

    if not tickers:
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if not v:
                break
            tickers.append(str(v).strip())
            r += 1

    return list(dict.fromkeys(tickers))


def download_closes(t):
    try:
        df = yf.download(t, period=PERIOD,
                         interval="1d",
                         auto_adjust=True,
                         progress=False)
        if df.empty:
            return None
        s = df["Close"].dropna()
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:, 0]
        return s
    except:
        return None


def quantize_4(p):
    if p <= B1: return 1
    if p <= B2: return 2
    if p <= B3: return 3
    return 4


def entropy(counts):
    tot = counts.sum()
    if tot == 0:
        return 0.0
    p = counts / tot
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())


def mutual_info(digits):
    chrono = digits[::-1]
    n = len(chrono)

    state_map = {}
    joint = []
    next_counts = np.zeros(K)

    def idx(s):
        if s not in state_map:
            state_map[s] = len(joint)
            joint.append(np.zeros(K))
        return state_map[s]

    for i in range(ORDER, n):
        s = chrono[i-ORDER:i]
        x = int(chrono[i]) - 1
        j = idx(s)
        joint[j][x] += 1
        next_counts[x] += 1

    Hx = entropy(next_counts)

    HxS = 0
    total = next_counts.sum()
    for row in joint:
        rs = row.sum()
        if rs > 0:
            HxS += (rs/total) * entropy(row)

    MI = Hx - HxS
    return MI, Hx


# ============================================================
# MAIN
# ============================================================

def main():

    tickers = read_tickers()
    print("Tickers listed:", len(tickers))

    portfolio_MI = 0
    portfolio_bias = 0
    portfolio_corr = 0
    portfolio_z = 0
    used = 0

    for t in tickers:

        closes = download_closes(t)
        if closes is None or len(closes) < N_DIGITS+1:
            continue

        closes = closes.tail(N_DIGITS+1)
        pct = closes.pct_change().iloc[1:] * 100.0
        if len(pct) != N_DIGITS:
            continue

        digits = "".join(str(quantize_4(float(x))) for x in pct)[::-1]

        # REAL MI
        MI_real, Hx = mutual_info(digits)

        # SHUFFLE MI
        shuffle_vals = []
        arr = np.array(list(digits))

        for _ in range(N_SHUFFLES):
            np.random.shuffle(arr)
            shuffled = "".join(arr)
            MI_s, _ = mutual_info(shuffled)
            shuffle_vals.append(MI_s)

        shuffle_mean = np.mean(shuffle_vals)
        shuffle_std = np.std(shuffle_vals)

        MI_corr = MI_real - shuffle_mean
        z = MI_corr / shuffle_std if shuffle_std > 0 else 0

        print(f"{t:8s}  "
              f"MI_real={MI_real:.4f}  "
              f"Bias={shuffle_mean:.4f}  "
              f"MI_corr={MI_corr:.4f}  "
              f"Z={z:.2f}")

        portfolio_MI += MI_real
        portfolio_bias += shuffle_mean
        portfolio_corr += MI_corr
        portfolio_z += z
        used += 1

    print("\n================ PORTFOLIO ================")

    if used == 0:
        print("No valid tickers.")
        return

    print(f"Tickers used: {used}")
    print(f"Avg MI_real  : {portfolio_MI/used:.4f}")
    print(f"Avg Bias     : {portfolio_bias/used:.4f}")
    print(f"Avg MI_corr  : {portfolio_corr/used:.4f}")
    print(f"Avg Z-score  : {portfolio_z/used:.2f}")
    print("GeneratedUTC:", datetime.now(UTC).isoformat())


if __name__ == "__main__":
    main()