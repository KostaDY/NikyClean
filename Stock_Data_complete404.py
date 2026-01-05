import os
import pandas as pd
import pytz
import traceback
import yfinance as yf
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# SETTINGS
# ============================================================
WORKBOOK = "ASXData.xlsx"
INPUT_SHEET = "Tickers"
OUTPUT_SHEET = "LatestData"
TIMEZONE = pytz.timezone("Australia/Sydney")
HISTORY_PERIOD = "7d"
HISTORY_INTERVAL = "1d"
AVG_VOLUME_DAYS = 7
INTRADAY_PERIOD = "1d"
INTRADAY_INTERVAL = "1m"
SKIP_FULL_INFO_TICKERS = {"LOCK.L"}

# ============================================================
# VERIFY FILE
# ============================================================
if not os.path.exists(WORKBOOK):
    raise FileNotFoundError(f"Workbook '{WORKBOOK}' not found")

# ============================================================
# READ TICKERS
# ============================================================
tickers_df = pd.read_excel(WORKBOOK, sheet_name=INPUT_SHEET)
col_name = "Ticker" if "Ticker" in tickers_df.columns else "Tickers"

tickers = (
    tickers_df[col_name]
    .astype(str)
    .str.strip()
    .replace("", pd.NA)
    .dropna()
    .tolist()
)

if not tickers:
    raise ValueError("No tickers found")

print(f"{len(tickers)} tickers loaded")

# ============================================================
# FETCH DATA (ONE TICKER = ONE ROW, ALWAYS)
# ============================================================
rows = []

for ticker in tickers:
    print(f"Fetching {ticker}")
    row = {
        "Ticker": ticker,
        "Status": "OK",
        "MissingFields": ""
    }

    try:
        tk = yf.Ticker(ticker)
        info = tk.fast_info if hasattr(tk, "fast_info") else {}
        full = {}
        if ticker not in SKIP_FULL_INFO_TICKERS:
            try:
                full = tk.info or {}
            except Exception:
                full = {}

        # --- Market prices ---
        row["Open"] = (
            info.get("open", "")
            or full.get("regularMarketOpen", "")
            or full.get("open", "")
        )
        row["MarketPrice"] = info.get("lastPrice", "")
        row["MarketDayHigh"] = info.get("dayHigh", "")
        row["MarketDayLow"] = info.get("dayLow", "")
        row["MarketPreviousClose"] = info.get("previousClose", "")
        row["Volume"] = info.get("volume", "")
        row["MarketTime"] = (
            info.get("lastTradeTime", "")
            or full.get("regularMarketTime", "")
            or full.get("marketTime", "")
        )
        row["AverageVolume"] = ""

        # --- Valuation / fundamentals ---
        row["marketCap"] = full.get("marketCap", "")
        row["fiftyTwoWeekHigh"] = full.get("fiftyTwoWeekHigh", "")
        row["fiftyTwoWeekLow"] = full.get("fiftyTwoWeekLow", "")
        row["currency"] = full.get("currency", "")

        # --- Analyst data ---
        row["targetMeanPrice"] = full.get("targetMeanPrice", "")
        row["recommendationMean"] = full.get("recommendationMean", "")

        # --- Dividends ---
        row["dividendRate"] = full.get("dividendRate", "")
        row["dividendYield"] = full.get("dividendYield", "")
        row["exDividendDate"] = full.get("exDividendDate", "")

        # --- Earnings ---
        row["earningsDate"] = ""
        cal = tk.calendar
        if isinstance(cal, pd.DataFrame) and not cal.empty:
            # Some tickers return earnings dates as index, others as column; values can be lists
            dates = None
            if "Earnings Date" in cal.index:
                dates = cal.loc["Earnings Date"].tolist()
            elif "Earnings Date" in cal.columns:
                dates = cal["Earnings Date"].tolist()

            if dates:
                first = dates[0]
                if isinstance(first, (list, tuple)) and first:
                    first = first[0]
                row["earningsDate"] = first
        elif isinstance(cal, dict) and "Earnings Date" in cal:
            # Newer yfinance versions return a dict for calendar
            dates = cal.get("Earnings Date")
            if isinstance(dates, (list, tuple)) and dates:
                row["earningsDate"] = dates[0]
            elif dates:
                row["earningsDate"] = dates

        # --- Average volume, fallback volume, and fallback market time ---
        try:
            hist = tk.history(period=HISTORY_PERIOD, interval=HISTORY_INTERVAL)
            if isinstance(hist, pd.DataFrame) and not hist.empty:
                if row["Open"] in ("", None) and "Open" in hist.columns:
                    open_series = hist["Open"].dropna()
                    if not open_series.empty:
                        row["Open"] = open_series.iloc[-1]
                if "Volume" in hist.columns:
                    if row["Volume"] in ("", None) and not hist["Volume"].dropna().empty:
                        row["Volume"] = hist["Volume"].dropna().iloc[-1]
                    avg_vol = hist["Volume"].dropna().tail(AVG_VOLUME_DAYS).mean()
                    row["AverageVolume"] = avg_vol if pd.notna(avg_vol) else ""
                if row["MarketTime"] in ("", None) and len(hist.index) > 0:
                    row["MarketTime"] = hist.index[-1]
        except Exception:
            row["AverageVolume"] = ""

        if row["AverageVolume"] in ("", None):
            row["AverageVolume"] = full.get("averageVolume", "") or full.get("averageVolume10days", "")

        if row["MarketTime"] in ("", None):
            try:
                intraday = tk.history(period=INTRADAY_PERIOD, interval=INTRADAY_INTERVAL)
                if isinstance(intraday, pd.DataFrame) and not intraday.empty:
                    row["MarketTime"] = intraday.index[-1]
            except Exception:
                pass

        # --- Diagnostics ---
        missing = [k for k, v in row.items() if v in ("", None)]
        if len(missing) > 3:
            row["Status"] = "PARTIAL"
            row["MissingFields"] = ",".join(missing)

    except Exception as e:
        row["Status"] = "NO_DATA"
        row["MissingFields"] = str(e)

    rows.append(row)

# ============================================================
# BUILD DATAFRAME
# ============================================================
df = pd.DataFrame(rows)

# ============================================================
# TIME NORMALIZATION
# ============================================================
def normalize_time(x):
    try:
        if isinstance(x, pd.Timestamp):
            if x.tzinfo is None:
                return x.tz_localize("UTC").tz_convert(TIMEZONE)
            return x.tz_convert(TIMEZONE)
        if isinstance(x, datetime):
            ts = pd.Timestamp(x)
            if ts.tzinfo is None:
                ts = ts.tz_localize("UTC")
            return ts.tz_convert(TIMEZONE)
        if isinstance(x, (int, float)) and not pd.isna(x):
            unit = "ms" if x >= 1e12 else "s"
            return pd.to_datetime(x, unit=unit, utc=True).tz_convert(TIMEZONE)
        return pd.to_datetime(x, utc=True, errors="coerce").tz_convert(TIMEZONE)
    except Exception:
        return pd.NaT

if "MarketTime" in df.columns:
    df["MarketTime"] = df["MarketTime"].apply(normalize_time)
    df["MarketTime"] = df["MarketTime"].dt.strftime("%d-%b-%Y %H:%M")

def normalize_date(x):
    try:
        if x in ("", None):
            return pd.NaT
        if isinstance(x, (int, float)) and not pd.isna(x):
            if x == 0:
                return pd.NaT
            unit = "ms" if x >= 1e12 else "s"
            return pd.to_datetime(x, unit=unit, utc=True)
        return pd.to_datetime(x, utc=True, errors="coerce")
    except Exception:
        return pd.NaT

for col in ["exDividendDate", "earningsDate"]:
    if col in df.columns:
        df[col] = df[col].apply(normalize_date).dt.strftime("%d-%b-%Y")

df = df.fillna("")
df.columns = [str(col) for col in df.columns]

# ============================================================
# EXCEL TABLE HELPERS
# ============================================================
def add_or_replace_table(ws, table_name):
    try:
        if hasattr(ws, "tables") and table_name in ws.tables:
            ws.tables.pop(table_name, None)
    except Exception:
        pass

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

# ============================================================
# WRITE TO EXCEL
# ============================================================
try:
    with pd.ExcelWriter(WORKBOOK, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
        ws = writer.sheets.get(OUTPUT_SHEET)
        if ws is not None:
            add_or_replace_table(ws, "RTdata")
    print(f"Updated '{OUTPUT_SHEET}' with {len(df)} rows")
except Exception:
    print("Excel writing failed:")
    traceback.print_exc()

# ============================================================
# SUMMARY
# ============================================================
print("\n=== SUMMARY ===")
print(f"Tickers processed : {len(tickers)}")
print(f"Rows written      : {len(df)}")
print(df["Status"].value_counts())
print(f"Workbook          : {os.path.abspath(WORKBOOK)}")
