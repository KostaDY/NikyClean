import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, timedelta

# ============================================================
# SETTINGS
# ============================================================

INPUT_WB    = "Solid_REG_ALL.xlsm"
NAMED_RANGE = "Register"

OUTPUT_WB = "priceFIFO.xlsx"
OUTPUT_WS = "PriceFIFO"

# ============================================================
# DATE NORMALIZATION
# ============================================================

def normalize_excel_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(v))
    return pd.to_datetime(v, errors="coerce")

# ============================================================
# LOAD NAMED RANGE (READ-ONLY)
# ============================================================

wb = load_workbook(INPUT_WB, data_only=True, read_only=True)

if NAMED_RANGE not in wb.defined_names:
    raise ValueError(f'Named range "{NAMED_RANGE}" not found')

defn = wb.defined_names[NAMED_RANGE]
sheet_name, ref = next(defn.destinations)
ws = wb[sheet_name]

rows = [[cell.value for cell in row] for row in ws[ref]]

# ============================================================
# PARSE TICKER GROUPS
# ============================================================

groups = {}
current_ticker = None

for r in rows:
    if len(r) < 3:
        continue

    d, n, p = r[0], r[1], r[2]

    # Ticker header
    if isinstance(d, str) and isinstance(n, str) and isinstance(p, str):
        current_ticker = d.strip()
        groups[current_ticker] = []
        continue

    if current_ticker and d is not None:
        groups[current_ticker].append((
            normalize_excel_date(d),
            int(n),
            float(p)
        ))

# ============================================================
# FIFO (MODEL F) + BEHAVIORAL VIOLATIONS (MODEL V)
# ============================================================

def fifo_and_behavioral_viol(trades):
    df = pd.DataFrame(trades, columns=["Date", "Qty", "Price"])

    purchases = df[df["Price"] < 0].copy()
    sales     = df[df["Price"] > 0].copy()

    purchases["Cost"] = -purchases["Price"]
    purchases.sort_values(["Cost", "Date"], inplace=True)
    sales.sort_values("Date", inplace=True)

    purchases["Remaining"] = purchases["Qty"]
    sales["Remaining"]     = sales["Qty"]

    viol_qty   = 0
    viol_value = 0.0

    # ----- Behavioral violation pass (independent of FIFO matching)
    for _, s in sales.iterrows():
        available = purchases[purchases["Remaining"] > 0]
        if available.empty:
            continue

        vwap = (available["Cost"] * available["Remaining"]).sum() / available["Remaining"].sum()

        if s["Price"] < vwap:
            viol_qty   += s["Qty"]
            viol_value += (vwap - s["Price"]) * s["Qty"]

        # simulate inventory availability decay
        qty = s["Qty"]
        for i in purchases.index:
            take = min(qty, purchases.at[i, "Remaining"])
            purchases.at[i, "Remaining"] -= take
            qty -= take
            if qty == 0:
                break

    # ----- Reset for FIFO reconstruction
    purchases["Remaining"] = purchases["Qty"]
    sales["Remaining"]     = sales["Qty"]

    p_idx = s_idx = 0
    while p_idx < len(purchases) and s_idx < len(sales):
        take = min(purchases.iloc[p_idx]["Remaining"], sales.iloc[s_idx]["Remaining"])
        purchases.iat[p_idx, purchases.columns.get_loc("Remaining")] -= take
        sales.iat[s_idx, sales.columns.get_loc("Remaining")] -= take

        if purchases.iloc[p_idx]["Remaining"] == 0:
            p_idx += 1
        if sales.iloc[s_idx]["Remaining"] == 0:
            s_idx += 1

    remaining = purchases[purchases["Remaining"] > 0][
        ["Date", "Remaining", "Price"]
    ].rename(columns={"Remaining": "Number"})

    summary = {
        "TotalPurchaseQty": int(df[df["Price"] < 0]["Qty"].sum()),
        "TotalSaleQty": int(df[df["Price"] > 0]["Qty"].sum()),
        "RemainingQty": int(remaining["Number"].sum()),
        "BehavioralViolationQty": int(viol_qty),
        "BehavioralViolationValue": round(viol_value, 2),
    }

    return remaining, summary

# ============================================================
# BUILD OUTPUT SHEET
# ============================================================

out_rows = []

for ticker, trades in groups.items():
    if not trades:
        continue

    out_rows.append([ticker, None, None])

    fifo_df, summary = fifo_and_behavioral_viol(trades)

    for _, r in fifo_df.iterrows():
        out_rows.append([r["Date"], int(r["Number"]), float(r["Price"])])

    out_rows.append(["MODEL F — Global Price-FIFO (canonical)", None, None])
    out_rows.append(["MODEL V — Behavioral Violation Metrics", None, None])

    for k, v in summary.items():
        out_rows.append([k, v, None])

    out_rows.append([None, None, None])

# ============================================================
# WRITE OUTPUT
# ============================================================

out_df = pd.DataFrame(out_rows, columns=["Date", "Number", "Price"])

with pd.ExcelWriter(OUTPUT_WB, engine="openpyxl", mode="w") as writer:
    out_df.to_excel(writer, sheet_name=OUTPUT_WS, index=False)

print(f"✅ Written: {Path(OUTPUT_WB).resolve()}")