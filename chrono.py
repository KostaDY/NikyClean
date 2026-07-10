#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ChronoCube snapshot updater.

The program:

1. Recalculates and saves the active Excel workbook.
2. Reads the Excel table "TabCalc" from Entry_TICK.xlsm.
3. Reads the trading date from the workbook-level named range "TradingDate".
4. Stores the current TabCalc values in a 16-slot circular history cube.
5. If the program is run again for the same TradingDate, it overwrites the
   previous snapshot instead of advancing to another slot.
6. Saves:
      - ChronoCube.csv
      - ChronoCube_meta.json
      - ChronoLines.xlsx

Important:
The named range "TradingDate" must refer to one Excel cell containing a valid
Excel date or date-like value.
"""

import json
from pathlib import Path
import subprocess

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


# ==========================================================
# CONFIG
# ==========================================================

WORKBOOK = Path("Entry_TICK.xlsm")
SHEET_NAME = "Calc"
TABLE_NAME = "TabCalc"

TRADING_DATE_NAME = "TradingDate"

CSV_FILE = Path("ChronoCube.csv")
META_FILE = Path("ChronoCube_meta.json")
EXCEL_FILE = Path("ChronoLines.xlsx")

MAX_SLOTS = 16


# ==========================================================
# RECALCULATE EXCEL
# ==========================================================

def recalc_excel():
    """
    Recalculate and save the currently active Microsoft Excel workbook.
    """

    script = '''
    tell application "Microsoft Excel"
        calculate full
        save active workbook
    end tell
    '''

    result = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True
    )

    if result.returncode != 0:
        raise RuntimeError(
            "Excel recalculation failed:\n"
            f"{result.stderr.strip()}"
        )


# ==========================================================
# NORMALIZE
# ==========================================================

def normalize(v):
    """
    Normalize values read from Excel.

    Blank cells and Excel error strings become NaN.
    Percentage text is converted to decimal form.
    Numeric text is converted to float.
    """

    if v is None:
        return np.nan

    if isinstance(v, str):
        s = v.strip()

        if s == "":
            return np.nan

        if s.startswith("#"):
            return np.nan

        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100
            except ValueError:
                return s

        try:
            return float(s)
        except ValueError:
            return s

    return v


# ==========================================================
# READ TABCALC
# ==========================================================

def read_tabcalc():
    """
    Read the Excel table TabCalc into a pandas DataFrame.
    """

    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(
        WORKBOOK,
        data_only=True,
        read_only=False
    )

    try:
        if SHEET_NAME not in wb.sheetnames:
            raise KeyError(
                f'Worksheet "{SHEET_NAME}" was not found in {WORKBOOK}.'
            )

        ws = wb[SHEET_NAME]

        if TABLE_NAME not in ws.tables:
            raise KeyError(
                f'Excel table "{TABLE_NAME}" was not found '
                f'on worksheet "{SHEET_NAME}".'
            )

        ref = ws.tables[TABLE_NAME].ref

        min_col, min_row, max_col, max_row = range_boundaries(ref)

        headers = [
            str(ws.cell(row=min_row, column=c).value)
            for c in range(min_col, max_col + 1)
        ]

        rows = []

        for r in range(min_row + 1, max_row + 1):
            row_values = [
                normalize(ws.cell(row=r, column=c).value)
                for c in range(min_col, max_col + 1)
            ]

            rows.append(row_values)

    finally:
        wb.close()

    df = pd.DataFrame(rows, columns=headers)

    if df.empty:
        raise ValueError(f'Excel table "{TABLE_NAME}" contains no data rows.')

    # Remove a final totals row, if present.
    if df.iloc[-1].astype(str).str.contains(
        "Total",
        case=False,
        na=False
    ).any():
        df = df.iloc[:-1].reset_index(drop=True)

    if df.empty:
        raise ValueError(
            f'Excel table "{TABLE_NAME}" contains no usable data rows.'
        )

    return df


# ==========================================================
# READ TRADING DATE
# ==========================================================

def read_trading_date():
    """
    Read the trading date from the workbook-level named range TradingDate.

    Returns:
        Date in ISO format: YYYY-MM-DD
    """

    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(
        WORKBOOK,
        data_only=True,
        read_only=False
    )

    try:
        defined_name = wb.defined_names.get(TRADING_DATE_NAME)

        if defined_name is None:
            raise KeyError(
                f'Workbook-level named range "{TRADING_DATE_NAME}" '
                f'was not found in {WORKBOOK}.'
            )

        destinations = list(defined_name.destinations)

        if len(destinations) != 1:
            raise ValueError(
                f'Named range "{TRADING_DATE_NAME}" must refer to '
                f'exactly one cell.'
            )

        sheet_name, cell_reference = destinations[0]

        # Remove Excel quoting around sheet names, if present.
        sheet_name = sheet_name.strip("'")

        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f'Named range "{TRADING_DATE_NAME}" refers to missing '
                f'worksheet "{sheet_name}".'
            )

        value = wb[sheet_name][cell_reference].value

    finally:
        wb.close()

    if value is None or str(value).strip() == "":
        raise ValueError(
            f'Named range "{TRADING_DATE_NAME}" is empty.'
        )

    try:
        trading_date = pd.to_datetime(value, errors="raise").date()
    except Exception as exc:
        raise ValueError(
            f'Named range "{TRADING_DATE_NAME}" does not contain '
            f'a valid date: {value!r}'
        ) from exc

    return trading_date.isoformat()


# ==========================================================
# LOAD META
# ==========================================================

def load_meta():
    """
    Read the next-write slot and slot dates from the metadata JSON file.
    """

    if not META_FILE.exists():
        return 0, [""] * MAX_SLOTS

    try:
        with open(META_FILE, "r", encoding="utf-8") as f:
            meta = json.load(f)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Invalid JSON in metadata file: {META_FILE}"
        ) from exc

    slot = meta.get("slot")
    dates = meta.get("dates")

    if not isinstance(slot, int):
        raise ValueError(
            f'Invalid "slot" value in {META_FILE}: {slot!r}'
        )

    if not 0 <= slot < MAX_SLOTS:
        raise ValueError(
            f'"slot" must be between 0 and {MAX_SLOTS - 1}. '
            f"Current value: {slot}"
        )

    if not isinstance(dates, list):
        raise ValueError(
            f'Invalid "dates" value in {META_FILE}.'
        )

    if len(dates) != MAX_SLOTS:
        raise ValueError(
            f'"dates" must contain exactly {MAX_SLOTS} entries. '
            f"Current number: {len(dates)}"
        )

    dates = [
        "" if value is None else str(value)
        for value in dates
    ]

    return slot, dates


# ==========================================================
# SAVE META
# ==========================================================

def save_meta(slot, dates):
    """
    Save the next-write slot and slot dates.
    """

    meta = {
        "slot": slot,
        "dates": dates
    }

    with open(META_FILE, "w", encoding="utf-8") as f:
        json.dump(
            meta,
            f,
            indent=2,
            ensure_ascii=False
        )


# ==========================================================
# LOAD / CREATE CUBE
# ==========================================================

def load_cube(n_param, n_row):
    """
    Load the existing cube from CSV or create an empty cube.
    """

    empty_cube = np.full(
        (n_param, n_row, MAX_SLOTS),
        np.nan,
        dtype=object
    )

    if not CSV_FILE.exists():
        return empty_cube

    try:
        df = pd.read_csv(CSV_FILE)
    except Exception as exc:
        raise RuntimeError(
            f"Could not read {CSV_FILE}"
        ) from exc

    expected_columns = 2 + MAX_SLOTS

    if df.shape[1] != expected_columns:
        print("CSV structure mismatch -> rebuilding cube")
        return empty_cube

    expected_rows = n_param * n_row

    if len(df) != expected_rows:
        print("Cube row count mismatch -> rebuilding cube")
        return empty_cube

    cube = df.iloc[:, 2:].to_numpy(dtype=object)

    try:
        cube = cube.reshape(
            n_param,
            n_row,
            MAX_SLOTS
        )
    except ValueError:
        print("Cube size mismatch -> rebuilding cube")
        return empty_cube

    return cube


# ==========================================================
# SAVE CSV
# ==========================================================

def save_cube_csv(cube, params, rows):
    """
    Save the complete cube to CSV.
    """

    records = []

    for p, parameter in enumerate(params):
        for r, row_number in enumerate(rows):
            record = [
                parameter,
                row_number
            ]

            record.extend(cube[p, r, :])
            records.append(record)

    columns = (
        ["Parameter", "Row"]
        + [f"Slot{i}" for i in range(MAX_SLOTS)]
    )

    df = pd.DataFrame(
        records,
        columns=columns
    )

    df.to_csv(
        CSV_FILE,
        index=False
    )


# ==========================================================
# EXPORT XLSX
# ==========================================================

def export_excel(cube, params, rows, dates, slot):
    """
    Export the cube in chronological order.

    The supplied slot is the next slot that will be written. Therefore,
    slot is also the beginning of the chronological circular-buffer view.
    """

    if slot == 0:
        chrono_cube = cube
        chrono_dates = dates
    else:
        chrono_cube = np.concatenate(
            (
                cube[:, :, slot:],
                cube[:, :, :slot]
            ),
            axis=2
        )

        chrono_dates = dates[slot:] + dates[:slot]

    with pd.ExcelWriter(
        EXCEL_FILE,
        engine="openpyxl"
    ) as writer:

        start_row = 0

        for p, parameter in enumerate(params):

            block = pd.DataFrame(
                chrono_cube[p, :, :].T,
                columns=[f"T{r}" for r in rows]
            ).T

            block.columns = chrono_dates
            block.insert(0, "Ticker", rows)

            title = pd.DataFrame(
                [[f"Parameter: {parameter}"]]
            )

            title.to_excel(
                writer,
                startrow=start_row,
                index=False,
                header=False
            )

            start_row += 1

            block.to_excel(
                writer,
                startrow=start_row,
                index=False
            )

            start_row += len(block) + 3

    subprocess.run(
        ["open", str(EXCEL_FILE)],
        check=False
    )


# ==========================================================
# MAIN
# ==========================================================

def main():

    print("Recalculating Excel...")
    recalc_excel()

    print("Reading TabCalc...")
    df = read_tabcalc()

    print(f'Reading named range "{TRADING_DATE_NAME}"...')
    snapshot_date = read_trading_date()

    params = list(df.columns)
    rows = list(range(1, len(df) + 1))

    n_param = len(params)
    n_row = len(rows)

    print("Parameters:", n_param)
    print("Rows:", n_row)
    print("Trading date:", snapshot_date)

    cube = load_cube(n_param, n_row)
    slot, dates = load_meta()

    snapshot = df.to_numpy(dtype=object).T

    if snapshot.shape != (n_param, n_row):
        raise ValueError(
            "Snapshot dimensions do not match the cube dimensions."
        )

    # The metadata slot always identifies the next slot to be used.
    previous_slot = (slot - 1) % MAX_SLOTS

    if dates[previous_slot] == snapshot_date:
        # Secondary run for the same trading date:
        # overwrite the previously stored snapshot and do not advance slot.
        write_slot = previous_slot

        print(
            f"Trading date {snapshot_date} already exists "
            f"in slot {write_slot}."
        )
        print("Overwriting the existing snapshot.")
    else:
        # First run for this trading date:
        # write to the next slot and advance the circular buffer.
        write_slot = slot

        print(
            f"Creating snapshot for {snapshot_date} "
            f"in slot {write_slot}."
        )

        slot = (slot + 1) % MAX_SLOTS

    cube[:, :, write_slot] = snapshot
    dates[write_slot] = snapshot_date

    save_cube_csv(
        cube,
        params,
        rows
    )

    save_meta(
        slot,
        dates
    )

    export_excel(
        cube,
        params,
        rows,
        dates,
        slot
    )

    print()
    print("Cube updated")
    print("Cube shape:", cube.shape)
    print("Written slot:", write_slot)
    print("Next slot:", slot)
    print("Trading date:", snapshot_date)
    print("CSV file written:", CSV_FILE)
    print("Metadata file written:", META_FILE)
    print("Excel file written:", EXCEL_FILE)


if __name__ == "__main__":
    main()