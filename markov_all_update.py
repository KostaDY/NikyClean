#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Multi-Ticker Daily Markov Engine
Now includes:
- Backward in-sample evaluation
- Accuracy counts
- Brier score
"""

# ============================================================
# IMPORTS
# ============================================================

import math
import shutil
import subprocess
from pathlib import Path
from datetime import datetime, UTC

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

# ============================================================
# FILES & SETTINGS
# ============================================================

CSV  = Path("markov.csv")
XLSX = Path("markov.xlsx")

CFG_SHEET = "M_Config"
RES_SHEET = "M_Results"
print("Workbook path used:", XLSX.resolve())
LOOKBACK_DAYS = 10
SEED_DAYS = 260

PCT_FMT = numbers.FORMAT_PERCENTAGE

# ============================================================
# QUANTIZATION
# ============================================================

def quantize(pct: float) -> str:
    if pct <= -1.0:
        return "1"
    if pct >= 1.0:
        return "3"
    return "2"

# ============================================================
# DATA DOWNLOAD
# ============================================================

def download_recent_closes(ticker: str, days: int):
    df = yf.download(
        ticker,
        period=f"{days}d",
        interval="1d",
        auto_adjust=True,
        progress=False
    )
    closes = df["Close"].dropna()
    if isinstance(closes, pd.DataFrame):
        closes = closes.iloc[:, 0]
    return closes

# ============================================================
# SEEDING
# ============================================================

def seed_digits_from_history(ticker: str) -> str:
    closes = download_recent_closes(ticker, SEED_DAYS).tail(200)
    pct = closes.pct_change() * 100.0
    digits = [quantize(float(x)) for x in pct.iloc[1:]]
    return "".join(digits[::-1])  # youngest first

# ============================================================
# MARKOV BUILD
# ============================================================

def build_markov(digits_youngest_first: str, order: int):
    chrono = digits_youngest_first[::-1]
    transitions = []

    for i in range(len(chrono) - order):
        a = chrono[i:i+order]
        b = chrono[i+1:i+order+1]
        transitions.append((a, b))

    # FULL state space (a ∪ b)
    states = sorted(set(a for a, _ in transitions) |
                    set(b for _, b in transitions))

    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))

    for a, b in transitions:
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        if C[i].sum():
            P[i] = C[i] / C[i].sum()
        else:
            P[i, i] = 1.0

    return chrono, states, idx, P

# ============================================================
# BACKTEST
# ============================================================

def evaluate_backtest(chrono, states, idx, P, order):
    correct = 0
    total = 0
    brier_total = 0.0

    for i in range(order, len(chrono)-1):
        current = chrono[i-order:i]
        if current not in idx:
            continue

        probs = {"1":0.0,"2":0.0,"3":0.0}
        row = P[idx[current]]

        for j, s in enumerate(states):
            probs[s[-1]] += row[j]

        factual = chrono[i]

        pred = max(probs, key=probs.get)
        if pred == factual:
            correct += 1

        y = {"1":0,"2":0,"3":0}
        y[factual] = 1
        brier_total += sum((probs[d]-y[d])**2 for d in "123")

        total += 1

    accuracy = correct/total if total else 0
    brier = brier_total/total if total else 0

    return total, correct, accuracy, brier

# ============================================================
# STATE MANAGEMENT
# ============================================================

def read_state():
    state = {}
    cur = None
    if not CSV.exists():
        return state

    for line in CSV.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith("[TICKER"):
            cur = line.split(",",1)[1].rstrip("]")
            state[cur] = {}
        elif cur and "," in line:
            k,v = line.split(",",1)
            state[cur][k] = v
    return state

def write_state(state):
    rows = []
    for t,s in state.items():
        rows += [
            f"[TICKER,{t}]",
            f"DIGITS,{s['DIGITS']}",
            ""
        ]
    CSV.write_text("\n".join(rows))

# ============================================================
# MAIN
# ============================================================

def main():

    # backup CSV
    if CSV.exists():
        ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        shutil.copy2(CSV, CSV.with_suffix(f".csv.bak_{ts}"))

    wb = load_workbook(XLSX)

    dn = wb.defined_names.get("TICKERS")
    sheet_name, cell_range = next(dn.destinations)
    ws_rng = wb[sheet_name]

    tickers = [
        cell.value.strip()
        for row in ws_rng[cell_range]
        for cell in row
        if isinstance(cell.value,str) and cell.value.strip()
    ]

    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)

    ws.append(["GeneratedUTC", datetime.now(UTC).isoformat()])
    ws.append([])

    state = read_state()

    row = 3

    for t in tickers:

        if t not in state:
            state[t] = {"DIGITS": seed_digits_from_history(t)}

        digits = state[t]["DIGITS"]

        ws.cell(row,1,t).font = Font(bold=True,size=14)
        row += 2

        for order in (2,3):

            chrono, states, idx, P = build_markov(digits, order)
            total, correct, acc, brier = evaluate_backtest(
                chrono, states, idx, P, order
            )

            ws.cell(row,1,f"ORDER {order}").font = bold
            row += 1

            ws.append(["Observations", total])
            ws.append(["Correct", correct])
            ws.append(["Accuracy", acc])
            ws.append(["BrierScore", brier])
            row += 4

        row += 2

    write_state(state)
    wb.save(XLSX)
    subprocess.run(["open", str(XLSX)], check=False)

    print("Daily multi-ticker Markov evaluation completed.")

# ============================================================

if __name__ == "__main__":
    main()