#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime, UTC
import math
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"          # enough to get stable history
N_DIGITS = 400           # digits length
WINDOW = 200             # rolling train window
ORDER = 2                # set 2 or 3 (recommend 2 for 5-state)
ALPHA = 1e-3             # stronger smoothing helps stability

K = 5  # 5 states

# Portfolio weights: equal by default
WEIGHTS = None  # dict ticker->weight, or None for equal

# Grid search for symmetric 5-state: [-a, -c, +c, +a], with a>c>0
A_GRID = np.arange(1.0, 4.1, 0.25)   # outer threshold (%)
C_GRID = np.arange(0.25, 2.1, 0.25)  # inner threshold (%)

# =========================
# Helpers
# =========================
def read_tickers_from_excel(
    xlsx_path: Path,
    cfg_sheet: str,
    named_range: str,
    fallback_col: str = "A",
    fallback_start_row: int = 2,
    fallback_max_rows: int = 2000,
) -> list[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    tickers: list[str] = []

    dn = None
    try:
        dn = wb.defined_names.get(named_range)
    except Exception:
        dn = None

    if dn is not None:
        try:
            dests = list(dn.destinations)
            if dests:
                sheet_name, cell_range = dests[0]
                ws = wb[sheet_name]
                for row in ws[cell_range]:
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.strip():
                            tickers.append(v.strip())
        except Exception:
            tickers = []

    if not tickers:
        if cfg_sheet not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{cfg_sheet}' not found in {xlsx_path}")
        ws = wb[cfg_sheet]
        col_idx = ord(fallback_col.upper()) - ord("A") + 1
        for r in range(fallback_start_row, fallback_start_row + fallback_max_rows):
            v = ws.cell(r, col_idx).value
            if v is None or (isinstance(v, str) and not v.strip()):
                break
            tickers.append(str(v).strip())

    # de-dupe preserve order
    seen = set()
    out = []
    for t in tickers:
        if t and t not in seen:
            out.append(t)
            seen.add(t)
    if not out:
        raise RuntimeError("No tickers found.")
    return out


def download_closes(ticker: str, period: str) -> pd.Series:
    df = yf.download(ticker, period=period, interval="1d", auto_adjust=True, progress=False, threads=True)
    if df is None or df.empty:
        raise RuntimeError(f"{ticker}: empty data")
    s = df["Close"].dropna()
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0].dropna()
    if s.empty:
        raise RuntimeError(f"{ticker}: no Close")
    return s


def quantize_5_fixed(pct: float, b1: float, b2: float, b3: float, b4: float) -> str:
    # boundaries must satisfy b1<b2<b3<b4
    if pct <= b1:
        return "1"
    if pct <= b2:
        return "2"
    if pct < b3:
        return "3"
    if pct < b4:
        return "4"
    return "5"


def build_digits_from_thresholds(closes: pd.Series, b: tuple[float,float,float,float], n_digits: int) -> str:
    need = n_digits + 1
    closes = closes.dropna()
    if len(closes) < need:
        raise RuntimeError(f"Need {need} closes, have {len(closes)}")
    closes = closes.tail(need)
    pct = closes.pct_change().iloc[1:] * 100.0  # length n_digits
    b1,b2,b3,b4 = b
    digits_chrono = [quantize_5_fixed(float(x), b1,b2,b3,b4) for x in pct.tolist()]
    return "".join(digits_chrono[::-1])  # youngest-first


def build_markov(train_chrono: str, order: int):
    # train_chrono: oldest->youngest
    transitions = [(train_chrono[i:i+order], train_chrono[i+1:i+order+1])
                   for i in range(len(train_chrono) - order)]
    states = sorted(set(a for a,_ in transitions) | set(b for _,b in transitions))
    idx = {s:i for i,s in enumerate(states)}
    n = len(states)

    C = np.zeros((n,n), dtype=float)
    for a,b in transitions:
        C[idx[a], idx[b]] += 1.0

    denom = C.sum(axis=1, keepdims=True) + ALPHA * n
    P = (C + ALPHA) / denom
    return states, idx, P


def probs_next_digit(states: list[str], row: np.ndarray, K: int) -> np.ndarray:
    probs = np.zeros(K, dtype=float)
    for j,s in enumerate(states):
        probs[int(s[-1]) - 1] += float(row[j])
    return probs


def rolling_brier_for_digits(digits_youngest: str, order: int, window: int, K: int) -> float:
    chrono = digits_youngest[::-1]  # oldest->youngest
    brier_sum = 0.0
    n = 0

    for i in range(window, len(chrono)):
        train = chrono[i-window:i]
        current = train[-order:]
        factual = int(chrono[i]) - 1

        states, idx, P = build_markov(train, order)
        if current not in idx:
            continue

        probs = probs_next_digit(states, P[idx[current]], K)

        y = np.zeros(K, dtype=float)
        y[factual] = 1.0
        brier_sum += float(np.sum((probs - y)**2))
        n += 1

    return brier_sum / n if n else float("nan")


def portfolio_objective(thresholds: tuple[float,float,float,float],
                        tickers: list[str],
                        closes_map: dict[str, pd.Series]) -> float:
    # weighted mean Brier across tickers
    scores = []
    weights = []
    for t in tickers:
        try:
            digits = build_digits_from_thresholds(closes_map[t], thresholds, N_DIGITS)
            b = rolling_brier_for_digits(digits, ORDER, WINDOW, K)
            if not np.isfinite(b):
                continue
            w = float(WEIGHTS.get(t, 1.0)) if WEIGHTS else 1.0
            scores.append(b * w)
            weights.append(w)
        except Exception:
            continue

    if not weights:
        return float("inf")
    return sum(scores) / sum(weights)


# =========================
# MAIN: grid search
# =========================
def main():
    tickers = read_tickers_from_excel(XLSX, CFG_SHEET, TICKERS_NAME)
    print(f"Tickers: {len(tickers)}")

    # Download once (expensive step)
    closes_map = {}
    for t in tickers:
        try:
            closes_map[t] = download_closes(t, PERIOD)
        except Exception as e:
            print(f"Skip {t}: {e}")

    good = [t for t in tickers if t in closes_map]
    print(f"Have closes for: {len(good)}")

    best = (float("inf"), None)

    # symmetric grid
    for a in A_GRID:
        for c in C_GRID:
            if c >= a:
                continue
            b = (-a, -c, c, a)
            val = portfolio_objective(b, good, closes_map)
            if val < best[0]:
                best = (val, b)
                print(f"NEW BEST J={best[0]:.6f}  thresholds={best[1]}")

    print("\nDONE")
    print(f"Best portfolio Brier = {best[0]:.6f}")
    print(f"Best thresholds       = {best[1]}")
    print(f"GeneratedUTC          = {datetime.now(UTC).isoformat()}")


if __name__ == "__main__":
    main()