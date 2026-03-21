#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Robust full-sample Mutual Information analysis for K=4 states.

States:
1 : pct <= -1
2 : -1 < pct <= 0
3 :  0 < pct <  1
4 : pct >= 1

Features:
- Validates data length
- Per-ticker diagnostics
- Never crashes
- Reports skipped tickers
"""

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
import math
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400
ORDER = 3
K = 4

# ============================================================
# TICKERS
# ============================================================

def read_tickers():
    wb = load_workbook(XLSX, data_only=True)
    tickers = []

    dn = wb.defined_names.get(TICKERS_NAME)

    if dn:
        try:
            sheet_name, cell_range = next(dn.destinations)
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value:
                        tickers.append(str(cell.value).strip())
        except Exception as e:
            print("Named range read error:", e)

    if not tickers:
        print("Falling back to column A in M_Config")
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if not v:
                break
            tickers.append(str(v).strip())
            r += 1

    return list(dict.fromkeys(tickers))

# ============================================================
# DATA
# ============================================================

def download_closes(ticker):
    try:
        df = yf.download(
            ticker,
            period=PERIOD,
            interval="1d",
            auto_adjust=True,
            progress=False
        )
        if df.empty:
            raise ValueError("Empty download")

        s = df["Close"].dropna()
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:, 0]

        return s

    except Exception as e:
        print(f"[{ticker}] Download error:", e)
        return None

# ============================================================
# QUANTIZATION
# ============================================================

def quantize_4(pct):
    if pct <= -1: return 1
    if pct <=  0: return 2
    if pct <   1: return 3
    return 4

# ============================================================
# INFORMATION THEORY
# ============================================================

def entropy(counts):
    total = counts.sum()
    if total <= 0:
        return 0.0
    p = counts / total
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())

def mutual_information(digits, order):
    chrono = digits[::-1]
    n = len(chrono)

    if n <= order:
        return 0.0, 0.0, 0

    state_map = {}
    joint = []
    next_counts = np.zeros(K)

    def idx(s):
        if s not in state_map:
            state_map[s] = len(joint)
            joint.append(np.zeros(K))
        return state_map[s]

    transitions = 0

    for i in range(order, n):
        s = chrono[i-order:i]
        x = int(chrono[i]) - 1
        j = idx(s)
        joint[j][x] += 1
        next_counts[x] += 1
        transitions += 1

    if transitions == 0:
        return 0.0, 0.0, 0

    Hx = entropy(next_counts)

    HxS = 0.0
    total = next_counts.sum()
    for row in joint:
        rs = row.sum()
        if rs > 0:
            HxS += (rs / total) * entropy(row)

    MI = Hx - HxS
    return MI, Hx, transitions

# ============================================================
# MAIN
# ============================================================

def main():

    tickers = read_tickers()
    print(f"\nTickers listed: {len(tickers)}")

    MI_sum = 0.0
    Hx_sum = 0.0
    total_transitions = 0
    used = 0
    skipped = 0

    for ticker in tickers:

        closes = download_closes(ticker)

        if closes is None:
            skipped += 1
            continue

        if len(closes) < N_DIGITS + 1:
            print(f"[{ticker}] Skipped: insufficient data ({len(closes)} closes)")
            skipped += 1
            continue

        closes = closes.tail(N_DIGITS + 1)
        pct = closes.pct_change().iloc[1:] * 100.0

        if len(pct) != N_DIGITS:
            print(f"[{ticker}] Skipped: pct length mismatch")
            skipped += 1
            continue

        digits = "".join(str(quantize_4(float(x))) for x in pct)[::-1]

        MI, Hx, transitions = mutual_information(digits, ORDER)

        if transitions == 0:
            print(f"[{ticker}] Skipped: no transitions")
            skipped += 1
            continue

        print(f"[{ticker}] OK  transitions={transitions}  MI={MI:.4f}  Hx={Hx:.4f}")

        MI_sum += MI
        Hx_sum += Hx
        total_transitions += transitions
        used += 1

    print("\n================ SUMMARY ================")

    if used == 0:
        print("No valid tickers processed.")
        return

    MI_avg = MI_sum / used
    Hx_avg = Hx_sum / used
    NMI = MI_avg / Hx_avg if Hx_avg > 0 else 0

    print(f"Tickers processed: {used}")
    print(f"Tickers skipped:   {skipped}")
    print(f"Avg transitions:   {total_transitions / used:.1f}")
    print(f"\nMI_avg  = {MI_avg:.4f} bits")
    print(f"Hx_avg  = {Hx_avg:.4f} bits (max=2.000)")
    print(f"NMI_avg = {NMI:.4f}")

    print("\nGeneratedUTC:", datetime.now(UTC).isoformat())

if __name__ == "__main__":
    main()