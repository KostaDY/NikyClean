#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, UTC
from pathlib import Path
import numpy as np
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
STATE_CSV = Path("digits_400_5_state.csv")
RES_SHEET = "M_Results_5_Full"

ALPHABET = "12345"
K = len(ALPHABET)
PCT_FMT = numbers.FORMAT_PERCENTAGE

# ============================================================
# HELPERS
# ============================================================

def entropy(probs):
    return -sum(p * math.log(p, 2) for p in probs if p > 0)

def build_markov(chrono, order):
    transitions = [(chrono[i:i+order], chrono[i+1:i+order+1])
                   for i in range(len(chrono)-order)]
    states = sorted(set(a for a,_ in transitions) |
                    set(b for _,b in transitions))
    idx = {s:i for i,s in enumerate(states)}
    n = len(states)

    C = np.zeros((n,n))
    for a,b in transitions:
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        s = C[i].sum()
        if s > 0:
            P[i] = C[i] / s
        else:
            P[i,i] = 1.0

    return states, idx, P

def evaluate_fullsample(chrono, order):
    states, idx, P = build_markov(chrono, order)

    correct = 0
    total = 0
    brier_sum = 0.0
    entropy_sum = 0.0

    for i in range(order, len(chrono)):
        current = chrono[i-order:i]
        factual = int(chrono[i]) - 1

        if current not in idx:
            continue

        row = P[idx[current]]

        probs = [0.0]*K
        for j,s in enumerate(states):
            probs[int(s[-1])-1] += row[j]

        pred = np.argmax(probs)
        correct += int(pred == factual)
        total += 1

        y = [0.0]*K
        y[factual] = 1.0
        brier_sum += sum((probs[k]-y[k])**2 for k in range(K))
        entropy_sum += entropy(probs)

    acc = correct/total if total else 0
    brier = brier_sum/total if total else 0
    ent = entropy_sum/total if total else 0

    return total, correct, acc, brier, ent

# ============================================================
# MAIN
# ============================================================

def main():

    df = pd.read_csv(STATE_CSV, dtype={"Ticker":str,"Digits":str})
    df = df.set_index("Ticker")

    wb = load_workbook(XLSX)

    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)

    ws.append(["GeneratedUTC", datetime.now(UTC).isoformat(),
               "Alphabet", ALPHABET,
               "RandomBrier≈", 1-1/K])
    ws.append([])

    header = ["Ticker",
              "Obs_O2","Correct_O2","Acc_O2","Brier_O2","Entropy_O2",
              "Obs_O3","Correct_O3","Acc_O3","Brier_O3","Entropy_O3"]

    ws.append(header)
    for c in range(1,len(header)+1):
        ws.cell(ws.max_row,c).font = bold

    for t,row in df.iterrows():
        digits = row["Digits"]

        if len(digits)!=400 or any(ch not in ALPHABET for ch in digits):
            continue

        chrono = digits[::-1]

        o2 = evaluate_fullsample(chrono,2)
        o3 = evaluate_fullsample(chrono,3)

        ws.append([t,
                   *o2,
                   *o3])

    # format accuracy columns
    for r in range(4, ws.max_row+1):
        ws.cell(r,4).number_format = PCT_FMT
        ws.cell(r,9).number_format = PCT_FMT

    wb.save(XLSX)
    print("Phase3 5-state completed.")

if __name__=="__main__":
    main()