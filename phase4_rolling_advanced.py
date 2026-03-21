#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, UTC
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

from markov_core import read_tickers_from_excel, XLSX

STATE = Path("digits_400_state.csv")
RES_SHEET = "M_Results"

WINDOW = 200
ALPHA = 1e-6
Z = 1.96  # 95%


# ============================================================
# Utilities
# ============================================================

def entropy(probs):
    return -sum(p * np.log2(p) for p in probs if p > 0)


def wilson_interval(p, n):
    if n == 0:
        return 0.0, 0.0
    denom = 1 + Z**2 / n
    center = p + Z**2 / (2*n)
    margin = Z * np.sqrt(p*(1-p)/n + Z**2/(4*n**2))
    lower = (center - margin) / denom
    upper = (center + margin) / denom
    return lower, upper


def build_markov(train, order):
    transitions = [
        (train[i:i+order], train[i+1:i+order+1])
        for i in range(len(train) - order)
    ]

    states = sorted(set(a for a, _ in transitions) |
                    set(b for _, b in transitions))
    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))
    for a, b in transitions:
        C[idx[a], idx[b]] += 1

    P = (C + ALPHA) / (C.sum(axis=1, keepdims=True) + ALPHA * n)
    return states, idx, P


def evaluate_fixed_window(chrono, order):
    correct = 0
    total = 0
    brier_sum = 0.0
    entropy_sum = 0.0
    correctness_list = []

    for i in range(WINDOW, len(chrono)):
        train = chrono[i-WINDOW:i]
        current = train[-order:]
        factual = chrono[i]

        states, idx, P = build_markov(train, order)

        if current not in idx:
            continue

        row = P[idx[current]]
        probs = [0.0, 0.0, 0.0]

        for j, s in enumerate(states):
            probs[int(s[-1]) - 1] += row[j]

        pred = np.argmax(probs) + 1
        factual_int = int(factual)

        correct_flag = int(pred == factual_int)
        correctness_list.append(correct_flag)

        correct += correct_flag
        total += 1

        y = [0, 0, 0]
        y[factual_int - 1] = 1
        brier_sum += sum((probs[k] - y[k])**2 for k in range(3))
        entropy_sum += entropy(probs)

    acc = correct / total if total else 0
    brier = brier_sum / total if total else 0
    avg_entropy = entropy_sum / total if total else 0

    ci_low, ci_high = wilson_interval(acc, total)

    return total, acc, brier, avg_entropy, ci_low, ci_high, correctness_list


def mcnemar(o2_list, o3_list):
    b = c = 0
    for x, y in zip(o2_list, o3_list):
        if x == 1 and y == 0:
            b += 1
        elif x == 0 and y == 1:
            c += 1
    if b + c == 0:
        return 0.0
    return (b - c)**2 / (b + c)


# ============================================================
# MAIN
# ============================================================

def main():
    tickers = read_tickers_from_excel(XLSX)
    df_state = pd.read_csv(STATE, dtype={"Ticker": str, "Digits": str})
    df_state = df_state.set_index("Ticker")

    wb = load_workbook(XLSX)

    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)

    ws.append(["GeneratedUTC", datetime.now(UTC).isoformat(),
               "Mode", "RollingFixed200"])
    ws.append([])

    header = [
        "Ticker",
        "Obs",
        "Acc_O2", "CI_L_O2", "CI_U_O2", "Brier_O2", "Entropy_O2",
        "Acc_O3", "CI_L_O3", "CI_U_O3", "Brier_O3", "Entropy_O3",
        "McNemar_O2_vs_O3"
    ]

    ws.append(header)
    for c in range(1, len(header)+1):
        ws.cell(ws.max_row, c).font = bold

    for t in tickers:
        if t not in df_state.index:
            continue

        chrono = df_state.at[t, "Digits"][::-1]

        obs2, acc2, brier2, ent2, l2, u2, list2 = evaluate_fixed_window(chrono, 2)
        obs3, acc3, brier3, ent3, l3, u3, list3 = evaluate_fixed_window(chrono, 3)

        stat = mcnemar(list2, list3)

        ws.append([
            t,
            obs2,
            acc2, l2, u2, brier2, ent2,
            acc3, l3, u3, brier3, ent3,
            stat
        ])

    wb.save(XLSX)
    print("Advanced rolling backtest completed.")


if __name__ == "__main__":
    main()