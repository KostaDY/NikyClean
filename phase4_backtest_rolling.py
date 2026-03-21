#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from datetime import datetime, UTC
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

from markov_core import read_tickers_from_excel, XLSX

STATE = Path("digits_400_state.csv")
RES_SHEET = "M_Results"
PCT_FMT = numbers.FORMAT_PERCENTAGE

ALPHA = 1e-6          # tiny smoothing to avoid zero rows
MIN_TRAIN = 60        # minimum digits before starting evaluation (tune as you like)


def build_markov_from_prefix(chrono_prefix: str, order: int):
    transitions = []
    for i in range(len(chrono_prefix) - order):
        a = chrono_prefix[i:i+order]
        b = chrono_prefix[i+1:i+order+1]
        transitions.append((a, b))

    states = sorted(set(a for a, _ in transitions) | set(b for _, b in transitions))
    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))
    for a, b in transitions:
        C[idx[a], idx[b]] += 1

    # Laplace smoothing
    P = (C + ALPHA) / (C.sum(axis=1, keepdims=True) + ALPHA * n)
    return states, idx, P


def predict_next_digit(states, idx, P, current_state: str):
    """
    Return probs dict for next digit using the row of current_state.
    """
    if current_state not in idx:
        return {"1": 1/3, "2": 1/3, "3": 1/3}

    row = P[idx[current_state]]
    probs = {"1": 0.0, "2": 0.0, "3": 0.0}
    for j, s in enumerate(states):
        probs[s[-1]] += float(row[j])
    return probs


def eval_rolling(chrono: str, order: int):
    correct = 0
    total = 0
    brier_sum = 0.0

    # i is index of factual digit
    # train on chrono[:i] to predict chrono[i]
    start_i = max(order + 1, MIN_TRAIN)

    for i in range(start_i, len(chrono)):
        train = chrono[:i]  # up to i-1 inclusive
        current = train[-order:]  # state at time i-1 (order digits)
        factual = chrono[i]

        states, idx, P = build_markov_from_prefix(train, order)
        probs = predict_next_digit(states, idx, P, current)

        pred = max(probs, key=probs.get)
        if pred == factual:
            correct += 1

        y = {"1": 0, "2": 0, "3": 0}
        y[factual] = 1
        brier_sum += sum((probs[d] - y[d]) ** 2 for d in "123")

        total += 1

    acc = correct / total if total else 0.0
    brier = brier_sum / total if total else 0.0
    return total, correct, acc, brier


def main():
    tickers = read_tickers_from_excel(XLSX)

    if not STATE.exists():
        raise RuntimeError("digits_400_state.csv not found. Run phase1 first.")

    df_state = pd.read_csv(STATE, dtype={"Ticker": str, "Digits": str})
    df_state = df_state.set_index("Ticker")

    wb = load_workbook(XLSX)

    # Append to existing M_Results (don’t delete; comparison is useful)
    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.append([])
        ws.append([])
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)

    ws.append(["GeneratedUTC", datetime.now(UTC).isoformat(), "Mode", "Rolling(WalkForward)", "Alpha", ALPHA, "MinTrain", MIN_TRAIN])
    ws.append(["Ticker", "Order", "Observations", "Correct", "Accuracy", "BrierScore"])
    for c in range(1, 7):
        ws.cell(ws.max_row, c).font = bold

    for t in tickers:
        if t not in df_state.index:
            continue
        digits_youngest = str(df_state.at[t, "Digits"])
        if len(digits_youngest) == 0 or any(ch not in "123" for ch in digits_youngest):
            continue

        chrono = digits_youngest[::-1]  # oldest -> youngest

        for order in (2, 3):
            obs, cor, acc, brier = eval_rolling(chrono, order)
            ws.append([t, order, obs, cor, acc, brier])

    # Format Accuracy as %
    # Find rows with numeric Accuracy in column 5
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, (float, int)) and 0 <= float(cell.value) <= 1:
                cell.number_format = PCT_FMT

    wb.save(XLSX)
    print("OK: appended rolling backtest results to markov.xlsx / M_Results")


if __name__ == "__main__":
    main()