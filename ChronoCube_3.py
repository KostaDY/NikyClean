#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime as dt
import json
from pathlib import Path
import subprocess

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


# ==========================================================
# CONFIG
# ==========================================================

WORKBOOK = Path("Entry_CorrectHistory.xlsm")
SHEET_NAME = "Calc"
TABLE_NAME = "TabCalc"

CSV_FILE = Path("ChronoCube.csv")
META_FILE = Path("ChronoCube_meta.json")
EXCEL_FILE = Path("ChronoLines.xlsx")

MAX_SLOTS = 16


# ==========================================================
# RECALCULATE EXCEL
# ==========================================================

def recalc_excel():

    script = '''
    tell application "Microsoft Excel"
        calculate full
        save active workbook
    end tell
    '''

    subprocess.run(["osascript","-e",script])


# ==========================================================
# NORMALIZE
# ==========================================================

def normalize(v):

    if v is None:
        return np.nan

    if isinstance(v,str):

        s=v.strip()

        if s=="":
            return np.nan

        if s.startswith("#"):
            return np.nan

        if s.endswith("%"):
            try:
                return float(s[:-1])/100
            except:
                return s

        try:
            return float(s)
        except:
            return s

    return v


# ==========================================================
# READ TABCALC
# ==========================================================

def read_tabcalc():

    wb=load_workbook(WORKBOOK,data_only=True)
    ws=wb[SHEET_NAME]

    ref=ws.tables[TABLE_NAME].ref

    min_col,min_row,max_col,max_row=range_boundaries(ref)

    headers=[
        str(ws.cell(row=min_row,column=c).value)
        for c in range(min_col,max_col+1)
    ]

    rows=[]

    for r in range(min_row+1,max_row+1):

        row_values=[
            normalize(ws.cell(row=r,column=c).value)
            for c in range(min_col,max_col+1)
        ]

        rows.append(row_values)

    wb.close()

    df=pd.DataFrame(rows,columns=headers)

    if df.iloc[-1].astype(str).str.contains("Total",case=False,na=False).any():
        df=df.iloc[:-1]

    return df


# ==========================================================
# LOAD META
# ==========================================================

def load_meta():

    if META_FILE.exists():

        with open(META_FILE,"r") as f:
            meta=json.load(f)

        return meta["slot"],meta["dates"]

    return 0,[""]*MAX_SLOTS


# ==========================================================
# SAVE META
# ==========================================================

def save_meta(slot,dates):

    meta={"slot":slot,"dates":dates}

    with open(META_FILE,"w") as f:
        json.dump(meta,f,indent=2)


# ==========================================================
# LOAD CUBE (INDEX BASED)
# ==========================================================

def load_cube():

    cube = {}

    if not CSV_FILE.exists():
        return cube

    df = pd.read_csv(CSV_FILE)

    for _, row in df.iterrows():

        key = (row["Parameter"], int(row["Row"]))

        cube[key] = list(row.iloc[2:].values)

    return cube


# ==========================================================
# SAVE CUBE
# ==========================================================

def save_cube(cube):

    records=[]

    for (param,row),values in cube.items():

        rec=[param,row]
        rec.extend(values)

        records.append(rec)

    columns=["Parameter","Row"]+[f"Slot{i}" for i in range(MAX_SLOTS)]

    df=pd.DataFrame(records,columns=columns)

    df.to_csv(CSV_FILE,index=False)


# ==========================================================
# UPDATE CUBE
# ==========================================================

def update_cube(cube,df,slot):

    params=list(df.columns)
    rows=list(range(1,len(df)+1))

    snapshot=df.to_numpy(dtype=object).T

    for p,param in enumerate(params):
        for r,row in enumerate(rows):

            key=(param,row)

            if key not in cube:
                cube[key]=[np.nan]*MAX_SLOTS

            cube[key][slot]=snapshot[p,r]

    return cube


# ==========================================================
# EXPORT XLSX
# ==========================================================

def export_excel(cube,params,rows,dates,slot):

    # build chronological order
    order = list(range(slot,MAX_SLOTS)) + list(range(0,slot))

    writer=pd.ExcelWriter(EXCEL_FILE,engine="openpyxl")

    start_row=0

    for param in params:

        block=[]

        for row in rows:

            key=(param,row)

            if key in cube:
                values=[cube[key][i] for i in order]
            else:
                values=[np.nan]*MAX_SLOTS

            block.append(values)

        block_df=pd.DataFrame(block,columns=[dates[i] for i in order])
        block_df.insert(0,"Ticker",rows)

        title=pd.DataFrame([[f"Parameter: {param}"]])

        title.to_excel(writer,startrow=start_row,index=False,header=False)
        start_row+=1

        block_df.to_excel(writer,startrow=start_row,index=False)
        start_row+=len(block_df)+3

    writer.close()

    subprocess.run(["open",EXCEL_FILE])


# ==========================================================
# MAIN
# ==========================================================

def main():

    print("Recalculating Excel...")
    recalc_excel()

    print("Reading TabCalc...")
    df=read_tabcalc()

    params=list(df.columns)
    rows=list(range(1,len(df)+1))

    print("Parameters:",len(params))
    print("Rows:",len(rows))

    cube=load_cube()

    slot,dates=load_meta()

    print("Updating slot:",slot)

    cube=update_cube(cube,df,slot)

    dates[slot]=dt.date.today().isoformat()

    slot=(slot+1)%MAX_SLOTS

    save_cube(cube)
    save_meta(slot,dates)

    export_excel(cube,params,rows,dates,slot)

    print("\nCube updated")
    print("Keys:",len(cube))
    print("Excel file written:",EXCEL_FILE)


if __name__=="__main__":
    main()