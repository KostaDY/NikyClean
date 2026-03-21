#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime as dt
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


# ==========================================================
# CONFIG
# ==========================================================

WORKBOOK = Path("Entry_CorrectHistory.xlsm")
SHEET_NAME = "Calc"
TABLE_NAME = "TabCalc"

CUBE_FILE = Path("ChronoCube.npz")
EXCEL_FILE = Path("ChronoLines.xlsx")

MAX_SLOTS = 16
print("Workbook path:", WORKBOOK.resolve())

# ==========================================================
# NORMALIZE VALUES
# ==========================================================

def normalize(v):

    if v is None:
        return np.nan

    if isinstance(v, str):

        s = v.strip()

        if s == "":
            return np.nan

        if s.startswith("#"):
            return np.nan

        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100
            except:
                return s

        try:
            return float(s)
        except:
            return s

    return v


# ==========================================================
# READ TABCALC
# ==========================================================

def read_tabcalc():

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET_NAME]

    ref = ws.tables[TABLE_NAME].ref

    min_col, min_row, max_col, max_row = range_boundaries(ref)

    headers = [
        str(ws.cell(row=min_row, column=c).value)
        for c in range(min_col, max_col + 1)
    ]

    rows = []

    for r in range(min_row + 1, max_row + 1):

        row_values = [
            normalize(ws.cell(row=r, column=c).value)
            for c in range(min_col, max_col + 1)
        ]

        rows.append(row_values)

    wb.close()

    df = pd.DataFrame(rows, columns=headers)

    # remove totals row
    if df.iloc[-1].astype(str).str.contains("Total", case=False, na=False).any():
        df = df.iloc[:-1]

    return df


# ==========================================================
# LOAD / CREATE CUBE
# ==========================================================

def load_cube(n_param, n_row):

    if CUBE_FILE.exists():

        bundle = np.load(CUBE_FILE, allow_pickle=True)

        cube = bundle["cube"]

        if cube.ndim != 3:
            print("Old cube format detected → rebuilding")
            cube = np.full((n_param, n_row, MAX_SLOTS), np.nan, dtype=object)
            slot = 0
            dates = [""] * MAX_SLOTS
            return cube, slot, dates

        old_param, old_row, _ = cube.shape

        if old_param != n_param or old_row != n_row:
            print("TabCalc structure changed → rebuilding cube")
            cube = np.full((n_param, n_row, MAX_SLOTS), np.nan, dtype=object)
            slot = 0
            dates = [""] * MAX_SLOTS
            return cube, slot, dates

        slot = int(bundle["slot"])
        dates = list(bundle["dates"])

    else:

        cube = np.full((n_param, n_row, MAX_SLOTS), np.nan, dtype=object)
        slot = 0
        dates = [""] * MAX_SLOTS

    return cube, slot, dates


# ==========================================================
# SAVE CUBE
# ==========================================================

def save_cube(cube, slot, dates, params, rows):

    np.savez(
        CUBE_FILE,
        cube=cube,
        slot=slot,
        dates=np.array(dates, dtype=object),
        params=np.array(params, dtype=object),
        rows=np.array(rows, dtype=object)
    )


# ==========================================================
# EXPORT EXCEL
# ==========================================================

def export_excel(cube, params, rows, dates, slot, df):

    n_param, n_row, n_time = cube.shape

    # ----------------------------------------------
    # rebuild chronological order
    # ----------------------------------------------

    if slot == 0:
        chrono_cube = cube
        chrono_dates = dates
    else:
        chrono_cube = np.concatenate(
            (cube[:, :, slot:], cube[:, :, :slot]),
            axis=2
        )

        chrono_dates = dates[slot:] + dates[:slot]

    # ----------------------------------------------
    # flatten cube
    # ----------------------------------------------

    records = []

    for p in range(n_param):
        for r in range(n_row):

            rec = [params[p], rows[r]]

            rec.extend(chrono_cube[p, r, :])

            records.append(rec)

    columns = ["Parameter", "Row"] + chrono_dates

    df_cube = pd.DataFrame(records, columns=columns)

    # ----------------------------------------------
    # snapshot sheet
    # ----------------------------------------------

    df_snapshot = df.copy()
    df_snapshot.insert(0, "Row", rows)

    # ----------------------------------------------
    # write Excel
    # ----------------------------------------------

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:

        df_cube.to_excel(writer, sheet_name="CubeChronological", index=False)

        df_snapshot.to_excel(writer, sheet_name="TickerSnapshot", index=False)


# ==========================================================
# MAIN
# ==========================================================

def main():

    print("Reading TabCalc...")

    df = read_tabcalc()

    params = list(df.columns)

    rows = list(range(1, len(df) + 1))

    n_param = len(params)
    n_row = len(rows)

    print("Parameters:", n_param)
    print("Rows:", n_row)

    cube, slot, dates = load_cube(n_param, n_row)

    print("Updating slot:", slot)

    snapshot = df.to_numpy(dtype=object).T

    cube[:, :, slot] = snapshot

    dates[slot] = dt.date.today().isoformat()

    slot = (slot + 1) % MAX_SLOTS

    save_cube(cube, slot, dates, params, rows)

    export_excel(cube, params, rows, dates, slot, df)

    print()
    print("Cube updated")
    print("Cube shape:", cube.shape)
    print("Excel file written:", EXCEL_FILE)


if __name__ == "__main__":
    main()