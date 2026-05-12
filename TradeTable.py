#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# SETTINGS
# ============================================================

CSV_PATH = Path("/Users/kostayanev/NikyClean/TradeArchive.csv")
OUT_XLSX = Path("/Users/kostayanev/NikyClean/TradeTable.xlsx")

# CSV columns expected:
# Date, Ticker, Number, Price
#
# Rule:
# negative Price = purchase
# positive Price = sale
# sale reduces purchase lots with nearest lower absolute purchase price
# ============================================================


def load_trades(csv_path: Path) -> pd.DataFrame:

    # CSV WITHOUT HEADERS:
    # col0 = Date
    # col1 = Ticker
    # col2 = Number
    # col3 = Price

    df = pd.read_csv(
        csv_path,
        header=None,
        names=["Date", "Ticker", "Number", "Price"]
    )

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    df["Ticker"] = (
        df["Ticker"]
        .astype(str)
        .str.strip()
    )

    df["Number"] = pd.to_numeric(
        df["Number"],
        errors="coerce"
    )

    df["Price"] = pd.to_numeric(
        df["Price"],
        errors="coerce"
    )

    df = df.dropna(
        subset=["Date", "Ticker", "Number", "Price"]
    )

    df = df[df["Ticker"] != ""]
    df = df[df["Number"] != 0]
    df = df[df["Price"] != 0]

    df = df.sort_values(
        ["Ticker", "Date"]
    ).reset_index(drop=True)

    return df


def apply_sales_for_ticker(tdf: pd.DataFrame):

    lots = []

    for _, row in tdf.iterrows():

        date = row["Date"]

        ticker = row["Ticker"]

        number = abs(float(row["Number"]))

        price = float(row["Price"])

        # PURCHASE

        if price < 0:

            lots.append({

                "date": date,

                "number": number,

                "price": price

            })

        # SALE

        else:

            sale_qty = number

            sale_price_abs = abs(price)

            while sale_qty > 0:

                available = [

                    i for i, lot in enumerate(lots)

                    if lot["number"] > 0

                ]

                if not available:

                    print(

                        f"WARNING: Sale on {date.date()} cannot be fully matched "

                        f"for ticker {ticker}. Remaining sale qty: {sale_qty}"

                    )

                    break

                # Prefer purchase lots with price <= sale price.

                lower_or_equal = [

                    i for i in available

                    if abs(lots[i]["price"]) <= sale_price_abs

                ]

                if lower_or_equal:

                    # nearest lower purchase price = highest purchase price <= sale price

                    best_i = max(

                        lower_or_equal,

                        key=lambda i: abs(lots[i]["price"])

                    )

                else:

                    # if sale is below all purchases, use nearest higher purchase price

                    best_i = min(

                        available,

                        key=lambda i: abs(lots[i]["price"])

                    )

                lot = lots[best_i]

                matched_qty = min(sale_qty, lot["number"])

                lot["number"] -= matched_qty

                sale_qty -= matched_qty

                if lot["number"] <= 1e-12:

                    lots.pop(best_i)

    return lots


def build_trade_table(df: pd.DataFrame) -> dict:
    result = {}

    for ticker, tdf in df.groupby("Ticker", sort=True):
        result[ticker] = apply_sales_for_ticker(tdf)

    return result


def write_xlsx(result: dict, out_xlsx: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TradeTable"

    max_lots = max((len(lots) for lots in result.values()), default=0)

    headers = ["Ticker"]
    for i in range(1, max_lots + 1):
        headers += [f"Date_{i}", f"Number_{i}", f"Price_{i}"]

    ws.append(headers)

    for ticker, lots in result.items():
        row = [ticker]

        for lot in lots:
            row += [
                lot["date"].date(),
                lot["number"],
                lot["price"]
            ]

        ws.append(row)

    # formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column == 1:
                continue

            header = ws.cell(row=1, column=cell.column).value

            if header and header.startswith("Date_"):
                cell.number_format = "yyyy-mm-dd"

            elif header and header.startswith("Number_"):
                cell.number_format = "0"

            elif header and header.startswith("Price_"):
                cell.number_format = "0.00"

    wb.save(out_xlsx)


def main():
    df = load_trades(CSV_PATH)
    result = build_trade_table(df)
    write_xlsx(result, OUT_XLSX)

    print(f"✅ Created: {OUT_XLSX}")


if __name__ == "__main__":
    main()