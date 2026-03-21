#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
import math
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400
ORDER = 3
K = 4

# Grid ranges (coarse first)
B1_GRID = np.arange(-3.0, -0.25, 0.25)
B2_GRID = np.arange(-1.5,  0.75, 0.25)
B3_GRID = np.arange( 0.5,  3.25, 0.25)

MIN_HX = 1.2  # avoid collapsed digitizations

# ============================================================
# UTILITIES
# ============================================================

def read_tickers():
    wb = load_workbook(XLSX, data_only=True)
    tickers = []
    dn = wb.defined_names.get(TICKERS_NAME)

    if dn:
        try:
            sheet_name, cell_range = next(dn.destinations)
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value:
                        tickers.append(str(cell.value).strip())
        except:
            pass

    if not tickers:
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if not v:
                break
            tickers.append(str(v).strip())
            r += 1

    return list(dict.fromkeys(tickers))

def download_closes(t):
    try:
        df = yf.download(t, period=PERIOD,
                         interval="1d",
                         auto_adjust=True,
                         progress=False)
        if df.empty:
            return None
        s = df["Close"].dropna()
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:,0]
        return s
    except:
        return None

def quantize_4(pct, b1, b2, b3):
    if pct <= b1: return 1
    if pct <= b2: return 2
    if pct <= b3: return 3
    return 4

def entropy(counts):
    tot = counts.sum()
    if tot == 0: return 0
    p = counts / tot
    p = p[p>0]
    return float(-(p*np.log2(p)).sum())

def mutual_info(digits, order):
    chrono = digits[::-1]
    n = len(chrono)

    state_map = {}
    joint = []
    next_counts = np.zeros(K)

    def idx(s):
        if s not in state_map:
            state_map[s] = len(joint)
            joint.append(np.zeros(K))
        return state_map[s]

    for i in range(order, n):
        s = chrono[i-order:i]
        x = int(chrono[i]) - 1
        j = idx(s)
        joint[j][x] += 1
        next_counts[x] += 1

    Hx = entropy(next_counts)

    HxS = 0
    total = next_counts.sum()
    for row in joint:
        rs = row.sum()
        if rs > 0:
            HxS += (rs/total) * entropy(row)

    MI = Hx - HxS
    return MI, Hx

# ============================================================
# MAIN
# ============================================================

def main():

    tickers = read_tickers()
    print(f"Tickers listed: {len(tickers)}")

    closes_map = {}
    for t in tickers:
        c = download_closes(t)
        if c is not None and len(c) >= N_DIGITS+1:
            closes_map[t] = c

    print(f"Tickers usable: {len(closes_map)}")

    best = (-1e9, None)

    for b1 in B1_GRID:
        for b2 in B2_GRID:
            for b3 in B3_GRID:

                if not (b1 < b2 < b3):
                    continue

                MI_sum = 0
                Hx_sum = 0
                used = 0

                for t, closes in closes_map.items():

                    closes2 = closes.tail(N_DIGITS+1)
                    pct = closes2.pct_change().iloc[1:] * 100.0

                    if len(pct) != N_DIGITS:
                        continue

                    digits = "".join(
                        str(quantize_4(float(x), b1, b2, b3))
                        for x in pct
                    )[::-1]

                    MI, Hx = mutual_info(digits, ORDER)

                    if Hx < MIN_HX:
                        continue

                    MI_sum += MI
                    Hx_sum += Hx
                    used += 1

                if used == 0:
                    continue

                MI_avg = MI_sum / used
                Hx_avg = Hx_sum / used
                NMI = MI_avg / Hx_avg if Hx_avg>0 else 0

                if MI_sum > best[0]:
                    best = (MI_sum, (b1,b2,b3))
                    print(
                        f"NEW BEST  "
                        f"b1={b1:.2f}  b2={b2:.2f}  b3={b3:.2f}  "
                        f"MI_avg={MI_avg:.4f}  "
                        f"Hx_avg={Hx_avg:.4f}  "
                        f"NMI={NMI:.4f}"
                    )

    print("\n================ FINAL BEST ================")
    print(f"Best thresholds: {best[1]}")
    print(f"GeneratedUTC: {datetime.now(UTC).isoformat()}")

if __name__ == "__main__":
    main()