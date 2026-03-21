import pandas as pd
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule


# =========================
# SETTINGS
# =========================
INPUT_XLSX  = "trades.xlsx"
INPUT_SHEET = "Trades"
OUTPUT_XLSX = "FIFO_Dashboard.xlsx"


# =========================
# FIFO ENGINE (parametric)
# =========================
def run_fifo(purchases: pd.DataFrame, sales: pd.DataFrame):
    """
    purchases: columns [Date, Number, PPrice] where PPrice is positive cost (abs)
    sales: columns [Date(optional), Number, SPrice] where SPrice is positive sale price

    Returns:
      purchases_out: purchases with RemainingQty, Sales_OK, Sales_VIOL
      metrics: dict
    """
    p = purchases.copy().reset_index(drop=True)
    s = sales.copy().reset_index(drop=True)

    p["RemainingQty"] = p["Number"].astype(int)
    s["OriginalQty"] = s["Number"].astype(int)
    s["RemainingQty"] = s["Number"].astype(int)

    p["Sales_OK"] = ""
    p["Sales_VIOL"] = ""

    def add_trace(cur: str, txt: str) -> str:
        return txt if not cur else f"{cur} | {txt}"

    p_idx = 0
    s_idx = 0

    ok_value = 0.0
    viol_value = 0.0
    viol_qty = 0

    while p_idx < len(p) and s_idx < len(s):
        p_price = float(p.at[p_idx, "PPrice"])
        s_price = float(s.at[s_idx, "SPrice"])
        p_rem = int(p.at[p_idx, "RemainingQty"])
        s_rem = int(s.at[s_idx, "RemainingQty"])

        take = min(p_rem, s_rem)
        if take <= 0:
            # advance any exhausted pointers
            if p_rem <= 0:
                p_idx += 1
            if s_rem <= 0:
                s_idx += 1
            continue

        if s_price >= p_price:
            p.at[p_idx, "Sales_OK"] = add_trace(p.at[p_idx, "Sales_OK"], f"{s_price}:-{take}")
            ok_value += (s_price - p_price) * take
        else:
            p.at[p_idx, "Sales_VIOL"] = add_trace(p.at[p_idx, "Sales_VIOL"], f"{s_price}:-{take}")
            viol_qty += take
            viol_value += (p_price - s_price) * take

        p.at[p_idx, "RemainingQty"] -= take
        s.at[s_idx, "RemainingQty"] -= take

        if int(p.at[p_idx, "RemainingQty"]) == 0:
            p_idx += 1
        if int(s.at[s_idx, "RemainingQty"]) == 0:
            s_idx += 1

    metrics = {
        "TotalPurchaseQty": int(p["Number"].sum()),
        "TotalSaleQty": int(s["OriginalQty"].sum()),
        "RemainingPurchaseQty": int(p["RemainingQty"].sum()),
        "ViolationQty": int(viol_qty),
        "OK_Value": float(ok_value),
        "ViolationValue": float(viol_value),
    }
    metrics["NetValue"] = metrics["OK_Value"] - metrics["ViolationValue"]
    metrics["ViolationToOKRatio"] = (metrics["ViolationValue"] / metrics["OK_Value"]) if metrics["OK_Value"] else 0.0
    metrics["ViolationQtyRate"] = (metrics["ViolationQty"] / metrics["TotalSaleQty"]) if metrics["TotalSaleQty"] else 0.0

    return p, metrics


# =========================
# LOAD + PREP DATA
# =========================
df = pd.read_excel(INPUT_XLSX, sheet_name=INPUT_SHEET)
df = df.iloc[:, :3]
df.columns = ["Date", "Number", "Price"]
df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
df["Number"] = df["Number"].astype(int)
df["Price"] = df["Price"].astype(float)

p_raw = df[df["Price"] < 0].copy()
s_raw = df[df["Price"] > 0].copy()

# Build purchase price as positive abs
p_raw["PPrice"] = -p_raw["Price"]
s_raw["SPrice"] = s_raw["Price"]


# =========================
# MODEL A: Classical FIFO (date-based)
# =========================
p_A = p_raw.sort_values(["Date"]).loc[:, ["Date", "Number", "PPrice"]]
s_A = s_raw.sort_values(["Date"]).loc[:, ["Date", "Number", "SPrice"]]
out_A, metrics_A = run_fifo(p_A, s_A)
metrics_A["Model"] = "Classical_FIFO"

# =========================
# MODEL B: Global Price-FIFO (price-based, globally)
# Purchases: increasing PPrice
# Sales: increasing SPrice (global exhaustion of lower sales before higher)
# =========================
p_B = p_raw.sort_values(["PPrice", "Date"]).loc[:, ["Date", "Number", "PPrice"]]
s_B = s_raw.sort_values(["SPrice", "Date"]).loc[:, ["Date", "Number", "SPrice"]]
out_B, metrics_B = run_fifo(p_B, s_B)
metrics_B["Model"] = "Global_Price_FIFO"

metrics_df = pd.DataFrame([metrics_A, metrics_B])[[
    "Model",
    "TotalPurchaseQty",
    "TotalSaleQty",
    "RemainingPurchaseQty",
    "ViolationQty",
    "ViolationQtyRate",
    "OK_Value",
    "ViolationValue",
    "NetValue",
    "ViolationToOKRatio"
]]

# =========================
# WRITE BASE WORKBOOK WITH PANDAS
# =========================
with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
    out_A.to_excel(w, sheet_name="FIFO_Classical_Purchases", index=False)
    out_B.to_excel(w, sheet_name="FIFO_PriceFIFO_Purchases", index=False)
    metrics_df.to_excel(w, sheet_name="Metrics_Comparison", index=False)

# =========================
# ADD DASHBOARD + FORMATTING + CHARTS (openpyxl)
# =========================
wb = load_workbook(OUTPUT_XLSX)
ws_m = wb["Metrics_Comparison"]

# Create/replace Dashboard
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_d = wb.create_sheet("Dashboard", 0)

# Styling helpers
bold = Font(bold=True)
title_font = Font(bold=True, size=18)
kpi_font = Font(bold=True, size=14)
small = Font(size=10, color="666666")
hdr_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
hdr_font = Font(bold=True, color="FFFFFF")
kpi_fill = PatternFill("solid", fgColor="E8F0FE")  # light blue
delta_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")

# Title
ws_d["A1"] = "FIFO Monitoring Dashboard (Single Ticker)"
ws_d["A1"].font = title_font
ws_d["A2"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws_d["A2"].font = small
ws_d.merge_cells("A1:H1")

# KPI table header
ws_d["A4"] = "KPI"
ws_d["B4"] = "Classical FIFO"
ws_d["C4"] = "Global Price-FIFO"
ws_d["D4"] = "Δ (Price-FIFO − Classical)"
for cell in ("A4","B4","C4","D4"):
    ws_d[cell].fill = hdr_fill
    ws_d[cell].font = hdr_font
    ws_d[cell].alignment = center
    ws_d[cell].border = border

# Link KPIs from Metrics_Comparison (assumes 2 rows: row2 Classical, row3 PriceFIFO)
# Columns in Metrics_Comparison:
# A Model
# B TotalPurchaseQty
# C TotalSaleQty
# D RemainingPurchaseQty
# E ViolationQty
# F ViolationQtyRate
# G OK_Value
# H ViolationValue
# I NetValue
# J ViolationToOKRatio
kpis = [
    ("Total Sale Qty", "C"),
    ("Violation Qty", "E"),
    ("Violation Qty Rate", "F"),
    ("Violation Value", "H"),
    ("OK Value", "G"),
    ("Net Value (OK−Viol)", "I"),
    ("Violation/OK Ratio", "J"),
]

start_row = 5
for i, (label, col) in enumerate(kpis):
    r = start_row + i
    ws_d[f"A{r}"] = label
    ws_d[f"A{r}"].font = bold
    ws_d[f"A{r}"].alignment = left
    ws_d[f"A{r}"].border = border

    # Classical
    ws_d[f"B{r}"] = f"=Metrics_Comparison!{col}2"
    # Price-FIFO
    ws_d[f"C{r}"] = f"=Metrics_Comparison!{col}3"
    # Delta
    ws_d[f"D{r}"] = f"=C{r}-B{r}"

    for c in ("B","C","D"):
        ws_d[f"{c}{r}"].fill = kpi_fill if c in ("B","C") else delta_fill
        ws_d[f"{c}{r}"].alignment = right
        ws_d[f"{c}{r}"].border = border

# Formats
money_rows = [start_row + 3, start_row + 4, start_row + 5]  # ViolationValue, OK_Value, NetValue
pct_rows = [start_row + 2, start_row + 6]  # ViolationQtyRate, Violation/OK
int_rows = [start_row + 0, start_row + 1]  # TotalSaleQty, ViolationQty

for r in int_rows:
    for c in ("B","C","D"):
        ws_d[f"{c}{r}"].number_format = "#,##0"

for r in money_rows:
    for c in ("B","C","D"):
        ws_d[f"{c}{r}"].number_format = "#,##0.00"

for r in pct_rows:
    for c in ("B","C","D"):
        ws_d[f"{c}{r}"].number_format = "0.00%"

# Notes
ws_d["A13"] = "Notes:"
ws_d["A13"].font = bold
ws_d["A14"] = "• Classical FIFO: purchases matched by time (purchase date order)."
ws_d["A15"] = "• Global Price-FIFO: purchases matched by increasing purchase price; sales consumed by increasing sale price."
ws_d["A16"] = "• Violation = selling below matched purchase price; ViolationValue = Σ(purchase−sale)*qty."
for r in range(14, 17):
    ws_d[f"A{r}"].font = small
ws_d.merge_cells("A14:H14")
ws_d.merge_cells("A15:H15")
ws_d.merge_cells("A16:H16")

# Column widths
ws_d.column_dimensions["A"].width = 26
ws_d.column_dimensions["B"].width = 18
ws_d.column_dimensions["C"].width = 18
ws_d.column_dimensions["D"].width = 22

# -------------------------
# Chart data block (hidden-ish)
# -------------------------
ws_d["F4"] = "Model"
ws_d["G4"] = "ViolationValue"
ws_d["H4"] = "ViolationQty"
for cell in ("F4","G4","H4"):
    ws_d[cell].fill = hdr_fill
    ws_d[cell].font = hdr_font
    ws_d[cell].alignment = center
    ws_d[cell].border = border

ws_d["F5"] = "=Metrics_Comparison!A2"
ws_d["F6"] = "=Metrics_Comparison!A3"
ws_d["G5"] = "=Metrics_Comparison!H2"
ws_d["G6"] = "=Metrics_Comparison!H3"
ws_d["H5"] = "=Metrics_Comparison!E2"
ws_d["H6"] = "=Metrics_Comparison!E3"

for r in (5,6):
    for c in ("F","G","H"):
        ws_d[f"{c}{r}"].border = border
        ws_d[f"{c}{r}"].alignment = right

ws_d["G5"].number_format = "#,##0.00"
ws_d["G6"].number_format = "#,##0.00"
ws_d["H5"].number_format = "#,##0"
ws_d["H6"].number_format = "#,##0"

# -------------------------
# Charts
# -------------------------
# ViolationValue bar chart
bar1 = BarChart()
bar1.type = "col"
bar1.title = "Violation Value (Lower is Better)"
bar1.y_axis.title = "Value"
bar1.x_axis.title = "Model"
data = Reference(ws_d, min_col=7, min_row=4, max_row=6)  # G4:G6 including header
cats = Reference(ws_d, min_col=6, min_row=5, max_row=6)  # F5:F6
bar1.add_data(data, titles_from_data=True)
bar1.set_categories(cats)
bar1.dataLabels = DataLabelList()
bar1.dataLabels.showVal = True
ws_d.add_chart(bar1, "A18")

# ViolationQty bar chart
bar2 = BarChart()
bar2.type = "col"
bar2.title = "Violation Qty (Lower is Better)"
bar2.y_axis.title = "Shares"
bar2.x_axis.title = "Model"
data2 = Reference(ws_d, min_col=8, min_row=4, max_row=6)  # H4:H6
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats)
bar2.dataLabels = DataLabelList()
bar2.dataLabels.showVal = True
ws_d.add_chart(bar2, "E18")

# Conditional formatting on ratios (delta column)
ratio_row = start_row + 6
ws_d.conditional_formatting.add(
    f"B{ratio_row}:C{ratio_row}",
    ColorScaleRule(start_type="num", start_value=0, start_color="C6EFCE",
                   mid_type="num", mid_value=0.05, mid_color="FFEB9C",
                   end_type="num", end_value=0.25, end_color="FFC7CE")
)

# Freeze panes
ws_d.freeze_panes = "A5"
ws_m.freeze_panes = "A2"

# Save
wb.save(OUTPUT_XLSX)
print(f"✅ Dashboard workbook written: {Path(OUTPUT_XLSX).resolve()}")