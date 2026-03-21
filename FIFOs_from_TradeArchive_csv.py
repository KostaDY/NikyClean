import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ============================================================
# SETTINGS
# ============================================================

CSV_PATH  = Path("TradeArchive.csv")
XLSX_PATH = Path("TradeArchive.xlsx")

TRADES_WS = "Trades"
FIFOS_WS  = "FIFOs"

# ============================================================
# STEP 1: READ & SANITIZE CSV
# ============================================================

df = pd.read_csv(
    CSV_PATH,
    header=0,
    names=["Date", "Ticker", "Number", "Price"]
)

# Parse & coerce
df["Date"]   = pd.to_datetime(df["Date"], errors="coerce")
df["Ticker"] = df["Ticker"].astype(str)
df["Number"] = pd.to_numeric(df["Number"], errors="coerce")
df["Price"]  = pd.to_numeric(df["Price"], errors="coerce")

# Drop invalid rows explicitly
df = df.dropna(subset=["Date", "Ticker", "Number", "Price"])

# Enforce final types
df["Number"] = df["Number"].astype(int)
df["Price"]  = df["Price"].astype(float)

# ============================================================
# STEP 2: CREATE TradeArchive.xlsx WITH ws Trades
# ============================================================

with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name=TRADES_WS, index=False)

# ============================================================
# PRICE-CENTRIC FIFO (PER TICKER)
# ============================================================

def price_fifo(trades: pd.DataFrame):
    purchases = trades[trades["Price"] < 0].copy()
    sales     = trades[trades["Price"] > 0].copy()

    # Sort purchases by increasing absolute price (cheapest first)
    purchases["AbsPrice"] = -purchases["Price"]
    purchases.sort_values(["AbsPrice", "Date"], inplace=True)

    # Sort sales by increasing price
    sales.sort_values("Price", inplace=True)

    purchases["RemainingQty"] = purchases["Number"]
    purchases["SalesTrace"]   = ""

    def add_trace(cur, txt):
        return txt if not cur else f"{cur} | {txt}"

    p = s = 0
    purchases = purchases.reset_index(drop=True)
    sales = sales.reset_index(drop=True)

    while p < len(purchases) and s < len(sales):
        take = min(
            purchases.at[p, "RemainingQty"],
            sales.at[s, "Number"]
        )

        purchases.at[p, "RemainingQty"] -= take
        sales.at[s, "Number"] -= take

        purchases.at[p, "SalesTrace"] = add_trace(
            purchases.at[p, "SalesTrace"],
            f"{sales.at[s, 'Price']}:{take}"
        )

        if purchases.at[p, "RemainingQty"] == 0:
            p += 1
        if sales.at[s, "Number"] == 0:
            s += 1

    return purchases[purchases["RemainingQty"] > 0]

# ============================================================
# STEP 3: BUILD FIFOs OUTPUT (GROUPED BY TICKER)
# ============================================================

out_rows = []

for ticker, grp in df.groupby("Ticker"):
    # Ticker separator row
    out_rows.append([ticker, None, None, None, None, None])

    fifo = price_fifo(grp)

    for _, r in fifo.iterrows():
        out_rows.append([
            r["Date"],
            ticker,
            r["Number"],
            r["Price"],
            r["RemainingQty"],
            r["SalesTrace"]
        ])

    out_rows.append([None, None, None, None, None, None])

# ============================================================
# STEP 4: WRITE ws FIFOs
# ============================================================

wb = load_workbook(XLSX_PATH)

if FIFOS_WS in wb.sheetnames:
    del wb[FIFOS_WS]

ws = wb.create_sheet(FIFOS_WS)

ws.append([
    "Date",
    "Ticker",
    "PurchaseQty",
    "PurchasePrice",
    "RemainingQty",
    "SalesTrace"
])

for row in out_rows:
    ws.append(row)

wb.save(XLSX_PATH)
wb.close()

print("✅ TradeArchive.xlsx created from CSV and price-FIFO written to ws 'FIFOs'")