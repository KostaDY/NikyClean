
import pandas as pd
from openpyxl import load_workbook
import os
import traceback

# --------------------------------------------------------
# --- User configuration
# --------------------------------------------------------
INPUT_FILE = "MEDY.xlsx"        # workbook containing ticker list
TICKER_SHEET = "TickerList"     # sheet with tickers (first column)
OUTPUT_SHEET = "YahooData"      # sheet where results are saved
TEMP_FILE = "YahooTemp.xlsx"    # temporary fast-write file

# --------------------------------------------------------
# --- Your existing data fetch function
# --------------------------------------------------------
def fetch_ticker_data(ticker: str) -> dict:
    """
    Replace this stub with your actual Yahoo data fetch logic.
    It must return a dictionary of fields for each ticker.
    """
    # Example placeholder (replace with your real logic)
    return {
        "Ticker": ticker,
        "Price": 123.45,
        "Dividend Yield (%)": 0.0234,
        "Target Mean Price": 150.0,
    }

# --------------------------------------------------------
# --- Main logic
# --------------------------------------------------------
def main():
    try:
        print("📘 Reading tickers from MEDY.xlsx ...")
        df_tickers = pd.read_excel(INPUT_FILE, sheet_name=TICKER_SHEET)
        tickers = df_tickers.iloc[:, 0].dropna().astype(str).tolist()

        if not tickers:
            print("⚠️ No tickers found in sheet 'TickerList'.")
            return

        print(f"✅ {len(tickers)} tickers loaded: {', '.join(tickers[:10])}"
              + ("..." if len(tickers) > 10 else ""))

        # --- Fetch data for each ticker ---
        results = []
        for t in tickers:
            try:
                data = fetch_ticker_data(t)
                results.append(data)
            except Exception as e:
                print(f"⚠️ Skipped {t}: {e}")

        df_result = pd.DataFrame(results)
        if df_result.empty:
            print("⚠️ No data returned, aborting.")
            return

        # --- 1. FAST write to temporary file with xlsxwriter ---
        print("💾 Writing temporary sheet...")
        with pd.ExcelWriter(TEMP_FILE, engine="xlsxwriter") as writer:
            df_result.to_excel(writer, index=False, sheet_name=OUTPUT_SHEET)
            workbook = writer.book
            worksheet = writer.sheets[OUTPUT_SHEET]

            # Apply percent formatting if the column exists
            if "Dividend Yield (%)" in df_result.columns:
                percent_format = workbook.add_format({"num_format": "0.00%"})
                col_idx = df_result.columns.get_loc("Dividend Yield (%)")
                worksheet.set_column(col_idx, col_idx, 18, percent_format)

        # --- 2. Safely merge the sheet into MEDY.xlsx ---
        print("🔁 Merging into MEDY.xlsx ...")
        wb_target = load_workbook(INPUT_FILE)
        wb_temp = load_workbook(TEMP_FILE)
        ws_temp = wb_temp[OUTPUT_SHEET]

        # Remove old YahooData if exists
        if OUTPUT_SHEET in wb_target.sheetnames:
            del wb_target[OUTPUT_SHEET]

        # Create a new sheet and copy data
        ws_new = wb_target.create_sheet(OUTPUT_SHEET)
        for row in ws_temp.iter_rows(values_only=True):
            ws_new.append(row)

        # Save changes
        wb_target.save(INPUT_FILE)
        wb_target.close()
        wb_temp.close()

        # --- 3. Clean up temporary file ---
        if os.path.exists(TEMP_FILE):
            os.remove(TEMP_FILE)

        print(f"\n✅ Final data saved safely in {INPUT_FILE} → sheet '{OUTPUT_SHEET}'")

    except Exception as e:
        print(f"\n❌ An unexpected error occurred:\n{e}")
        print(traceback.format_exc())

# --------------------------------------------------------
# --- Run
# --------------------------------------------------------
if __name__ == "__main__":
    main()