#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import math
import subprocess
from pathlib import Path
from datetime import datetime
import shutil

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

# ============================================================
# FILES & SETTINGS
# ============================================================

CSV  = Path("markov.csv")      # canonical evolving state
XLSX = Path("markov.xlsx")     # presentation workbook (also contains M_Config)

CFG_SHEET = "M_Config"
RES_SHEET = "M_Results"

# used for daily factual change + safe weekend/holiday coverage
LOOKBACK_DAYS = 10

# used for initial seeding of DIGITS if not present
SEED_DAYS = 260  # enough to reliably get 200 trading days

PCT_FMT = numbers.FORMAT_PERCENTAGE

# ============================================================
# QUANTIZATION (3 STATES)
# ============================================================

def quantize(pct: float) -> str:
    if pct <= -1.0:
        return "1"
    if pct >= 1.0:
        return "3"
    return "2"

# ============================================================
# MARKOV (order k) – prediction only (no stationary here)
# ============================================================

def markov_predict(digits_youngest_first: str, order: int):
    # convert to chronological (oldest -> youngest) for transition counting
    chrono = list(digits_youngest_first)[::-1]

    # build complete state space a ∪ b
    states = set()
    for i in range(len(chrono) - order):
        states.add("".join(chrono[i:i+order]))
        states.add("".join(chrono[i+1:i+order+1]))
    states = sorted(states)
    idx = {s: i for i, s in enumerate(states)}
    n = len(states)

    C = np.zeros((n, n))
    for i in range(len(chrono) - order):
        a = "".join(chrono[i:i+order])
        b = "".join(chrono[i+1:i+order+1])
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        if C[i].sum():
            P[i] = C[i] / C[i].sum()
        else:
            P[i, i] = 1.0

    current = "".join(chrono[-order:])  # youngest group
    row = P[idx[current]]

    probs = {"1": 0.0, "2": 0.0, "3": 0.0}
    for j, s in enumerate(states):
        probs[s[-1]] += row[j]

    pred = max(probs, key=probs.get)
    return current, probs, pred

# ============================================================
# CSV STATE STORE
# ============================================================

def read_state():
    state = {}
    cur = None
    if not CSV.exists():
        return state

    for line in CSV.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith("[TICKER"):
            cur = line.split(",", 1)[1].rstrip("]")
            state[cur] = {}
        elif cur and "," in line:
            k, v = line.split(",", 1)
            state[cur][k] = v
    return state

def write_state(state: dict):
    out = []
    for t, s in state.items():
        out += [
            f"[TICKER,{t}]",
            f"DIGITS,{s['DIGITS']}",
            f"LAST_FACT_DATE,{s.get('LAST_FACT_DATE','')}",
            f"LAST_PRED_O2,{s['LAST_PRED_O2']}",
            f"LAST_PRED_O3,{s['LAST_PRED_O3']}",
            f"ACC_O2,{s['ACC_O2']}",
            f"ACC_O3,{s['ACC_O3']}",
            ""
        ]
    CSV.write_text("\n".join(out))

# ============================================================
# DATA HELPERS
# ============================================================

def download_recent_closes(ticker: str, days: int) -> pd.Series:
    df = yf.download(
        ticker,
        period=f"{days}d",
        interval="1d",
        auto_adjust=False,
        progress=False
    )
    closes = df["Close"].dropna()

    # If yfinance returns a DataFrame with multiindex columns, ensure we have a Series
    if isinstance(closes, pd.DataFrame):
        closes = closes.iloc[:, 0].dropna()

    return closes

def get_factual_digit_and_date(ticker: str):
    closes = download_recent_closes(ticker, LOOKBACK_DAYS)

    if len(closes) < 2:
        raise RuntimeError("Insufficient close data to compute daily change")

    last_close = closes.iloc[-1].item()
    prev_close = closes.iloc[-2].item()

    pct = (last_close / prev_close - 1.0) * 100.0
    digit = quantize(float(pct))

    # trading date for the last close
    last_dt = closes.index[-1]
    fact_date = pd.to_datetime(last_dt).date().isoformat()

    return digit, fact_date

def seed_digits_from_history(ticker: str) -> str:
    # Need 200 trading closes -> 199 pct-changes
    closes = download_recent_closes(ticker, SEED_DAYS)
    closes = closes.tail(200)

    if len(closes) < 200:
        raise RuntimeError("Not enough history to seed 200 trading days")

    pct = closes.pct_change() * 100.0
    digits = [quantize(float(x)) for x in pct.iloc[1:].tolist()]  # chronological oldest->youngest, len=199

    if len(digits) != 199:
        raise RuntimeError("Seeding failed: expected 199 digits")

    # store youngest-first string (head is newest)
    return "".join(digits[::-1])

def init_ticker_state(ticker: str) -> dict:
    digits = seed_digits_from_history(ticker)
    # initialize predictions from current seeded buffer
    _, _, pred2 = markov_predict(digits, 2)
    _, _, pred3 = markov_predict(digits, 3)

    return {
        "DIGITS": digits,
        "LAST_FACT_DATE": "",                 # empty until first daily update
        "LAST_PRED_O2": pred2,                # prediction for *next* day (as of now)
        "LAST_PRED_O3": pred3,
        "ACC_O2": "0:0:0:0:0:0:0:0:0",        # 3x3 flattened
        "ACC_O3": "0:0:0:0:0:0:0:0:0",
    }

def acc_update(acc_str: str, last_pred: str, factual: str) -> str:
    acc = list(map(int, acc_str.split(":")))
    acc[(int(last_pred) - 1) * 3 + (int(factual) - 1)] += 1
    return ":".join(map(str, acc))

def acc_to_pct_vector(acc_str: str):
    acc = list(map(int, acc_str.split(":")))
    tot = sum(acc)
    if tot == 0:
        return [0.0] * 9
    return [x / tot for x in acc]

# ============================================================
# MAIN
# ============================================================

def main():
    # --- backup markov.csv if exists ---
    if CSV.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(CSV, CSV.with_suffix(f".csv.bak_{ts}"))

    # --- open workbook (tickers source) ---
    wb = load_workbook(XLSX)

    if CFG_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{CFG_SHEET}' not found in {XLSX.name}")

    # --- named range TICKERS ---
    dn = wb.defined_names.get("TICKERS")
    if dn is None:
        raise RuntimeError("Named range 'TICKERS' not found")

    sheet_name, cell_range = next(dn.destinations)
    ws_rng = wb[sheet_name]

    tickers = [
        cell.value.strip()
        for row in ws_rng[cell_range]
        for cell in row
        if isinstance(cell.value, str) and cell.value.strip()
    ]
    if not tickers:
        raise RuntimeError("Named range 'TICKERS' is empty")

    # --- results sheet ---
    if RES_SHEET in wb.sheetnames:
        ws = wb[RES_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(RES_SHEET)

    bold = Font(bold=True)

    # --- load persisted state ---
    state = read_state()

    # --- ensure ticker states exist (seed from history if missing/bad) ---
    for t in tickers:
        need_init = False
        if t not in state:
            need_init = True
        else:
            digits = state[t].get("DIGITS", "")
            if not isinstance(digits, str) or len(digits) != 199 or any(ch not in "123" for ch in digits):
                need_init = True
            for k in ("LAST_PRED_O2", "LAST_PRED_O3", "ACC_O2", "ACC_O3"):
                if k not in state[t]:
                    need_init = True

        if need_init:
            state[t] = init_ticker_state(t)

    # ========================================================
    # ORDER 2 BLOCK
    # ========================================================
    headers = [
        "Ticker", "Youngest_O2", "P1", "P2", "P3",
        "P1→F1", "P1→F2", "P1→F3",
        "P2→F1", "P2→F2", "P2→F3",
        "P3→F1", "P3→F2", "P3→F3"
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = bold

    out_row = 2

    for t in tickers:
        s = state[t]

        # compute factual digit + date
        try:
            factual, fact_date = get_factual_digit_and_date(t)
        except Exception as e:
            # still output based on existing digits, but no update
            factual, fact_date = None, None

        # update once per trading day only
        if factual is not None:
            last_fact_date = s.get("LAST_FACT_DATE", "")
            if last_fact_date != fact_date:
                # 1) score yesterday's prediction (O2 + O3) against today's factual digit
                s["ACC_O2"] = acc_update(s["ACC_O2"], s["LAST_PRED_O2"], factual)
                s["ACC_O3"] = acc_update(s["ACC_O3"], s["LAST_PRED_O3"], factual)

                # 2) update cyclic buffer (newest at head)
                s["DIGITS"] = factual + s["DIGITS"][:-1]

                # 3) store date to prevent double update
                s["LAST_FACT_DATE"] = fact_date

        # compute today's predictions (for next day)
        young2, probs2, pred2 = markov_predict(s["DIGITS"], 2)
        young3, probs3, pred3 = markov_predict(s["DIGITS"], 3)
        s["LAST_PRED_O2"] = pred2
        s["LAST_PRED_O3"] = pred3

        acc2_pct = acc_to_pct_vector(s["ACC_O2"])

        ws.append([t, young2, probs2["1"], probs2["2"], probs2["3"]] + acc2_pct)
        out_row += 1

    # ========================================================
    # ORDER 3 BLOCK
    # ========================================================
    ws.append([])
    hdr_row = ws.max_row + 1
    ws.append([
        "Ticker", "Youngest_O3", "P1", "P2", "P3",
        "P1→F1", "P1→F2", "P1→F3",
        "P2→F1", "P2→F2", "P2→F3",
        "P3→F1", "P3→F2", "P3→F3"
    ])
    for c in range(1, len(headers) + 1):
        ws.cell(hdr_row, c).font = bold

    for t in tickers:
        s = state[t]
        young3, probs3, _ = markov_predict(s["DIGITS"], 3)
        acc3_pct = acc_to_pct_vector(s["ACC_O3"])
        ws.append([t, young3, probs3["1"], probs3["2"], probs3["3"]] + acc3_pct)

    # --------------------------------------------------------
    # Format percentages (P1..P3 and accuracy vector)
    # --------------------------------------------------------
    for row in ws.iter_rows(min_row=2, min_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = PCT_FMT

    # --- save state and workbook ---
    write_state(state)
    wb.save(XLSX)
    subprocess.run(["open", str(XLSX)], check=False)

    print("Daily Markov update completed successfully.")

# ============================================================

if __name__ == "__main__":
    main()