#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
TRADES – Transactions Manager GUI (SafeWrite, Reload-on-Transaction, Option C)

- Adds / deletes / undo deletes trades in TRADES.xlsx
- Strong validation with per-field GUI warnings
- Dates stored as true Excel dates
- Integers stored as integers, prices as floats
- Tickers loaded from Stock sheet (Ticker + Description)
- Safe delete (no spill through empty rows, Act ignored when matching)
- Logging to TRADES_log.csv and Log sheet
- DEBUG: prints key events and workbook/sheet identity
- No "Act" dropdown – buttons define the action (Add / Delete / Open / Undo / Exit)
- Option C: after writing Transactions (add/delete), we save and RELOAD workbook
"""

import os
import csv
from datetime import datetime, date
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl

# ======================================
# CONFIGURATION
# ======================================

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "TRADES.xlsx")
LOG_CSV_FILE = os.path.join(BASE_DIR, "TRADES_log.csv")

TRANS_HEADERS = ["Date", "Prefix", "Ticker", "Number", "Price", "Act"]
LOG_HEADERS = ["Timestamp", "Action", "Details"]

# Globals (workbook and sheets)
wb = None
ws_trans = None
ws_log = None

# Globals (tickers + undo)
TICKER_LIST = []
TICKER_DISPLAY = []
last_deleted_entry = None  # (row_values_list, original_row_index)


# ======================================
# DEBUG UTILITIES
# ======================================

def debug_print(message: str):
    """Simple debug print to terminal."""
    print("[DEBUG]", message)


def check_workbook_integrity(context: str = ""):
    """
    Ensure ws_trans and ws_log belong to the same workbook instance 'wb'.
    """
    global wb, ws_trans, ws_log
    if wb is None or ws_trans is None or ws_log is None:
        return

    wb_id = id(wb)
    trans_parent_id = id(ws_trans.parent)
    log_parent_id = id(ws_log.parent)

    msg = (f"Context={context} | wb_id={wb_id}, "
           f"ws_trans.parent={trans_parent_id}, ws_log.parent={log_parent_id}")
    debug_print(msg)

    # If these ever differ, something is seriously off.


def reload_workbook(context: str = ""):
    """
    Save current workbook, then reload it from disk and rebind ws_trans/ws_log.
    Used after Transactions changes (Option C).
    """
    global wb, ws_trans, ws_log

    if wb is None:
        return

    debug_print(f"Reload workbook ({context}) – saving first.")
    wb.save(EXCEL_FILE)

    wb_new = openpyxl.load_workbook(EXCEL_FILE)
    ws_trans_new = wb_new["Transactions"]
    ws_log_new = wb_new["Log"]

    wb = wb_new
    ws_trans = ws_trans_new
    ws_log = ws_log_new

    debug_print(
        f"Reloaded workbook ({context}). "
        f"wb_id={id(wb)}, ws_trans.parent={id(ws_trans.parent)}, ws_log.parent={id(ws_log.parent)}"
    )


# ======================================
# INITIALIZATION
# ======================================

def init_log_csv():
    """Ensure CSV log file exists with headers."""
    if not os.path.exists(LOG_CSV_FILE):
        with open(LOG_CSV_FILE, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(LOG_HEADERS)


def init_workbook():
    """Create/open TRADES.xlsx and ensure required sheets/headers."""
    global wb, ws_trans, ws_log

    debug_print(f"TRADES path: {EXCEL_FILE}")

    if not os.path.exists(EXCEL_FILE):
        debug_print("TRADES.xlsx not found, creating new workbook.")
        wb_new = openpyxl.Workbook()

        # Transactions sheet
        ws_t = wb_new.active
        ws_t.title = "Transactions"
        for col, h in enumerate(TRANS_HEADERS, start=1):
            ws_t.cell(row=1, column=col, value=h)

        # Log sheet
        ws_l = wb_new.create_sheet("Log")
        for col, h in enumerate(LOG_HEADERS, start=1):
            ws_l.cell(row=1, column=col, value=h)

        # Stock sheet (user-maintained)
        ws_s = wb_new.create_sheet("Stock")
        ws_s.cell(1, 1, "Ticker")
        ws_s.cell(1, 2, "Description")

        wb_new.save(EXCEL_FILE)
        debug_print("New TRADES.xlsx created.")

    wb_local = openpyxl.load_workbook(EXCEL_FILE)
    debug_print(f"Workbook loaded. id(wb_local) = {id(wb_local)}")

    # Ensure Transactions
    if "Transactions" not in wb_local.sheetnames:
        debug_print("Transactions sheet missing; creating.")
        ws_t = wb_local.create_sheet("Transactions")
        for col, h in enumerate(TRANS_HEADERS, start=1):
            ws_t.cell(row=1, column=col, value=h)

    # Ensure Log
    if "Log" not in wb_local.sheetnames:
        debug_print("Log sheet missing; creating.")
        ws_l = wb_local.create_sheet("Log")
        for col, h in enumerate(LOG_HEADERS, start=1):
            ws_l.cell(row=1, column=col, value=h)

    ws_trans_local = wb_local["Transactions"]
    ws_log_local = wb_local["Log"]

    # Ensure headers present in Log
    for col, h in enumerate(LOG_HEADERS, start=1):
        if ws_log_local.cell(row=1, column=col).value is None:
            ws_log_local.cell(row=1, column=col, value=h)

    wb_local.save(EXCEL_FILE)

    # Bind globals
    wb = wb_local
    ws_trans = ws_trans_local
    ws_log = ws_log_local

    debug_print(
        f"Bound globals: wb_id = {id(wb)}, "
        f"ws_trans.parent_id = {id(ws_trans.parent)}, "
        f"ws_log.parent_id = {id(ws_log.parent)}"
    )
    check_workbook_integrity("init_workbook")


def rebuild_log_from_csv():
    """Rebuilds Log sheet contents (rows 2+) from TRADES_log.csv."""
    global wb, ws_log
    check_workbook_integrity("rebuild_log_from_csv")

    # Clear rows except header
    while ws_log.max_row > 1:
        ws_log.delete_rows(2)

    if os.path.exists(LOG_CSV_FILE):
        with open(LOG_CSV_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)  # skip CSV header
            for row in reader:
                nr = ws_log.max_row + 1
                for c, v in enumerate(row, start=1):
                    ws_log.cell(row=nr, column=c, value=v)

    wb.save(EXCEL_FILE)
    debug_print("Log sheet rebuilt from CSV.")


# ======================================
# LOGGING (no save here, just write)
# ======================================

def append_log(action: str, details: str):
    """
    Append a log entry to CSV and to Log sheet.
    IMPORTANT: Does NOT call wb.save() to avoid complex interactions.
    """
    global wb, ws_log

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [ts, action, details]

    # CSV
    try:
        with open(LOG_CSV_FILE, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)
    except Exception as e:
        debug_print(f"Failed to write CSV log: {e}")

    # Excel (only if workbook is ready)
    if wb is not None and ws_log is not None:
        nr = ws_log.max_row + 1
        for c, v in enumerate(row, start=1):
            ws_log.cell(row=nr, column=c, value=v)
    # Workbook will be saved on next transaction or on exit.


# ======================================
# LOAD TICKERS FROM STOCK SHEET
# ======================================

def load_tickers_from_workbook():
    """
    Loads tickers and descriptions from current in-memory workbook wb.
    - If named range 'Tickers' exists: use it (A = ticker, B = description).
    - Else: read column A (ticker), B (description) starting from row 2.
    Returns (tickers_list, display_list).
    """
    global wb
    tickers = []
    display = []

    if wb is None:
        debug_print("load_tickers_from_workbook called but wb is None.")
        return [], []

    if "Stock" not in wb.sheetnames:
        debug_print("No Stock sheet; returning empty ticker list.")
        return [], []

    ws_s = wb["Stock"]

    # CASE 1 — Named range
    if "Tickers" in wb.defined_names:
        dn = wb.defined_names["Tickers"]
        for sheetname, cell_ref in dn.destinations:
            ws2 = wb[sheetname]
            rng = ws2[cell_ref]
            for row in rng:
                cell = row[0]
                if cell.value:
                    t = str(cell.value).strip().upper()
                    desc = ws2.cell(row=cell.row, column=cell.col_idx + 1).value
                    desc = str(desc).strip() if desc else ""
                    tickers.append(t)
                    display.append(f"{t} – {desc}" if desc else t)
    else:
        # CASE 2 — Column A+B
        for r in ws_s.iter_rows(min_row=2, max_col=2):
            c1 = r[0].value
            if c1:
                t = str(c1).strip().upper()
                desc = r[1].value
                desc = str(desc).strip() if desc else ""
                tickers.append(t)
                display.append(f"{t} – {desc}" if desc else t)

    # Deduplicate, preserve order
    final = list(dict.fromkeys(zip(tickers, display)))
    tick_list = [x[0] for x in final]
    disp_list = [x[1] for x in final]

    debug_print(f"Loaded {len(tick_list)} tickers from Stock.")
    return tick_list, disp_list


# ======================================
# VALIDATION HELPERS
# ======================================

def validate_date_field(value: str):
    v = value.strip()
    if not v:
        return False, "Date is required."
    try:
        d = datetime.strptime(v, "%m/%d/%y").date()
        return True, d
    except ValueError:
        return False, "Use mm/dd/yy."


def validate_prefix_field(value: str):
    v = value.strip()
    if not v:
        return False, "Prefix required."
    if not v.isdigit():
        return False, "Prefix must be integer."
    i = int(v)
    if not (1 <= i <= 9):
        return False, "Prefix 1–9."
    return True, i


def validate_ticker_field(value: str):
    v = value.strip().upper()
    if not v:
        return False, "Ticker required."
    if TICKER_LIST and v not in TICKER_LIST:
        return False, f"Unknown: {v}"
    return True, v


def validate_number_field(value: str):
    v = value.strip()
    if not v:
        return False, "Number required."
    if not v.isdigit():
        return False, "Must be integer."
    n = int(v)
    if n <= 0:
        return False, "> 0 required."
    return True, n


def validate_price_field(value: str):
    v = value.strip()
    if not v:
        return False, "Price required."
    try:
        p = float(v)
        return True, p
    except ValueError:
        return False, "Must be numeric."


# ======================================
# TRANSACTION OPERATIONS
# ======================================

def append_transaction(entry):
    """
    Append one transaction row (Date, Prefix, Ticker, Number, Price, Act)
    to Transactions sheet with correct typing.
    Then save & reload workbook (Option C).
    """
    global wb, ws_trans
    check_workbook_integrity("append_transaction")

    excel_date, prefix_i, ticker_s, number_i, price_f, act_s = entry

    nr = ws_trans.max_row + 1
    ws_trans.cell(row=nr, column=1, value=excel_date)          # date
    ws_trans.cell(row=nr, column=2, value=int(prefix_i))        # int
    ws_trans.cell(row=nr, column=3, value=str(ticker_s))        # str
    ws_trans.cell(row=nr, column=4, value=int(number_i))        # int
    ws_trans.cell(row=nr, column=5, value=float(price_f))       # float
    ws_trans.cell(row=nr, column=6, value=str(act_s))           # str

    debug_print(f"append_transaction: wrote row {nr}, entry={entry}")
    reload_workbook("after_append_transaction")  # save + reload

    append_log("add", f"Added: {entry}")


def delete_matching_transaction(entry):
    """
    Delete first row matching Date, Prefix, Ticker, Number, Price.
    Act column is ignored for matching.
    Safe shift without spilling through blank rows.
    Then save & reload workbook (Option C).
    """
    global wb, ws_trans, last_deleted_entry
    check_workbook_integrity("delete_matching_transaction")

    date_s, prefix_i, ticker_s, number_i, price_f, _ = entry
    max_row = ws_trans.max_row

    target = None
    for r in range(2, max_row + 1):
        cell_date = ws_trans.cell(r, 1).value
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()

        vals = [
            cell_date,
            ws_trans.cell(r, 2).value,
            ws_trans.cell(r, 3).value,
            ws_trans.cell(r, 4).value,
            ws_trans.cell(r, 5).value,
        ]

        price_cell = vals[4]
        try:
            price_cell_f = float(price_cell)
        except Exception:
            price_cell_f = price_cell

        if (vals[0] == date_s and
                vals[1] == prefix_i and
                vals[2] == ticker_s and
                vals[3] == number_i and
                price_cell_f == price_f):
            target = r
            break

    if not target:
        msg = f"No matching transaction for {entry}"
        debug_print(msg)
        append_log("delete_fail", msg)
        messagebox.showwarning("Not found", "No matching transaction found.")
        return False

    # Store for undo (all 6 columns)
    saved_row = [ws_trans.cell(target, c).value for c in range(1, 7)]
    last_deleted_entry = (saved_row, target)

    # Find last used row (ignore trailing blanks)
    last_used = max_row
    while last_used > 1:
        if any(ws_trans.cell(last_used, c).value not in ("", None)
               for c in range(1, 7)):
            break
        last_used -= 1

    # Shift rows between target and last_used - 1
    for r in range(target, last_used):
        for c in range(1, 7):
            ws_trans.cell(r, c).value = ws_trans.cell(r + 1, c).value

    # Clear last_used row
    for c in range(1, 7):
        ws_trans.cell(last_used, c).value = None

    debug_print(f"Deleted row {target}: {saved_row}")
    reload_workbook("after_delete_matching_transaction")  # save + reload

    append_log("delete", f"Deleted: {saved_row}")
    messagebox.showinfo("Deleted", "Transaction deleted.")
    return True


def undo_last_delete():
    """Restore the last deleted row if available."""
    global last_deleted_entry
    if not last_deleted_entry:
        messagebox.showinfo("Undo", "Nothing to undo.")
        return

    values, _ = last_deleted_entry
    append_transaction(values)
    append_log("undo_delete", f"Restored: {values}")
    last_deleted_entry = None
    messagebox.showinfo("Undo", "Last deleted transaction restored.")


def open_workbook():
    """Open TRADES.xlsx in the OS."""
    append_log("open", "Opened workbook from GUI.")
    # Save log updates before open
    wb.save(EXCEL_FILE)
    if os.name == "posix":
        subprocess.call(["open", EXCEL_FILE])
    else:
        os.startfile(EXCEL_FILE)


# ======================================
# GUI
# ======================================

def build_gui():
    """Build and return the Tk root window."""
    global TICKER_LIST, TICKER_DISPLAY

    root = tk.Tk()
    root.title("TRADES – Transactions Manager")

    frm = ttk.Frame(root, padding=10)
    frm.grid(row=0, column=0, sticky="nsew")

    # Variables (values)
    date_var = tk.StringVar(value=date.today().strftime("%m/%d/%y"))
    prefix_var = tk.StringVar(value="1")
    ticker_var = tk.StringVar()
    number_var = tk.StringVar()
    price_var = tk.StringVar()

    # Variables (errors)
    date_err = tk.StringVar(value="")
    prefix_err = tk.StringVar(value="")
    ticker_err = tk.StringVar(value="")
    number_err = tk.StringVar(value="")
    price_err = tk.StringVar(value="")

    # Load tickers initially
    TICKER_LIST[:], TICKER_DISPLAY[:] = load_tickers_from_workbook()
    if TICKER_LIST:
        ticker_var.set(TICKER_LIST[0])

    # Utility to create error label
    def make_error_label(parent, row, col):
        lbl = tk.Label(parent, text="", fg="red", anchor="w")
        lbl.grid(row=row, column=col, columnspan=2, sticky="w")
        return lbl

    # ---- Date ----
    ttk.Label(frm, text="Date (mm/dd/yy):").grid(row=0, column=0, sticky="w")
    date_entry = ttk.Entry(frm, textvariable=date_var, width=15)
    date_entry.grid(row=0, column=1, sticky="w")
    date_err_lbl = make_error_label(frm, 1, 1)

    # ---- Prefix ----
    ttk.Label(frm, text="Prefix (1–9):").grid(row=2, column=0, sticky="w")
    prefix_cb = ttk.Combobox(
        frm, textvariable=prefix_var,
        values=[str(i) for i in range(1, 10)],
        width=5, state="readonly"
    )
    prefix_cb.grid(row=2, column=1, sticky="w")
    prefix_err_lbl = make_error_label(frm, 3, 1)

    # ---- Ticker ----
    ttk.Label(frm, text="Ticker:").grid(row=4, column=0, sticky="w")
    ticker_cb = ttk.Combobox(
        frm, textvariable=ticker_var,
        values=TICKER_DISPLAY,
        width=30, state="readonly"
    )
    ticker_cb.grid(row=4, column=1, sticky="w")
    ticker_err_lbl = make_error_label(frm, 5, 1)

    def on_ticker_selected(event=None):
        sel = ticker_var.get()
        if "–" in sel:
            ticker_var.set(sel.split("–")[0].strip())
        ok, res = validate_ticker_field(ticker_var.get())
        ticker_err.set("" if ok else res)

    ticker_cb.bind("<<ComboboxSelected>>", on_ticker_selected)

    def reload_workbook(context: str = ""):
        global wb, ws_trans, ws_log

        try:
            debug_print(f"[RELOAD] Saving workbook ({context})")
            wb.save(EXCEL_FILE)
        except Exception as e:
            debug_print(f"[RELOAD ERROR] Failed to save workbook: {e}")
            return

        try:
            wb_new = openpyxl.load_workbook(EXCEL_FILE)
        except Exception as e:
            debug_print(f"[RELOAD ERROR] Failed to load workbook after save: {e}")
            return

        try:
            ws_trans_new = wb_new["Transactions"]
            ws_log_new = wb_new["Log"]
        except Exception as e:
            debug_print(f"[RELOAD ERROR] Missing sheet after reload: {e}")
            return

        wb = wb_new
        ws_trans = ws_trans_new
        ws_log = ws_log_new

        debug_print(
            f"[RELOAD] Done ({context}). "
            f"wb_id={id(wb)}, ws_trans.parent={id(ws_trans.parent)}, ws_log.parent={id(ws_log.parent)}"
        )

    # Bind field-level validation on focus-out
    def on_date_focus_out(event=None):
        ok, res = validate_date_field(date_var.get())
        date_err.set("" if ok else res)

    def on_prefix_focus_out(event=None):
        ok, res = validate_prefix_field(prefix_var.get())
        prefix_err.set("" if ok else res)

    def on_ticker_focus_out(event=None):
        ok, res = validate_ticker_field(ticker_var.get())
        ticker_err.set("" if ok else res)

    def on_number_focus_out(event=None):
        ok, res = validate_number_field(number_var.get())
        number_err.set("" if ok else res)

    def on_price_focus_out(event=None):
        ok, res = validate_price_field(price_var.get())
        price_err.set("" if ok else res)

    date_entry.bind("<FocusOut>", on_date_focus_out)
    prefix_cb.bind("<FocusOut>", on_prefix_focus_out)
    ticker_cb.bind("<FocusOut>", on_ticker_focus_out)
    number_entry.bind("<FocusOut>", on_number_focus_out)
    price_entry.bind("<FocusOut>", on_price_focus_out)

    # Helper to set error text
    def set_err(var, lbl, msg):
        var.set(msg)
        lbl.update_idletasks()

    # Clear errors
    def clear_errors():
        date_err.set("")
        prefix_err.set("")
        ticker_err.set("")
        number_err.set("")
        price_err.set("")

    # Clear fields after successful add/delete (keep ticker)
    def clear_fields():
        date_var.set(date.today().strftime("%m/%d/%y"))
        prefix_var.set("1")
        number_var.set("")
        price_var.set("")

    # --- ACTION HELPERS ---

    def run_add():
        clear_errors()

        ok_d, res_d = validate_date_field(date_var.get())
        if not ok_d:
            set_err(date_err, date_err_lbl, res_d)

        ok_pfx, res_pfx = validate_prefix_field(prefix_var.get())
        if not ok_pfx:
            set_err(prefix_err, prefix_err_lbl, res_pfx)

        ok_t, res_t = validate_ticker_field(ticker_var.get())
        if not ok_t:
            set_err(ticker_err, ticker_err_lbl, res_t)

        ok_n, res_n = validate_number_field(number_var.get())
        if not ok_n:
            set_err(number_err, number_err_lbl, res_n)

        ok_pr, res_pr = validate_price_field(price_var.get())
        if not ok_pr:
            set_err(price_err, price_err_lbl, res_pr)

        if not (ok_d and ok_pfx and ok_t and ok_n and ok_pr):
            messagebox.showerror("Validation", "Please correct the highlighted fields.")
            return

        excel_date, prefix_i, ticker_s, number_i, price_f = (
            res_d, res_pfx, res_t, res_n, res_pr
        )

        # Purchase / sale confirmation
        if price_f < 0:
            if not messagebox.askyesno(
                "Confirm Purchase",
                f"Price = {price_f}. Confirm PURCHASE?"
            ):
                return
        elif price_f > 0:
            if not messagebox.askyesno(
                "Confirm Sale",
                f"Price = {price_f}. Confirm SALE?"
            ):
                return

        entry = [excel_date, prefix_i, ticker_s, number_i, price_f, "add"]
        append_transaction(entry)
        messagebox.showinfo("Added", "Transaction added.")
        clear_fields()

    def run_delete():
        clear_errors()

        ok_d, res_d = validate_date_field(date_var.get())
        if not ok_d:
            set_err(date_err, date_err_lbl, res_d)

        ok_pfx, res_pfx = validate_prefix_field(prefix_var.get())
        if not ok_pfx:
            set_err(prefix_err, prefix_err_lbl, res_pfx)

        ok_t, res_t = validate_ticker_field(ticker_var.get())
        if not ok_t:
            set_err(ticker_err, ticker_err_lbl, res_t)

        ok_n, res_n = validate_number_field(number_var.get())
        if not ok_n:
            set_err(number_err, number_err_lbl, res_n)

        ok_pr, res_pr = validate_price_field(price_var.get())
        if not ok_pr:
            set_err(price_err, price_err_lbl, res_pr)

        if not (ok_d and ok_pfx and ok_t and ok_n and ok_pr):
            messagebox.showerror("Validation", "Please correct the highlighted fields.")
            return

        excel_date, prefix_i, ticker_s, number_i, price_f = (
            res_d, res_pfx, res_t, res_n, res_pr
        )

        entry = [excel_date, prefix_i, ticker_s, number_i, price_f, "add"]
        if delete_matching_transaction(entry):
            clear_fields()

    # BUTTON BAR
    btn = ttk.Frame(frm)
    btn.grid(row=10, column=0, columnspan=3, pady=10, sticky="w")

    ttk.Button(
        btn, text="Add",
        command=run_add
    ).grid(row=0, column=0, padx=5)

    ttk.Button(
        btn, text="Delete",
        command=run_delete
    ).grid(row=0, column=1, padx=5)

    ttk.Button(
        btn, text="Open WB",
        command=open_workbook
    ).grid(row=0, column=2, padx=5)

    ttk.Button(
        btn, text="Undo Delete",
        command=undo_last_delete
    ).grid(row=0, column=3, padx=5)

    ttk.Button(
        btn, text="Exit",
        command=lambda: [
            append_log("exit", "GUI session closed via Exit button."),
            wb.save(EXCEL_FILE),
            debug_print(f"Workbook saved on Exit: {EXCEL_FILE}"),
            root.destroy()
        ]
    ).grid(row=0, column=4, padx=5)

    return root


# ======================================
# MAIN
# ======================================

def main():
    init_log_csv()
    init_workbook()
    rebuild_log_from_csv()
    append_log("start", "GUI session started.")

    root = build_gui()
    root.mainloop()

    # Final save as safety net
    if wb is not None:
        wb.save(EXCEL_FILE)
        debug_print(f"Workbook saved on final exit: {EXCEL_FILE}")


if __name__ == "__main__":
    main()