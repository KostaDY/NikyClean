#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
import math
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400

K = 3
ORDER = 3   # try 2 or 3

A_GRID = np.arange(0.25, 3.01, 0.25)

MIN_HX = 0.5   # avoid trivial collapse

# ============================================================
# EXCEL TICKERS
# ============================================================

def read_tickers():
    wb = load_workbook(XLSX, data_only=True)
    tickers = []

    dn = wb.defined_names.get(TICKERS_NAME)

    if dn:
        sheet_name, cell_range = next(dn.destinations)
        ws = wb[sheet_name]
        for row in ws[cell_range]:
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip():
                    tickers.append(v.strip())

    if not tickers:
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if not v:
                break
            tickers.append(str(v).strip())
            r += 1

    return list(dict.fromkeys(tickers))

# ============================================================
# DATA
# ============================================================

def download_closes(ticker):
    df = yf.download(
        ticker,
        period=PERIOD,
        interval="1d",
        auto_adjust=True,
        progress=False
    )
    s = df["Close"].dropna()
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s

def quantize_3(pct, a):
    if pct <= -a:
        return 1
    if pct >= a:
        return 3
    return 2

def build_digits(closes, a):
    need = N_DIGITS + 1
    closes = closes.tail(need)
    pct = closes.pct_change().iloc[1:] * 100.0

    digits = [str(quantize_3(float(x), a)) for x in pct]
    return "".join(digits[::-1])

# ============================================================
# INFORMATION THEORY
# ============================================================

def entropy_from_counts(counts):
    tot = counts.sum()
    if tot == 0:
        return 0.0
    p = counts / tot
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())

def mutual_information(digits, order):
    chrono = digits[::-1]
    n = len(chrono)

    state_map = {}
    joint = []
    next_counts = np.zeros(K)

    def idx(s):
        if s not in state_map:
            state_map[s] = len(joint)
            joint.append(np.zeros(K))
        return state_map[s]

    obs = 0
    for i in range(order, n):
        s = chrono[i-order:i]
        x = int(chrono[i]) - 1
        j = idx(s)
        joint[j][x] += 1
        next_counts[x] += 1
        obs += 1

    Hx = entropy_from_counts(next_counts)

    HxS = 0.0
    total = obs
    for row in joint:
        rs = row.sum()
        if rs > 0:
            HxS += (rs / total) * entropy_from_counts(row)

    MI = Hx - HxS
    return MI, Hx, HxS

# ============================================================
# MAIN
# ============================================================

def main():

    tickers = read_tickers()
    print(f"Tickers listed: {len(tickers)}")

    closes_map = {}
    for t in tickers:
        try:
            closes_map[t] = download_closes(t)
        except:
            print("Skip", t)

    best = (-1e9, None)

    for a in A_GRID:

        MI_sum = 0
        Hx_sum = 0
        HxS_sum = 0
        used = 0

        for t in tickers:
            if t not in closes_map:
                continue

            try:
                digits = build_digits(closes_map[t], a)
                MI, Hx, HxS = mutual_information(digits, ORDER)

                if Hx < MIN_HX:
                    continue

                MI_sum += MI
                Hx_sum += Hx
                HxS_sum += HxS
                used += 1

            except:
                continue

        if used == 0:
            continue

        MI_avg = MI_sum / used
        Hx_avg = Hx_sum / used
        HxS_avg = HxS_sum / used
        NMI_avg = MI_avg / Hx_avg if Hx_avg > 0 else 0

        if MI_sum > best[0]:
            best = (MI_sum, a)
            print(
                f"NEW BEST a={a:.2f}  "
                f"MI_sum={MI_sum:.3f}  "
                f"Hx_avg={Hx_avg:.3f}  "
                f"Hx|S_avg={HxS_avg:.3f}  "
                f"MI_avg={MI_avg:.3f}  "
                f"NMI_avg={NMI_avg:.3f}"
            )

    print("\nDONE")
    print(f"Best threshold a = {best[1]}")
    print(f"GeneratedUTC = {datetime.now(UTC).isoformat()}")

if __name__ == "__main__":
    main()