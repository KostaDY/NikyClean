#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime as dt
import json
from pathlib import Path
import subprocess

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


# ==========================================================
# CONFIG
# ==========================================================

WORKBOOK = Path("Entry_RSI.xlsm")
SHEET_NAME = "Calc"
TABLE_NAME = "TabCalc"

CSV_FILE = Path("ChronoCube_RSI.csv")
META_FILE = Path("ChronoCube_meta.json")
EXCEL_FILE = Path("ChronoLines_RSI.xlsx")

MAX_SLOTS = 16


# ==========================================================
# RECALCULATE EXCEL
# ==========================================================

def recalc_excel():

    script = '''
    tell application "Microsoft Excel"
        calculate full
        save active workbook
    end tell
    '''

    subprocess.run(["osascript","-e",script])


# ==========================================================
# NORMALIZE
# ==========================================================

def normalize(v):

    if v is None:
        return np.nan

    if isinstance(v,str):

        s=v.strip()

        if s=="":
            return np.nan

        if s.startswith("#"):
            return np.nan

        if s.endswith("%"):
            try:
                return float(s[:-1])/100
            except:
                return s

        try:
            return float(s)
        except:
            return s

    return v


# ==========================================================
# READ TABCALC
# ==========================================================

def read_tabcalc():

    wb=load_workbook(WORKBOOK,data_only=True)
    ws=wb[SHEET_NAME]

    ref=ws.tables[TABLE_NAME].ref

    min_col,min_row,max_col,max_row=range_boundaries(ref)

    headers=[
        str(ws.cell(row=min_row,column=c).value)
        for c in range(min_col,max_col+1)
    ]

    rows=[]

    for r in range(min_row+1,max_row+1):

        row_values=[
            normalize(ws.cell(row=r,column=c).value)
            for c in range(min_col,max_col+1)
        ]

        rows.append(row_values)

    wb.close()

    df=pd.DataFrame(rows,columns=headers)

    if df.iloc[-1].astype(str).str.contains("Total",case=False,na=False).any():
        df=df.iloc[:-1]

    return df


# ==========================================================
# LOAD META
# ==========================================================

def load_meta():

    if META_FILE.exists():

        with open(META_FILE,"r") as f:
            meta=json.load(f)

        return meta["slot"],meta["dates"]

    return 0,[""]*MAX_SLOTS


# ==========================================================
# SAVE META
# ==========================================================

def save_meta(slot,dates):

    meta={"slot":slot,"dates":dates}

    with open(META_FILE,"w") as f:
        json.dump(meta,f,indent=2)


# ==========================================================
# LOAD / CREATE CUBE
# ==========================================================

def load_cube(n_param,n_row):

    if not CSV_FILE.exists():
        return np.full((n_param,n_row,MAX_SLOTS),np.nan,dtype=object)

    df=pd.read_csv(CSV_FILE)

    if df.shape[1]!=2+MAX_SLOTS:
        print("CSV structure mismatch → rebuilding cube")
        return np.full((n_param,n_row,MAX_SLOTS),np.nan,dtype=object)

    cube=df.iloc[:,2:].to_numpy(dtype=object)

    try:
        cube=cube.reshape(n_param,n_row,MAX_SLOTS)
    except:
        print("Cube size mismatch → rebuilding cube")
        cube=np.full((n_param,n_row,MAX_SLOTS),np.nan,dtype=object)

    return cube


# ==========================================================
# SAVE CSV
# ==========================================================

def save_cube_csv(cube,params,rows):

    n_param,n_row,n_slot=cube.shape

    records=[]

    for p in range(n_param):
        for r in range(n_row):

            rec=[params[p],rows[r]]
            rec.extend(cube[p,r,:])

            records.append(rec)

    columns=["Parameter","Row"]+[f"Slot{i}" for i in range(MAX_SLOTS)]

    df=pd.DataFrame(records,columns=columns)

    df.to_csv(CSV_FILE,index=False)


# ==========================================================
# EXPORT XLSX (BLOCK PER PARAMETER)
# ==========================================================

def export_excel(cube,params,rows,dates,slot):

    if slot==0:
        chrono_cube=cube
        chrono_dates=dates
    else:
        chrono_cube=np.concatenate((cube[:,:,slot:],cube[:,:,:slot]),axis=2)
        chrono_dates=dates[slot:]+dates[:slot]

    writer=pd.ExcelWriter(EXCEL_FILE,engine="openpyxl")

    start_row=0

    for p,param in enumerate(params):

        block=pd.DataFrame(
            chrono_cube[p,:,:].T,
            columns=[f"T{r}" for r in rows]
        ).T

        block.columns=chrono_dates
        block.insert(0,"Ticker",rows)

        title=pd.DataFrame([[f"Parameter: {param}"]])

        title.to_excel(writer,startrow=start_row,index=False,header=False)

        start_row+=1

        block.to_excel(writer,startrow=start_row,index=False)

        start_row+=len(block)+3

    writer.close()

    subprocess.run(["open",EXCEL_FILE])


# ==========================================================
# MAIN
# ==========================================================

def main():

    print("Recalculating Excel...")
    recalc_excel()

    print("Reading TabCalc...")

    df=read_tabcalc()

    params=list(df.columns)
    rows=list(range(1,len(df)+1))

    n_param=len(params)
    n_row=len(rows)

    print("Parameters:",n_param)
    print("Rows:",n_row)

    cube=load_cube(n_param,n_row)

    slot,dates=load_meta()

    print("Updating slot:",slot)

    snapshot=df.to_numpy(dtype=object).T

    cube[:,:,slot]=snapshot

    dates[slot]=dt.date.today().isoformat()

    slot=(slot+1)%MAX_SLOTS

    save_cube_csv(cube,params,rows)

    save_meta(slot,dates)

    export_excel(cube,params,rows,dates,slot)

    print("\nCube updated")
    print("Cube shape:",cube.shape)
    print("Excel file written:",EXCEL_FILE)


if __name__=="__main__":
    main()