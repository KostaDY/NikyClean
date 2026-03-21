#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, UTC
from pathlib import Path
import numpy as np
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
STATE_CSV = Path("digits_400_5_state.csv")
RES_SHEET = "M_Results_5_Roll200"

WINDOW = 200
ALPHA = 1e-6
Z = 1.96

ALPHABET = "12345"
K = len(ALPHABET)
PCT_FMT = numbers.FORMAT_PERCENTAGE

# ============================================================
# HELPERS
# ============================================================

def entropy(probs):
    return -sum(p*math.log(p,2) for p in probs if p>0)

def wilson_ci(p,n):
    if n==0: return 0,0
    denom = 1 + Z**2/n
    center = p + Z**2/(2*n)
    margin = Z*math.sqrt(p*(1-p)/n + Z**2/(4*n**2))
    return (center-margin)/denom, (center+margin)/denom

def build_markov(train,order):
    transitions=[(train[i:i+order],train[i+1:i+order+1])
                 for i in range(len(train)-order)]
    states=sorted(set(a for a,_ in transitions)|
                  set(b for _,b in transitions))
    idx={s:i for i,s in enumerate(states)}
    n=len(states)

    C=np.zeros((n,n))
    for a,b in transitions:
        C[idx[a],idx[b]]+=1

    denom=C.sum(axis=1,keepdims=True)+ALPHA*n
    P=(C+ALPHA)/denom
    return states,idx,P

def evaluate_fixed(chrono,order):

    correct=0
    total=0
    brier_sum=0
    entropy_sum=0
    correctness=[]

    for i in range(WINDOW,len(chrono)):
        train=chrono[i-WINDOW:i]
        current=train[-order:]
        factual=int(chrono[i])-1

        states,idx,P=build_markov(train,order)

        if current not in idx:
            continue

        row=P[idx[current]]

        probs=[0.0]*K
        for j,s in enumerate(states):
            probs[int(s[-1])-1]+=row[j]

        pred=np.argmax(probs)
        ok=int(pred==factual)

        correctness.append(ok)
        correct+=ok
        total+=1

        y=[0.0]*K
        y[factual]=1.0
        brier_sum+=sum((probs[k]-y[k])**2 for k in range(K))
        entropy_sum+=entropy(probs)

    acc=correct/total if total else 0
    brier=brier_sum/total if total else 0
    ent=entropy_sum/total if total else 0
    lo,hi=wilson_ci(acc,total)

    return total,acc,lo,hi,brier,ent,correctness

def mcnemar(a,b):
    n=min(len(a),len(b))
    B=C=0
    for i in range(n):
        if a[i]==1 and b[i]==0: B+=1
        elif a[i]==0 and b[i]==1: C+=1
    if B+C==0: return 0
    return (B-C)**2/(B+C)

# ============================================================
# MAIN
# ============================================================

def main():

    df=pd.read_csv(STATE_CSV,dtype={"Ticker":str,"Digits":str})
    df=df.set_index("Ticker")

    wb=load_workbook(XLSX)

    if RES_SHEET in wb.sheetnames:
        ws=wb[RES_SHEET]
        ws.delete_rows(1,ws.max_row)
    else:
        ws=wb.create_sheet(RES_SHEET)

    bold=Font(bold=True)

    ws.append(["GeneratedUTC",datetime.now(UTC).isoformat(),
               "Window",WINDOW,
               "RandomBrier≈",1-1/K])
    ws.append([])

    header=["Ticker","Obs",
            "Acc_O2","CI_L_O2","CI_U_O2","Brier_O2","Entropy_O2",
            "Acc_O3","CI_L_O3","CI_U_O3","Brier_O3","Entropy_O3",
            "McNemar_O2_vs_O3"]

    ws.append(header)
    for c in range(1,len(header)+1):
        ws.cell(ws.max_row,c).font=bold

    for t,row in df.iterrows():

        digits=row["Digits"]
        if len(digits)!=400 or any(ch not in ALPHABET for ch in digits):
            continue

        chrono=digits[::-1]

        o2=evaluate_fixed(chrono,2)
        o3=evaluate_fixed(chrono,3)

        stat=mcnemar(o2[6],o3[6])

        ws.append([t,o2[0],
                   o2[1],o2[2],o2[3],o2[4],o2[5],
                   o3[1],o3[2],o3[3],o3[4],o3[5],
                   stat])

    # format %
    for r in range(4,ws.max_row+1):
        for c in (3,4,5,8,9,10):
            ws.cell(r,c).number_format=PCT_FMT

    wb.save(XLSX)
    print("Phase4 5-state rolling completed.")

if __name__=="__main__":
    main()