import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, timedelta

# ============================================================
# SETTINGS
# ============================================================

INPUT_WB    = "Solid_REG_ALL.xlsm"
NAMED_RANGE = "Register"

OUTPUT_WB = "priceFIFO.xlsx"
OUTPUT_WS = "PriceFIFO"

# ============================================================
# EXCEL DATE NORMALIZATION
# ============================================================

def normalize_excel_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date (1900 system)
        return datetime(1899, 12, 30) + timedelta(days=float(v))
    return pd.to_datetime(v, errors="coerce")

# ============================================================
# LOAD NAMED RANGE (READ-ONLY)
# ============================================================

wb = load_workbook(INPUT_WB, data_only=True, read_only=True)

if NAMED_RANGE not in wb.defined_names:
    raise ValueError(f'Named range "{NAMED_RANGE}" not found')

defn = wb.defined_names[NAMED_RANGE]
sheet_name, ref = next(defn.destinations)
ws = wb[sheet_name]

rows = [[cell.value for cell in row] for row in ws[ref]]

# ============================================================
# PARSE TICKER GROUPS
# ============================================================

groups = {}
current_ticker = None

for r in rows:
    if len(r) < 3:
        continue

    date, number, price = r[0], r[1], r[2]

    # Ticker header row
    if isinstance(date, str) and isinstance(number, str) and isinstance(price, str):
        current_ticker = date.strip()
        groups[current_ticker] = []
        continue

    # Data row
    if current_ticker and date is not None:
        groups[current_ticker].append((
            normalize_excel_date(date),
            int(number),
            float(price)
        ))

# ============================================================
# PRICE-CENTRIC FIFO WITH VIOLATIONS
# ============================================================

def price_fifo_with_viol(trades):
    df = pd.DataFrame(trades, columns=["Date", "Number", "Price"])

    purchases = df[df["Price"] < 0].copy()
    sales     = df[df["Price"] > 0].copy()

    purchases["AbsPrice"] = -purchases["Price"]
    purchases.sort_values(["AbsPrice", "Date"], inplace=True)
    sales.sort_values("Price", inplace=True)

    purchases["Remaining"] = purchases["Number"]
    sales["Remaining"]     = sales["Number"]

    viol_qty   = 0
    viol_value = 0.0

    p_idx = s_idx = 0

    while p_idx < len(purchases) and s_idx < len(sales):
        p = purchases.iloc[p_idx]
        s = sales.iloc[s_idx]

        take = min(p["Remaining"], s["Remaining"])
        if take <= 0:
            break

        # Violation check
        if s["Price"] < p["AbsPrice"]:
            viol_qty += take
            viol_value += (p["AbsPrice"] - s["Price"]) * take

        purchases.iat[p_idx, purchases.columns.get_loc("Remaining")] -= take
        sales.iat[s_idx, sales.columns.get_loc("Remaining")] -= take

        if purchases.iloc[p_idx]["Remaining"] == 0:
            p_idx += 1
        if sales.iloc[s_idx]["Remaining"] == 0:
            s_idx += 1

    remaining = purchases[purchases["Remaining"] > 0][
        ["Date", "Remaining", "Price"]
    ].rename(columns={"Remaining": "Number"})

    summary = {
        "TotalPurchaseQty": int(purchases["Number"].sum()),
        "TotalSaleQty": int(sales["Number"].sum()),
        "RemainingQty": int(remaining["Number"].sum()),
        "ViolationQty": int(viol_qty),
        "ViolationValue": round(viol_value, 2),
    }

    return remaining, summary

# ============================================================
# BUILD SINGLE OUTPUT SHEET
# ============================================================

out_rows = []

for ticker, trades in groups.items():
    if not trades:
        continue

    # Ticker separator
    out_rows.append([ticker, None, None])

    fifo_df, summary = price_fifo_with_viol(trades)

    for _, r in fifo_df.iterrows():
        out_rows.append([
            r["Date"],
            int(r["Number"]),
            float(r["Price"])
        ])

    # Summary block
    out_rows.append([f"SUMMARY {ticker}", None, None])
    for k, v in summary.items():
        out_rows.append([k, v, None])

    out_rows.append([None, None, None])  # spacer

# ============================================================
# WRITE OUTPUT
# ============================================================

out_df = pd.DataFrame(out_rows, columns=["Date", "Number", "Price"])

with pd.ExcelWriter(OUTPUT_WB, engine="openpyxl", mode="w") as writer:
    out_df.to_excel(writer, sheet_name=OUTPUT_WS, index=False)

print(f"✅ Written: {Path(OUTPUT_WB).resolve()}")