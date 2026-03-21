import os
import csv
import subprocess
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ============================================================
# PATHS (THIS IS WHAT WAS MISSING)
# ============================================================
BASE_DIR = "/Users/kostayanev/NikyClean"
CSV_PATH = os.path.join(BASE_DIR, "TradeArchive.csv")
XLSX_PATH = os.path.join(BASE_DIR, "Parameters.xlsx")
SHEET_NAME = "TradeArchive"

# ============================================================
# CHECK CSV
# ============================================================
if not os.path.exists(CSV_PATH):
    raise FileNotFoundError(f"CSV not found: {CSV_PATH}")

# ============================================================
# OPEN OR CREATE WORKBOOK
# ============================================================
if os.path.exists(XLSX_PATH):
    wb = load_workbook(XLSX_PATH)
else:
    wb = Workbook()

# ============================================================
# RECREATE TradeArchive SHEET ONLY
# ============================================================
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME)

# ============================================================
# LOAD CSV WITH STRICT TYPES (NO append(row) FOR DATA)
# ============================================================
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader, start=1):

        # Header row (TEXT is correct)
        if r == 1:
            ws.append(row)
            continue

        # A — Date (Excel numeric date)
        try:
            dt = datetime.fromisoformat(row[0].strip())
            ws.cell(r, 1, dt)
            ws.cell(r, 1).number_format = "yyyy-mm-dd"
        except Exception:
            ws.cell(r, 1, None)

        # B — Ticker (text)
        ws.cell(r, 2, row[1].strip())

        # C — Integer (NUMERIC)
        try:
            ws.cell(r, 3, int(float(row[2])))
            ws.cell(r, 3).number_format = "0"
        except Exception:
            ws.cell(r, 3, None)

        # D — Decimal (NUMERIC, 2dp)
        try:
            ws.cell(r, 4, float(row[3]))
            ws.cell(r, 4).number_format = "0.00"
        except Exception:
            ws.cell(r, 4, None)

# ============================================================
# REMOVE DEFAULT EMPTY SHEET
# ============================================================
if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
    s = wb["Sheet"]
    if s.max_row == 1 and s.max_column == 1 and s["A1"].value is None:
        del wb["Sheet"]

# ============================================================
# SAVE & OPEN
# ============================================================
wb.save(XLSX_PATH)
subprocess.run(["open", XLSX_PATH])