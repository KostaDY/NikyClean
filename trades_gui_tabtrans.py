#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
TRADES – GUI for Excel table TabTrans in TRADES.xlsx

Workbook: TRADES.xlsx
Sheets:
  - Transactions: contains Excel Table TabTrans
      Columns (at least first 6):
         A: Date   (Excel date)
         B: Prefix (int)
         C: Ticker (text)
         D: Number (int)
         E: Price  (float)
         F: Act    (text: "add", etc.)
      Columns G onward are formula columns (Amount, Day, Day_Buys, ...)

  - Stock: tickers in A2:A... (non-empty cells) – used for dropdown
  - Log (optional): we append [Timestamp, Action, Details]

Operations:
  - Add: append a row to TabTrans (A–F) and copy formulas (G→last) from previous row
  - Delete: delete the first matching row (Date, Prefix, Ticker, Number, Price)
  - Open WB: open TRADES.xlsx
  - Exit: close GUI

No CSV, no reload gymnastics; every operation:
  load_workbook() → modify → save → done.
"""

import os
from datetime import datetime, date
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

# -------------------------------------------------
# CONFIG
# -------------------------------------------------

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "TRADES.xlsx")

TRANS_SHEET = "Transactions"
TABLE_NAME = "TabTrans"
STOCK_SHEET = "Stock"
LOG_SHEET = "Log"

# -------------------------------------------------
# BASIC HELPERS
# -------------------------------------------------

def debug(msg: str):
    print("[DEBUG]", msg)


def load_trades_workbook():
    """
    Load TRADES.xlsx and return (wb, ws_trans, ws_stock, ws_log_or_None).
    Raises if file/sheets are missing.
    """
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"TRADES.xlsx not found at {EXCEL_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    if TRANS_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{TRANS_SHEET}' not found in TRADES.xlsx")

    ws_trans = wb[TRANS_SHEET]
    ws_stock = wb[STOCK_SHEET] if STOCK_SHEET in wb.sheetnames else None
    ws_log = wb[LOG_SHEET] if LOG_SHEET in wb.sheetnames else None

    return wb, ws_trans, ws_stock, ws_log


def get_tabtrans(ws_trans):
    """
    Return the openpyxl Table object for TabTrans in ws_trans.
    """
    tables = ws_trans._tables if hasattr(ws_trans, "_tables") else ws_trans.tables
    # supports both dict-like and list-like storage
    if isinstance(tables, dict):
        for t in tables.values():
            if t.name == TABLE_NAME:
                return t
    else:
        for t in tables:
            if t.name == TABLE_NAME:
                return t
    raise KeyError(f"Table '{TABLE_NAME}' not found in sheet '{TRANS_SHEET}'")


def append_log_entry(action: str, details: str):
    """
    Append a log line to Log sheet if it exists: [Timestamp, Action, Details].
    Uses a separate load/save to avoid interference with transaction operations.
    """
    try:
        wb, _, _, ws_log = load_trades_workbook()
    except Exception as e:
        debug(f"append_log_entry: cannot load workbook: {e}")
        return

    if ws_log is None:
        return

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = ws_log.max_row + 1
    ws_log.cell(row=row, column=1, value=ts)
    ws_log.cell(row=row, column=2, value=action)
    ws_log.cell(row=row, column=3, value=details)

    try:
        wb.save(EXCEL_FILE)
    except Exception as e:
        debug(f"append_log_entry: save failed: {e}")


# -------------------------------------------------
# STOCK / TICKERS
# -------------------------------------------------

def load_tickers_from_stock():
    """
    Return list of tickers from Stock!A2:A... (non-empty, uppercase, de-duplicated).
    """
    try:
        wb, _, ws_stock, _ = load_trades_workbook()
    except Exception as e:
        debug(f"load_tickers_from_stock: cannot load workbook: {e}")
        return []

    if ws_stock is None:
        debug("load_tickers_from_stock: sheet 'Stock' missing.")
        return []

    tickers = []
    for row in range(2, ws_stock.max_row + 1):
        val = ws_stock.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            t = str(val).strip().upper()
            tickers.append(t)

    # Remove duplicates, preserve order
    seen = set()
    out = []
    for t in tickers:
        if t not in seen:
            seen.add(t)
            out.append(t)

    debug(f"load_tickers_from_stock: {len(out)} tickers loaded.")
    return out


# -------------------------------------------------
# VALIDATION
# -------------------------------------------------

def validate_date(value: str):
    v = value.strip()
    if not v:
        return False, "Date required."
    try:
        d = datetime.strptime(v, "%m/%d/%y").date()
        return True, d
    except ValueError:
        return False, "Use mm/dd/yy."


def validate_prefix(value: str):
    v = value.strip()
    if not v:
        return False, "Prefix required."
    if not v.isdigit():
        return False, "Prefix must be integer."
    i = int(v)
    if not (1 <= i <= 9):
        return False, "Prefix 1–9 only."
    return True, i


def validate_ticker(value: str, allowed):
    v = value.strip().upper()
    if not v:
        return False, "Ticker required."
    if allowed and v not in allowed:
        return False, f"Ticker '{v}' not in Stock."
    return True, v


def validate_number(value: str):
    v = value.strip()
    if not v:
        return False, "Number required."
    if not v.isdigit():
        return False, "Number must be integer."
    n = int(v)
    if n <= 0:
        return False, "> 0 required."
    return True, n


def validate_price(value: str):
    v = value.strip()
    if not v:
        return False, "Price required."
    try:
        p = float(v)
        return True, p
    except ValueError:
        return False, "Price must be numeric."


# -------------------------------------------------
# TABLE OPERATIONS – ADD / DELETE
# -------------------------------------------------

def add_transaction(excel_date, prefix_i, ticker_s, number_i, price_f):
    """
    Add a row to TabTrans with given values, adjust table ref, save.
    Act is set to "add".

    - Writes only into A–F (data columns).
    - Copies formulas in G..last column from previous data row if any.
    """
    wb, ws_trans, _, _ = load_trades_workbook()
    tbl = get_tabtrans(ws_trans)

    # Table ref boundaries
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    header_row = min_row      # should be 1
    last_data_row = max_row   # last data row within table

    # New row is next row after last_data_row
    new_row = last_data_row + 1

    # --- 1) Write values into A:F (first 6 columns)
    # Columns: A=Date, B=Prefix, C=Ticker, D=Number, E=Price, F=Act
    row_values = [excel_date, prefix_i, ticker_s, number_i, price_f, "add"]
    for offset, val in enumerate(row_values):
        col_idx = min_col + offset  # min_col should be 1
        ws_trans.cell(row=new_row, column=col_idx, value=val)

    # --- 2) Copy formulas from previous last row (if there is at least one data row)
    if last_data_row > header_row:
        for c in range(min_col + 6, max_col + 1):  # from column G (A=1, so G=7) to last column
            src = ws_trans.cell(row=last_data_row, column=c)
            dst = ws_trans.cell(row=new_row, column=c)
            dst.value = src.value

    # --- 3) Extend table ref by one row
    new_max_row = new_row
    tbl.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_max_row}"

    debug(f"add_transaction: wrote row {new_row}, values={row_values}, table_ref={tbl.ref}")
    wb.save(EXCEL_FILE)
    append_log_entry("add", f"Row {new_row}: {row_values}")


def delete_transaction(excel_date, prefix_i, ticker_s, number_i, price_f):
    """
    Delete first matching row in TabTrans by (Date, Prefix, Ticker, Number, Price).
    Delete the entire table row (A..last column), shrink table, save.
    """
    wb, ws_trans, _, _ = load_trades_workbook()
    tbl = get_tabtrans(ws_trans)

    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    header_row = min_row
    first_data_row = header_row + 1

    # Locate target row
    target_row = None
    for r in range(first_data_row, max_row + 1):
        cell_date = ws_trans.cell(r, min_col).value
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()
        elif isinstance(cell_date, date):
            pass  # already date
        # else: other types won't match excel_date

        cell_prefix = ws_trans.cell(r, min_col + 1).value
        cell_ticker = ws_trans.cell(r, min_col + 2).value
        cell_number = ws_trans.cell(r, min_col + 3).value
        cell_price = ws_trans.cell(r, min_col + 4).value

        # Normalize ticker and price
        cell_ticker_norm = (cell_ticker or "").strip().upper()
        try:
            cell_price_f = float(cell_price)
        except Exception:
            cell_price_f = cell_price

        if (cell_date == excel_date and
            cell_prefix == prefix_i and
            cell_ticker_norm == ticker_s and
            cell_number == number_i and
            cell_price_f == price_f):
            target_row = r
            break

    if target_row is None:
        msg = f"No match for ({excel_date}, {prefix_i}, {ticker_s}, {number_i}, {price_f})"
        debug(f"delete_transaction: {msg}")
        append_log_entry("delete_fail", msg)
        messagebox.showwarning("Delete", "No matching transaction found.")
        return False

    debug(f"delete_transaction: deleting row {target_row}")

    # Delete entire row in sheet (A..last column will shift up)
    ws_trans.delete_rows(target_row, 1)

    # Shrink table ref by one row
    new_max_row = max_row - 1
    if new_max_row < header_row:
        new_max_row = header_row  # at least header
    tbl.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_max_row}"

    wb.save(EXCEL_FILE)
    append_log_entry(
        "delete",
        f"Deleted row {target_row} for ({excel_date}, {prefix_i}, {ticker_s}, {number_i}, {price_f})"
    )
    messagebox.showinfo("Delete", "Transaction deleted.")
    return True


# -------------------------------------------------
# GUI
# -------------------------------------------------

def build_gui():
    root = tk.Tk()
    root.title("TRADES – TabTrans Manager")

    frm = ttk.Frame(root, padding=10)
    frm.grid(row=0, column=0, sticky="nsew")

    # Load tickers once at start (button can reload them)
    ticker_list = load_tickers_from_stock()

    # Variables
    date_var = tk.StringVar(value=date.today().strftime("%m/%d/%y"))
    prefix_var = tk.StringVar(value="1")
    ticker_var = tk.StringVar()
    number_var = tk.StringVar()
    price_var = tk.StringVar()

    date_err = tk.StringVar(value="")
    prefix_err = tk.StringVar(value="")
    ticker_err = tk.StringVar(value="")
    number_err = tk.StringVar(value="")
    price_err = tk.StringVar(value="")

    # --- Layout helpers
    def err_label(row):
        lbl = tk.Label(frm, text="", fg="red", anchor="w")
        lbl.grid(row=row, column=1, columnspan=2, sticky="w")
        return lbl

    # Date
    ttk.Label(frm, text="Date (mm/dd/yy):").grid(row=0, column=0, sticky="w")
    date_entry = ttk.Entry(frm, textvariable=date_var, width=15)
    date_entry.grid(row=0, column=1, sticky="w")
    date_err_lbl = err_label(1)

    # Prefix
    ttk.Label(frm, text="Prefix (1–9):").grid(row=2, column=0, sticky="w")
    prefix_cb = ttk.Combobox(
        frm, textvariable=prefix_var,
        values=[str(i) for i in range(1, 10)],
        width=5, state="readonly"
    )
    prefix_cb.grid(row=2, column=1, sticky="w")
    prefix_err_lbl = err_label(3)

    # Ticker
    ttk.Label(frm, text="Ticker:").grid(row=4, column=0, sticky="w")
    ticker_cb = ttk.Combobox(
        frm, textvariable=ticker_var,
        values=ticker_list,
        width=15, state="readonly"
    )
    ticker_cb.grid(row=4, column=1, sticky="w")
    ticker_err_lbl = err_label(5)

    # Number
    ttk.Label(frm, text="Number:").grid(row=6, column=0, sticky="w")
    number_entry = ttk.Entry(frm, textvariable=number_var, width=15)
    number_entry.grid(row=6, column=1, sticky="w")
    number_err_lbl = err_label(7)

    # Price
    ttk.Label(frm, text="Price:").grid(row=8, column=0, sticky="w")
    price_entry = ttk.Entry(frm, textvariable=price_var, width=15)
    price_entry.grid(row=8, column=1, sticky="w")
    price_err_lbl = err_label(9)

    # Validation on focus out
    def on_date_out(_=None):
        ok, res = validate_date(date_var.get())
        date_err.set("" if ok else res)

    def on_prefix_out(_=None):
        ok, res = validate_prefix(prefix_var.get())
        prefix_err.set("" if ok else res)

    def on_ticker_out(_=None):
        ok, res = validate_ticker(ticker_var.get(), ticker_list)
        ticker_err.set("" if ok else res)

    def on_number_out(_=None):
        ok, res = validate_number(number_var.get())
        number_err.set("" if ok else res)

    def on_price_out(_=None):
        ok, res = validate_price(price_var.get())
        price_err.set("" if ok else res)

    date_entry.bind("<FocusOut>", on_date_out)
    prefix_cb.bind("<FocusOut>", on_prefix_out)
    ticker_cb.bind("<FocusOut>", on_ticker_out)
    number_entry.bind("<FocusOut>", on_number_out)
    price_entry.bind("<FocusOut>", on_price_out)

    # Bind error vars to labels
    date_err.trace_add("write", lambda *_: date_err_lbl.config(text=date_err.get()))
    prefix_err.trace_add("write", lambda *_: prefix_err_lbl.config(text=prefix_err.get()))
    ticker_err.trace_add("write", lambda *_: ticker_err_lbl.config(text=ticker_err.get()))
    number_err.trace_add("write", lambda *_: number_err_lbl.config(text=number_err.get()))
    price_err.trace_add("write", lambda *_: price_err_lbl.config(text=price_err.get()))

    def clear_errors():
        date_err.set("")
        prefix_err.set("")
        ticker_err.set("")
        number_err.set("")
        price_err.set("")

    def clear_fields():
        date_var.set(date.today().strftime("%m/%d/%y"))
        prefix_var.set("1")
        number_var.set("")
        price_var.set("")

    # --- Actions

    def reload_tickers():
        nonlocal ticker_list
        ticker_list = load_tickers_from_stock()
        ticker_cb["values"] = ticker_list
        if ticker_list:
            ticker_var.set(ticker_list[0])
            ticker_err.set("")
        else:
            ticker_var.set("")
            ticker_err.set("No tickers in Stock.")

    ttk.Button(frm, text="Reload tickers", command=reload_tickers)\
        .grid(row=4, column=2, padx=5)

    def do_add():
        clear_errors()
        ok_d, res_d = validate_date(date_var.get())
        ok_pfx, res_pfx = validate_prefix(prefix_var.get())
        ok_t, res_t = validate_ticker(ticker_var.get(), ticker_list)
        ok_n, res_n = validate_number(number_var.get())
        ok_pr, res_pr = validate_price(price_var.get())

        if not ok_d:
            date_err.set(res_d)
        if not ok_pfx:
            prefix_err.set(res_pfx)
        if not ok_t:
            ticker_err.set(res_t)
        if not ok_n:
            number_err.set(res_n)
        if not ok_pr:
            price_err.set(res_pr)

        if not (ok_d and ok_pfx and ok_t and ok_n and ok_pr):
            messagebox.showerror("Validation", "Please correct highlighted fields.")
            return

        excel_date, prefix_i, ticker_s, number_i, price_f = \
            res_d, res_pfx, res_t, res_n, res_pr

        # Confirm purchase/sale
        if price_f < 0:
            if not messagebox.askyesno("Confirm", f"Price={price_f}. Confirm PURCHASE?"):
                return
        elif price_f > 0:
            if not messagebox.askyesno("Confirm", f"Price={price_f}. Confirm SALE?"):
                return

        try:
            add_transaction(excel_date, prefix_i, ticker_s, number_i, price_f)
            messagebox.showinfo("Add", "Transaction added.")
            clear_fields()
        except Exception as e:
            debug(f"do_add: error: {e}")
            messagebox.showerror("Error", f"Add failed:\n{e}")

    def do_delete():
        clear_errors()
        ok_d, res_d = validate_date(date_var.get())
        ok_pfx, res_pfx = validate_prefix(prefix_var.get())
        ok_t, res_t = validate_ticker(ticker_var.get(), ticker_list)
        ok_n, res_n = validate_number(number_var.get())
        ok_pr, res_pr = validate_price(price_var.get())

        if not ok_d:
            date_err.set(res_d)
        if not ok_pfx:
            prefix_err.set(res_pfx)
        if not ok_t:
            ticker_err.set(res_t)
        if not ok_n:
            number_err.set(res_n)
        if not ok_pr:
            price_err.set(res_pr)

        if not (ok_d and ok_pfx and ok_t and ok_n and ok_pr):
            messagebox.showerror("Validation", "Please correct highlighted fields.")
            return

        excel_date, prefix_i, ticker_s, number_i, price_f = \
            res_d, res_pfx, res_t, res_n, res_pr

        try:
            delete_transaction(excel_date, prefix_i, ticker_s, number_i, price_f)
            clear_fields()
        except Exception as e:
            debug(f"do_delete: error: {e}")
            messagebox.showerror("Error", f"Delete failed:\n{e}")

    def do_open_wb():
        append_log_entry("open", "Workbook opened from GUI.")
        try:
            if os.name == "posix":
                subprocess.call(["open", EXCEL_FILE])
            else:
                os.startfile(EXCEL_FILE)
        except Exception as e:
            debug(f"do_open_wb: {e}")
            messagebox.showerror("Error", f"Cannot open workbook:\n{e}")

    def do_exit():
        append_log_entry("exit", "GUI session closed.")
        root.destroy()

    # Button bar
    btn = ttk.Frame(frm)
    btn.grid(row=10, column=0, columnspan=3, pady=10, sticky="w")

    ttk.Button(btn, text="Add",    command=do_add).grid(row=0, column=0, padx=5)
    ttk.Button(btn, text="Delete", command=do_delete).grid(row=0, column=1, padx=5)
    ttk.Button(btn, text="Open WB",command=do_open_wb).grid(row=0, column=2, padx=5)
    ttk.Button(btn, text="Exit",   command=do_exit).grid(row=0, column=3, padx=5)

    return root


# -------------------------------------------------
# MAIN
# -------------------------------------------------

def main():
    debug(f"Using TRADES.xlsx at: {EXCEL_FILE}")
    # basic sanity check
    try:
        wb, ws_trans, ws_stock, ws_log = load_trades_workbook()
        debug(f"Workbook sheets: {wb.sheetnames}")
        _ = get_tabtrans(ws_trans)
        debug("TabTrans found OK.")
    except Exception as e:
        debug(f"Startup error: {e}")
        messagebox.showerror("Startup error", str(e))
        return

    root = build_gui()
    root.mainloop()


if __name__ == "__main__":
    main()