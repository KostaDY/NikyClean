import os
import time
import pandas as pd
import pytz
import traceback
import yfinance as yf
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess

# ============================================================
# SETTINGS
# ============================================================
WORKBOOK = "ASXData.xlsx"
INPUT_SHEET = "Tickers"
OUTPUT_SHEET = "LatestData"

TIMEZONE = pytz.timezone("Australia/Sydney")

HISTORY_PERIOD = "7d"
HISTORY_INTERVAL = "1d"
AVG_VOLUME_DAYS = 7

INTRADAY_PERIOD = "1d"
INTRADAY_INTERVAL = "1m"

SKIP_FULL_INFO_TICKERS = {"LOCK.L"}

YAHOO_DELAY_SEC = 0.3   # 300 ms abuse protection

# ============================================================
# EXCEL CONTROL (macOS)
# ============================================================
def close_excel_workbook(path):
    try:
        subprocess.run([
            "osascript",
            "-e",
            f'''
            tell application "Microsoft Excel"
                if exists workbook "{os.path.basename(path)}" then
                    close workbook "{os.path.basename(path)}" saving no
                end if
            end tell
            '''
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

def open_excel_workbook(path):
    try:
        subprocess.run(["open", path],
                       stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL)
    except Exception:
        pass

# ============================================================
# VERIFY FILE
# ============================================================
if not os.path.exists(WORKBOOK):
    raise FileNotFoundError(f"Workbook '{WORKBOOK}' not found")

close_excel_workbook(WORKBOOK)

# ============================================================
# READ TICKERS (HARD NAN FILTER)
# ============================================================
tickers_df = pd.read_excel(WORKBOOK, sheet_name=INPUT_SHEET)
col_name = "Ticker" if "Ticker" in tickers_df.columns else "Tickers"

tickers = (
    tickers_df[col_name]
    .dropna()
    .astype(str)
    .str.strip()
    .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA})
    .dropna()
    .tolist()
)

if not tickers:
    raise ValueError("No tickers found")

# ============================================================
# FETCH DATA
# ============================================================
rows = []

for ticker in tickers:
    time.sleep(YAHOO_DELAY_SEC)

    row = {
        "Ticker": ticker,
        "Status": "OK",
        "MissingFields": ""
    }

    try:
        tk = yf.Ticker(ticker)

        info = tk.fast_info if hasattr(tk, "fast_info") else {}
        full = {}

        if ticker not in SKIP_FULL_INFO_TICKERS:
            try:
                full = tk.info or {}
            except Exception:
                full = {}

        row["Open"] = (
            info.get("open")
            or full.get("regularMarketOpen")
            or full.get("open")
            or ""
        )

        row["MarketPrice"] = info.get("lastPrice", "")
        row["MarketDayHigh"] = info.get("dayHigh", "")
        row["MarketDayLow"] = info.get("dayLow", "")
        row["MarketPreviousClose"] = info.get("previousClose", "")
        row["Volume"] = info.get("volume", "")
        row["MarketTime"] = (
            info.get("lastTradeTime")
            or full.get("regularMarketTime")
            or full.get("marketTime")
            or ""
        )

        row["AverageVolume"] = ""

        row["marketCap"] = full.get("marketCap", "")
        row["fiftyTwoWeekHigh"] = full.get("fiftyTwoWeekHigh", "")
        row["fiftyTwoWeekLow"] = full.get("fiftyTwoWeekLow", "")
        row["currency"] = full.get("currency", "")

        row["targetMeanPrice"] = full.get("targetMeanPrice", "")
        row["recommendationMean"] = full.get("recommendationMean", "")

        row["dividendRate"] = full.get("dividendRate", "")
        row["dividendYield"] = full.get("dividendYield", "")
        row["exDividendDate"] = full.get("exDividendDate", "")

        # --- Earnings ---
        row["earningsDate"] = ""
        cal = tk.calendar

        if isinstance(cal, pd.DataFrame) and not cal.empty:
            if "Earnings Date" in cal.index:
                dates = cal.loc["Earnings Date"].tolist()
            elif "Earnings Date" in cal.columns:
                dates = cal["Earnings Date"].tolist()
            else:
                dates = None

            if dates:
                first = dates[0]
                if isinstance(first, (list, tuple)) and first:
                    first = first[0]
                row["earningsDate"] = first

        elif isinstance(cal, dict) and "Earnings Date" in cal:
            d = cal.get("Earnings Date")
            if isinstance(d, (list, tuple)) and d:
                row["earningsDate"] = d[0]
            else:
                row["earningsDate"] = d

        # --- History ---
        try:
            hist = tk.history(period=HISTORY_PERIOD, interval=HISTORY_INTERVAL)
            if not hist.empty:
                if not row["Open"] and "Open" in hist:
                    row["Open"] = hist["Open"].dropna().iloc[-1]

                if "Volume" in hist:
                    vol = hist["Volume"].dropna()
                    if not vol.empty:
                        row["Volume"] = row["Volume"] or vol.iloc[-1]
                        row["AverageVolume"] = vol.tail(AVG_VOLUME_DAYS).mean()

                if not row["MarketTime"]:
                    row["MarketTime"] = hist.index[-1]
        except Exception:
            pass

        if not row["AverageVolume"]:
            row["AverageVolume"] = (
                full.get("averageVolume")
                or full.get("averageVolume10days")
                or ""
            )

        if not row["MarketTime"]:
            try:
                intraday = tk.history(period=INTRADAY_PERIOD, interval=INTRADAY_INTERVAL)
                if not intraday.empty:
                    row["MarketTime"] = intraday.index[-1]
            except Exception:
                pass

        missing = [k for k, v in row.items() if v in ("", None)]
        if len(missing) > 3:
            row["Status"] = "PARTIAL"
            row["MissingFields"] = ",".join(missing)

    except Exception as e:
        row["Status"] = "NO_DATA"
        row["MissingFields"] = str(e)

    rows.append(row)

# ============================================================
# DATAFRAME
# ============================================================
df = pd.DataFrame(rows)

# ============================================================
# TIME NORMALIZATION
# ============================================================
def normalize_time(x):
    try:
        ts = pd.to_datetime(x, utc=True, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.tz_convert(TIMEZONE).strftime("%d-%b-%Y %H:%M")
    except Exception:
        return ""

def normalize_date(x):
    try:
        ts = pd.to_datetime(x, utc=True, errors="coerce")
        return "" if pd.isna(ts) else ts.strftime("%d-%b-%Y")
    except Exception:
        return ""

if "MarketTime" in df:
    df["MarketTime"] = df["MarketTime"].apply(normalize_time)

for col in ("exDividendDate", "earningsDate"):
    if col in df:
        df[col] = df[col].apply(normalize_date)

df = df.fillna("")
df.columns = df.columns.astype(str)

# ============================================================
# EXCEL TABLE HELPERS
# ============================================================
def add_or_replace_table(ws, name):
    try:
        ws.tables.pop(name, None)
    except Exception:
        pass

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )
    ws.add_table(table)

# ============================================================
# WRITE TO EXCEL
# ============================================================
with pd.ExcelWriter(
    WORKBOOK,
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    df.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
    ws = writer.sheets[OUTPUT_SHEET]
    add_or_replace_table(ws, "RTdata")

open_excel_workbook(WORKBOOK)