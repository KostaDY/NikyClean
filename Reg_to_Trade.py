#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# SETTINGS
# ============================================================

CSV_PATH = Path("/Users/kostayanev/NikyClean/TradeArchive.csv")
OUT_XLSX = Path("/Users/kostayanev/NikyClean/TradeTable.xlsx")

SHEET_NAME = "TradeTable"
TABLE_NAME = "TabTradeTable"


# ============================================================
# LOAD CSV WITHOUT HEADERS
# ============================================================

def load_trades(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(
        csv_path,
        header=None,
        names=["Date", "Ticker", "Number", "Price"],
        skip_blank_lines=True
    )

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Ticker"] = df["Ticker"].astype(str).str.strip()
    df["Number"] = pd.to_numeric(df["Number"], errors="coerce")
    df["Price"] = pd.to_numeric(df["Price"], errors="coerce")

    df = df.dropna(subset=["Date", "Ticker", "Number", "Price"])

    df = df[df["Ticker"] != ""]
    df = df[df["Number"] != 0]
    df = df[df["Price"] != 0]

    # keep original file order inside each ticker
    df["_Order"] = range(len(df))

    return df


# ============================================================
# PRICE-BASED SALE MATCHING
# ============================================================

def match_ticker_lots(tdf: pd.DataFrame):
    """
    For one ticker:
    - purchases are all negative-price rows
    - sales are all positive-price rows
    - sales reduce purchase lots by price logic only

    Matching rule:
    1. Prefer nearest lower/equal buy price by absolute value.
       Example sale 195, buys 180 and 150 -> match 180 first.
    2. If sale price is below all buy prices, use nearest buy price.
       Example sale 170, buys 180 and 200 -> match 180 first.
    """

    purchases = []
    sales = []

    for _, row in tdf.sort_values("_Order").iterrows():
        date = row["Date"]
        number = abs(float(row["Number"]))
        price = float(row["Price"])

        if price < 0:
            purchases.append({
                "date": date,
                "number": number,
                "price": price
            })
        else:
            sales.append({
                "number": number,
                "price": price
            })

    lots = purchases[:]

    total_buy = sum(lot["number"] for lot in lots)
    total_sale = sum(sale["number"] for sale in sales)

    if total_sale > total_buy + 1e-9:
        raise ValueError(
            f"Sales exceed purchases: BUY={total_buy}, SALE={total_sale}"
        )

    for sale in sales:
        sale_qty = sale["number"]
        sale_price = abs(sale["price"])

        while sale_qty > 1e-12:

            if not lots:
                raise ValueError("Internal error: no lots left to match sale.")

            # Lots with buy price <= sale price
            lower_or_equal = [
                i for i, lot in enumerate(lots)
                if abs(lot["price"]) <= sale_price
            ]

            if lower_or_equal:
                # nearest lower/equal = highest buy price below/equal sale price
                best_i = max(
                    lower_or_equal,
                    key=lambda i: abs(lots[i]["price"])
                )
            else:
                # sale below all buy prices:
                # match nearest buy price by absolute distance
                best_i = min(
                    range(len(lots)),
                    key=lambda i: abs(abs(lots[i]["price"]) - sale_price)
                )

            lot = lots[best_i]
            matched = min(sale_qty, lot["number"])

            lot["number"] -= matched
            sale_qty -= matched

            if lot["number"] <= 1e-12:
                lots.pop(best_i)

    return lots


def build_result(df: pd.DataFrame) -> dict:
    result = {}

    for ticker, tdf in df.groupby("Ticker", sort=True):
        result[ticker] = match_ticker_lots(tdf)

    return result


# ============================================================
# WRITE XLSX
# ============================================================

def write_trade_table(result: dict, out_xlsx: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    max_lots = max((len(lots) for lots in result.values()), default=0)

    headers = ["Ticker"]

    for i in range(1, max_lots + 1):
        headers += [f"D{i}", f"N{i}", f"P{i}"]

    ws.append(headers)

    for ticker, lots in result.items():
        row = [ticker]

        for lot in lots:
            row += [
                lot["date"].date(),
                lot["number"],
                lot["price"]
            ]

        # pad row so Excel table is rectangular
        row += [""] * (len(headers) - len(row))

        ws.append(row)

    # Header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Column formatting
    ws.column_dimensions["A"].width = 14

    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value

            if str(header).startswith("D"):
                cell.number_format = "yyyy-mm-dd"

            elif str(header).startswith("N"):
                cell.number_format = "0"

            elif str(header).startswith("P"):
                cell.number_format = "0.00"

    # Create Excel table
    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    tab = Table(
        displayName=TABLE_NAME,
        ref=table_ref
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.freeze_panes = "B2"

    wb.save(out_xlsx)


# ============================================================
# MAIN
# ============================================================

def main():
    df = load_trades(CSV_PATH)
    result = build_result(df)
    write_trade_table(result, OUT_XLSX)

    print(f"✅ Created: {OUT_XLSX}")


if __name__ == "__main__":
    main()