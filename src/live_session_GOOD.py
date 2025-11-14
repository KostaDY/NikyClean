#!/usr/bin/env python3
"""
Live session bridge:
- Read tickers from excel/DataSource.xlsx (DataRT / RTData / Ticker_Symbol)
- Fetch live data via yahooquery (+ HTML scraping in slow mode)
- Write clean excel/DataSource_Raw.xlsx with table RTData_Raw
"""

from __future__ import annotations

import argparse
import time
import subprocess
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any

import pandas as pd
import requests
from bs4 import BeautifulSoup
from yahooquery import Ticker

from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00


# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "excel"

DATASOURCE_FILE = EXCEL_DIR / "DataSource.xlsx"
RAW_FILE = EXCEL_DIR / "DataSource_Raw.xlsx"

DATASOURCE_SHEET = "DataRT"
DATASOURCE_TABLE = "RTData"

RAW_SHEET = "RTData_Raw"
RAW_TABLE = "RTData_Raw"

# Final column order (FAST + SLOW)
COLUMNS = [
    "Ticker",
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "Dividend",
    "RepDiv",
]


# ---------------------------------------------------------------------
# UTILS
# ---------------------------------------------------------------------

def ts() -> str:
    """Current time in HH:MM:SS."""
    return datetime.now().strftime("%H:%M:%S")


def open_excel_file(path: Path) -> None:
    """Open an Excel file in macOS Finder."""
    try:
        subprocess.run(["open", str(path)], check=False)
    except Exception:
        # Non-fatal: user can open manually.
        pass


# ---------------------------------------------------------------------
# STEP 1: READ TICKERS FROM DataSource.xlsx / DataRT / RTData
# ---------------------------------------------------------------------

def read_tickers() -> List[Optional[str]]:
    """
    Read tickers from DataSource.xlsx → DataRT → Excel table RTData → column 'Ticker_Symbol'.
    Preserve empty rows.
    """
    if not DATASOURCE_FILE.exists():
        raise FileNotFoundError(f"DataSource.xlsx not found at {DATASOURCE_FILE}")

    wb = load_workbook(DATASOURCE_FILE, data_only=True)
    if DATASOURCE_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{DATASOURCE_SHEET}' not found in {DATASOURCE_FILE.name}")

    ws: Worksheet = wb[DATASOURCE_SHEET]

    if isinstance(ws._tables, dict):
        tables = ws._tables
    else:
        # Very old openpyxl versions: list-like; convert to dict
        tables = {t.name: t for t in ws._tables}

    if DATASOURCE_TABLE not in tables:
        raise KeyError(f"Excel table '{DATASOURCE_TABLE}' not found in sheet '{DATASOURCE_SHEET}'")

    tbl = tables[DATASOURCE_TABLE]
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)

    # Header row is min_row
    header_row = ws[min_row]
    ticker_col_idx = None
    for cell in header_row:
        if str(cell.value).strip() == "Ticker_Symbol":
            ticker_col_idx = cell.column
            break

    if ticker_col_idx is None:
        raise KeyError("Column 'Ticker_Symbol' not found in RTData header row")

    tickers: List[Optional[str]] = []
    for r in range(min_row + 1, max_row + 1):
        cell = ws.cell(row=r, column=ticker_col_idx)
        v = cell.value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            tickers.append(None)
        else:
            tickers.append(str(v).strip())

    return tickers


# ---------------------------------------------------------------------
# STEP 2: FAST FETCH (always used)
# ---------------------------------------------------------------------

def fetch_fast(tickers: List[Optional[str]]) -> pd.DataFrame:
    """
    Fast fields via yahooquery:
    Close, Open, Last, Low, High, P/E, Change, ChangePct, Volume, VolumeAverage.
    FIXED:
    - Change and ChangePct correctly read from price OR summaryDetail.
    """
    syms = [t for t in tickers if t]
    if not syms:
        # Only empty rows
        return pd.DataFrame({col: [] for col in COLUMNS})

    tq = Ticker(syms, asynchronous=True)

    raw_price = tq.price or {}
    raw_sd = tq.summary_detail or {}

    def get_dict(d: Any, key: str) -> Dict[str, Any]:
        v = d.get(key)
        return v if isinstance(v, dict) else {}

    # Normalize per ticker
    price = {s: get_dict(raw_price, s) for s in syms}
    sd = {s: get_dict(raw_sd, s) for s in syms}

    rows: Dict[str, Dict[str, Any]] = {}

    for s in syms:
        p = price.get(s, {})
        d = sd.get(s, {})

        # Prefer summaryDetail with fallback to price
        close = d.get("regularMarketPreviousClose") or p.get("regularMarketPreviousClose")
        open_ = d.get("regularMarketOpen") or p.get("regularMarketOpen")
        last = p.get("regularMarketPrice") or d.get("regularMarketPrice")
        low = d.get("regularMarketDayLow") or p.get("regularMarketDayLow")
        high = d.get("regularMarketDayHigh") or p.get("regularMarketDayHigh")
        pe = d.get("trailingPE") or p.get("trailingPE")

        # --- FIXED CHANGE ---
        change = (
            d.get("regularMarketChange")
            or p.get("regularMarketChange")
        )

        # --- FIXED CHANGE % ---
        change_pct = (
            d.get("regularMarketChangePercent")
            or p.get("regularMarketChangePercent")
        )

        # Normalize percent if needed
        if change_pct is not None:
            try:
                if abs(change_pct) > 1:
                    change_pct = change_pct / 100.0
            except Exception:
                pass

        volume = d.get("regularMarketVolume") or p.get("regularMarketVolume")
        volavg = d.get("averageVolume") or p.get("averageVolume")

        rows[s] = {
            "Ticker": s,
            "Close": close,
            "Open": open_,
            "Last": last,
            "Low": low,
            "High": high,
            "P/E": pe,
            "Change": change,
            "ChangePct": change_pct,
            "Volume": volume,
            "VolumeAverage": volavg,
        }

    # Preserve empty ticker rows
    empty_fast = {col: None for col in COLUMNS}
    empty_fast["Ticker"] = None

    result_rows = []
    for t in tickers:
        if t and t in rows:
            result_rows.append(rows[t])
        else:
            result_rows.append(empty_fast.copy())

    df = pd.DataFrame(result_rows)

    # Ensure all final columns exist
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None

    return df[COLUMNS]


# ---------------------------------------------------------------------
# STEP 3: SLOW FETCH (optional, adds Beta, 1YT, Ddate, EarningDate, Dividend, RepDiv)
# ---------------------------------------------------------------------

def _parse_yahoo_date(s: Optional[str]) -> Optional[datetime]:
    """Parse Yahoo-like date strings into datetime, or return None on failure."""
    if not s:
        return None
    # Some earnings strings are ranges like "Nov 06, 2025 - Nov 10, 2025"
    s = s.split(" - ")[0].strip()
    s = s.split(" to ")[0].strip()

    for fmt in ("%b %d, %Y", "%b %d %Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def fetch_slow(tickers: List[Optional[str]]) -> pd.DataFrame:
    """
    Slow mode: fetch expensive/static fields (once per day).
    Beta, 1YT (target mean), Ddate (ex-div date), EarningDate, Dividend, RepDiv.
    Includes internal HTML scraping; no external modules.
    """
    syms = [t for t in tickers if t]
    if not syms:
        empty = {
            "Ticker": None,
            "Beta": None,
            "1YT": None,
            "Ddate": None,
            "EarningDate": None,
            "Dividend": None,
            "RepDiv": None,
        }
        return pd.DataFrame([empty for _ in tickers])

    tq = Ticker(syms, asynchronous=True)

    raw_sd = tq.summary_detail or {}
    raw_ks = tq.all_modules.get("keyStatistics", {})

    def safe_dict(src: Any, key: str) -> Dict[str, Any]:
        v = src.get(key)
        return v if isinstance(v, dict) else {}

    sd = {s: safe_dict(raw_sd, s) for s in syms}
    ks = {s: safe_dict(raw_ks, s) for s in syms}

    # HTML fallback for 1YT, ex-div, earnings date
    def scrape_html(ticker: str) -> Dict[str, Any]:
        url = f"https://finance.yahoo.com/quote/{ticker}"
        h = {"targetMeanPrice": None, "exDividendDate": None, "earningsDate": None}
        try:
            resp = requests.get(
                url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=7,
            )
            if resp.status_code != 200:
                return h
            soup = BeautifulSoup(resp.text, "lxml")

            # 1y Target Est
            t1 = soup.find("span", string="1y Target Est")
            if t1:
                v = t1.find_next("span")
                if v:
                    txt = v.text.replace(",", "").strip()
                    try:
                        h["targetMeanPrice"] = float(txt)
                    except Exception:
                        pass

            # Ex-Dividend Date
            t2 = soup.find("span", string="Ex-Dividend Date")
            if t2:
                v = t2.find_next("span")
                if v:
                    h["exDividendDate"] = v.text.strip()

            # Earnings Date
            t3 = soup.find("span", string="Earnings Date")
            if t3:
                v = t3.find_next("span")
                if v:
                    h["earningsDate"] = v.text.strip()

        except Exception:
            # Ignore HTML errors per ticker
            pass

        return h

    html = {s: scrape_html(s) for s in syms}

    rows: Dict[str, Dict[str, Any]] = {}
    for s in syms:
        d = sd.get(s, {})
        k = ks.get(s, {})
        h = html.get(s, {})

        beta = d.get("beta") or k.get("beta")
        div_yield = d.get("dividendYield") or k.get("dividendYield")

        target = h.get("targetMeanPrice")
        exdiv_raw = h.get("exDividendDate")
        ear_raw = h.get("earningsDate")

        exdiv_dt = _parse_yahoo_date(exdiv_raw)
        earn_dt = _parse_yahoo_date(ear_raw)

        rows[s] = {
            "Ticker": s,
            "Beta": beta,
            "1YT": target,
            "Ddate": exdiv_dt,
            "EarningDate": earn_dt,
            "Dividend": div_yield,
            "RepDiv": div_yield,  # Keep Excel's column; same semantic by default
        }

    empty_slow = {
        "Ticker": None,
        "Beta": None,
        "1YT": None,
        "Ddate": None,
        "EarningDate": None,
        "Dividend": None,
        "RepDiv": None,
    }

    result = []
    for t in tickers:
        if t and t in rows:
            result.append(rows[t])
        else:
            result.append(empty_slow.copy())

    return pd.DataFrame(result)


# ---------------------------------------------------------------------
# STEP 4: WRITE CLEAN Excel workbook DataSource_Raw.xlsx
# ---------------------------------------------------------------------

def write_raw_excel(df: pd.DataFrame) -> None:
    """
    Create a fresh Excel workbook with:
    - File: excel/DataSource_Raw.xlsx
    - Sheet: RTData_Raw
    - Table: RTData_Raw
    - Data from df (COLUMNS)
    """
    RAW_FILE.parent.mkdir(parents=True, exist_ok=True)
    if RAW_FILE.exists():
        RAW_FILE.unlink()

    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET

    # Write header
    for j, col in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=j, value=col)

    # Write data rows
    for i, row in enumerate(df[COLUMNS].itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    max_row = ws.max_row
    max_col = ws.max_column
    ref = f"A1:{chr(64 + max_col)}{max_row}"

    table = Table(displayName=RAW_TABLE, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Number formats
    # Map column name -> format
    col_fmt = {
        "Close": FORMAT_NUMBER_00,
        "Open": FORMAT_NUMBER_00,
        "Last": FORMAT_NUMBER_00,
        "Low": FORMAT_NUMBER_00,
        "High": FORMAT_NUMBER_00,
        "P/E": FORMAT_NUMBER_00,
        "Volume": "#,##0",
        "VolumeAverage": "#,##0",
        "Beta": FORMAT_NUMBER_00,
        "1YT": FORMAT_NUMBER_00,
        "Change": FORMAT_NUMBER_00,
        "ChangePct": FORMAT_PERCENTAGE_00,
        "Dividend": FORMAT_PERCENTAGE_00,
        "RepDiv": FORMAT_PERCENTAGE_00,
        "Ddate": "mmm d, yyyy",
        "EarningDate": "mmm d, yyyy",
    }

    # Apply formats by scanning header row
    header_row = ws[1]
    for cell in header_row:
        col_name = str(cell.value)
        if col_name in col_fmt:
            nf = col_fmt[col_name]
            col_letter = cell.column_letter
            for r in range(2, max_row + 1):
                ws[f"{col_letter}{r}"].number_format = nf

    wb.save(RAW_FILE)


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Live session fast/slow updater")
    parser.add_argument(
        "--slow",
        action="store_true",
        help="Fetch slow fields as well (Beta, 1YT, Ddate, EarningDate, Dividend, RepDiv)",
    )
    args = parser.parse_args()

    mode = "SLOW" if args.slow else "FAST"
    print(f"🕒 START at {ts()} — MODE: {mode}")

    t0 = time.time()
    print(f"📥 Reading tickers from {DATASOURCE_FILE.name} → {DATASOURCE_TABLE}…")
    tick0 = time.time()
    tickers = read_tickers()
    tick1 = time.time()
    print(f"✔ Loaded {len(tickers)} tickers (empty rows kept).")
    print(f"⏱ Tickers read in {tick1 - tick0:.3f} sec")

    # Fast data
    fast0 = time.time()
    df = fetch_fast(tickers)
    fast1 = time.time()
    print(f"⏱ Fast fetch in {fast1 - fast0:.3f} sec")

    # Slow mode: enrich df
    if args.slow:
        slow0 = time.time()
        df_s = fetch_slow(tickers)
        slow1 = time.time()
        print(f"⏱ Slow fetch in {slow1 - slow0:.3f} sec")

        for col in ["Beta", "1YT", "Ddate", "EarningDate", "Dividend", "RepDiv"]:
            if col in df_s.columns:
                df[col] = df_s[col]
            else:
                df[col] = None

    # Write Excel
    w0 = time.time()
    print(f"📝 Writing {RAW_FILE.name} …")
    write_raw_excel(df)
    w1 = time.time()
    print(f"⏱ Excel written in {w1 - w0:.3f} sec")

    total = time.time() - t0
    print(f"✅ Finished. Total time: {total:.3f} sec")

    # Auto-open result
    open_excel_file(RAW_FILE)


if __name__ == "__main__":
    main()