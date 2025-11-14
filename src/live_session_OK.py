from pathlib import Path
from datetime import datetime, timedelta
import random

import pandas as pd
from openpyxl import Workbook, load_workbook

# -------------------------------------------------------
# CONFIG
# -------------------------------------------------------
DATAFILE = Path("excel/DataSource.xlsx")      # READ-ONLY always
RAWFILE  = Path("excel/DataSource_Raw.xlsx")  # ALWAYS recreated
RAW_SHEET = "RTData_Raw"

RAW_COLUMNS = [
    "Ticker",
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "Dividend",
]

# -------------------------------------------------------
# 1. READ TICKERS FROM DATASOURCE (READ-ONLY)
# -------------------------------------------------------
def load_tickers():
    """Read tickers from DataSource.xlsx, table RTData, column Ticker_Symbol."""
    if not DATAFILE.exists():
        raise FileNotFoundError("DataSource.xlsx not found.")

    df = pd.read_excel(
        DATAFILE,
        sheet_name="DataRT",
        header=1,
        usecols=["Ticker_Symbol"],
    )
    # KEEP empty rows
    return df["Ticker_Symbol"].tolist()


# -------------------------------------------------------
# 2. SYNTHETIC MARKET DATA PER TICKER
# -------------------------------------------------------
def generate_fake_row(ticker):
    """Generate placeholder data for each ticker."""

    if pd.isna(ticker) or str(ticker).strip() == "":
        # Keep empty row but with all empty values
        return {col: None for col in RAW_COLUMNS}

    ticker = str(ticker).strip()
    base = random.uniform(10, 300)

    close = round(base + random.uniform(-2, 2), 2)
    open_ = round(base + random.uniform(-3, 3), 2)
    last = round(base + random.uniform(-1, 1), 2)
    low  = round(min(close, open_) - random.uniform(0, 3), 2)
    high = round(max(close, open_) + random.uniform(0, 3), 2)

    today = datetime.today().date()
    dd = today + timedelta(days=random.randint(-60, 60))
    ed = today + timedelta(days=random.randint(-120, 120))

    return {
        "Ticker": ticker,
        "Close": close,
        "Open": open_,
        "Last": last,
        "Low": low,
        "High": high,
        "P/E": round(random.uniform(5, 40), 2),
        "Change": round(last - close, 2),
        "ChangePct": round(random.uniform(-0.05, 0.05), 4),
        "Volume": random.randint(100_000, 50_000_000),
        "VolumeAverage": random.randint(100_000, 50_000_000),
        "Beta": round(random.uniform(0.5, 2.0), 2),
        "1YT": round(base * random.uniform(0.7, 1.5), 2),
        "Ddate": dd,
        "EarningDate": ed,
        "Dividend": round(random.uniform(0.0, 0.05), 4),
    }


# -------------------------------------------------------
# 3. CREATE A NEW, CLEAN DATASOURCE_RAW WORKBOOK
# -------------------------------------------------------
def create_clean_raw_workbook():
    """Always recreate DataSource_Raw.xlsx fresh and clean."""

    if RAWFILE.exists():
        RAWFILE.unlink()  # HARD DELETE

    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET

    # Insert header row
    ws.append(RAW_COLUMNS)

    # Save now (empty except header)
    wb.save(RAWFILE)
    print(f"✔ Created new clean workbook: {RAWFILE}")


# -------------------------------------------------------
# 4. WRITE RAW DATA INTO CLEAN WORKBOOK
# -------------------------------------------------------
def write_raw_data(df):
    wb = load_workbook(RAWFILE)
    ws = wb[RAW_SHEET]

    start_row = 2

    # Insert rows
    for i, rec in enumerate(df.to_dict(orient="records"), start=start_row):
        for col_idx, col_name in enumerate(RAW_COLUMNS, start=1):
            ws.cell(row=i, column=col_idx, value=rec[col_name])

    # Apply formatting
    float_cols = {"Close","Open","Last","Low","High","P/E","Change","Beta","1YT"}
    int_cols   = {"Volume","VolumeAverage"}
    pct_cols   = {"ChangePct","Dividend"}
    date_cols  = {"Ddate","EarningDate"}

    col_map = {i+1: name for i, name in enumerate(RAW_COLUMNS)}
    last_row = start_row + len(df) - 1

    for row in range(start_row, last_row + 1):
        for col in range(1, len(RAW_COLUMNS) + 1):
            name = col_map[col]
            cell = ws.cell(row=row, column=col)

            if name in float_cols:
                cell.number_format = "0.00"
            elif name in int_cols:
                cell.number_format = "#,##0"
            elif name in pct_cols:
                cell.number_format = "0.00%"
            elif name in date_cols:
                cell.number_format = "mmm d, yyyy"

    wb.save(RAWFILE)
    print(f"✔ Wrote {len(df)} rows into {RAW_SHEET}")


# -------------------------------------------------------
# MAIN
# -------------------------------------------------------
def main():
    print("Reading tickers from DataSource.xlsx …")
    tickers = load_tickers()
    print(f"✔ Loaded {len(tickers)} tickers")

    print("Generating synthetic data …")
    rows = [generate_fake_row(t) for t in tickers]
    df = pd.DataFrame(rows, columns=RAW_COLUMNS)

    create_clean_raw_workbook()
    write_raw_data(df)

    print("✔ Finished clean export into DataSource_Raw.xlsx")


if __name__ == "__main__":
    main()