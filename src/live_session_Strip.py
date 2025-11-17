from __future__ import annotations

import argparse
import platform
import subprocess
import time
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from yahooquery import Ticker

# --------------------------------------------------------------------
# Paths & constants
# --------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "excel"
DATASOURCE_XLSX = EXCEL_DIR / "DataSource.xlsx"
RAW_XLSX = EXCEL_DIR / "DataSource_Raw.xlsx"

TICKER_SHEET = "TickerList"
TICKER_TABLE_COL = "Ticker"   # column name in sheet TickerList

RAW_SHEET = "RTData_Raw"
RAW_TABLE_NAME = "RTData_Raw"  # (only used for naming, no real table object)

# Column layout for the output
FAST_COLS = [
    "Ticker",
    "RefreshTime",      # from regularMarketTime
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
]

SLOW_COLS = [
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "RepDiv",
]

ALL_COLS = FAST_COLS + SLOW_COLS


# --------------------------------------------------------------------
# Utility: Open file on macOS
# --------------------------------------------------------------------

def open_excel_file(path: Path) -> None:
    """Open a file in the default app (macOS only)."""
    if platform.system() == "Darwin":
        try:
            subprocess.Popen(["open", str(path)])
        except Exception as e:
            print(f"⚠ Could not auto-open Excel file: {e}")


# --------------------------------------------------------------------
# Step 1: Read tickers from DataSource.xlsx / TickerList
# --------------------------------------------------------------------

def read_tickers() -> List[Optional[str]]:
    """
    Read tickers from DataSource.xlsx, sheet 'TickerList', column 'Ticker'.

    - Keeps row order exactly as in Excel.
    - Keeps empty rows (returns None there) so output row count matches.
    """
    if not DATASOURCE_XLSX.exists():
        raise FileNotFoundError(f"DataSource.xlsx not found at {DATASOURCE_XLSX}")

    df = pd.read_excel(DATASOURCE_XLSX, sheet_name=TICKER_SHEET)
    if TICKER_TABLE_COL not in df.columns:
        raise KeyError(
            f"Column '{TICKER_TABLE_COL}' not found in sheet '{TICKER_SHEET}' "
            f"of {DATASOURCE_XLSX}"
        )

    tickers: List[Optional[str]] = []
    for val in df[TICKER_TABLE_COL].tolist():
        if pd.isna(val):
            tickers.append(None)
        else:
            s = str(val).strip()
            tickers.append(s if s else None)

    return tickers


# --------------------------------------------------------------------
# Helpers for yahooquery
# --------------------------------------------------------------------

def _safe_get(d: Any, key: str, default: Any = None) -> Any:
    if isinstance(d, dict):
        return d.get(key, default)
    return default


def _epoch_to_datetime(ts: Any) -> Optional[datetime]:
    if ts in (None, "", 0):
        return None
    try:
        # yahooquery returns seconds since epoch
        return datetime.fromtimestamp(float(ts))
    except Exception:
        return None


def _epoch_to_date(ts: Any) -> Optional[date]:
    dt = _epoch_to_datetime(ts)
    return dt.date() if dt else None


# --------------------------------------------------------------------
# Step 2: FAST fetch — prices / volumes / RefreshTime
# --------------------------------------------------------------------

def fetch_fast(tickers_raw: List[Optional[str]]) -> pd.DataFrame:
    """
    FAST mode:
    - Fetches market data for all distinct, non-empty tickers via yahooquery.
    - Returns a DataFrame with one row per *input* row (keeps blanks).
    """
    # Build list of valid tickers (non-empty)
    valid = [t for t in tickers_raw if t]
    unique_valid = sorted(set(valid))

    # Prepare empty DataFrame if no tickers
    if not unique_valid:
        return pd.DataFrame(columns=FAST_COLS)

    tq = Ticker(unique_valid, asynchronous=True)

    price_data = tq.price or {}
    summary_detail = tq.summary_detail or {}

    rows: List[Dict[str, Any]] = []

    for raw_ticker in tickers_raw:
        if not raw_ticker:
            # Preserve blank row
            rows.append({col: None for col in FAST_COLS})
            continue

        p = price_data.get(raw_ticker, {})
        sd = summary_detail.get(raw_ticker, {})

        last = _safe_get(p, "regularMarketPrice")
        close = _safe_get(p, "regularMarketPreviousClose")
        open_ = _safe_get(p, "regularMarketOpen")
        low = _safe_get(p, "regularMarketDayLow")
        high = _safe_get(p, "regularMarketDayHigh")
        change = _safe_get(p, "regularMarketChange")
        change_pct = _safe_get(p, "regularMarketChangePercent")
        volume = _safe_get(p, "regularMarketVolume")

        # Use trailingPE as main P/E, fallback to forwardPE
        pe = _safe_get(sd, "trailingPE")
        if pe is None:
            pe = _safe_get(sd, "forwardPE")

        # Volume average: try a few common keys
        vol_avg = (
            _safe_get(sd, "averageVolume")
            or _safe_get(sd, "averageDailyVolume3Month")
            or _safe_get(sd, "averageDailyVolume10Day")
        )

        # Refresh time from regularMarketTime
        rmt = _safe_get(p, "regularMarketTime")
        refresh_dt = _epoch_to_datetime(rmt)

        rows.append(
            {
                "Ticker": raw_ticker,
                "RefreshTime": refresh_dt,
                "Close": close,
                "Open": open_,
                "Last": last,
                "Low": low,
                "High": high,
                "P/E": pe,
                "Change": change,
                "ChangePct": change_pct,
                "Volume": volume,
                "VolumeAverage": vol_avg,
            }
        )

    return pd.DataFrame(rows, columns=FAST_COLS)


# --------------------------------------------------------------------
# Step 3: SLOW fetch — Beta, 1YT, Ddate, EarningDate, RepDiv
# --------------------------------------------------------------------

def fetch_slow(tickers_raw: List[Optional[str]]) -> pd.DataFrame:
    """
    SLOW mode:
    - Fetch Beta, 1YT, Ddate, EarningDate, RepDiv via yahooquery.
    - Still uses only the API (no HTML scraping).
    """
    valid = [t for t in tickers_raw if t]
    unique_valid = sorted(set(valid))

    if not unique_valid:
        return pd.DataFrame(columns=["Ticker"] + SLOW_COLS)

    tq = Ticker(unique_valid, asynchronous=True)

    summary_detail = tq.summary_detail or {}
    financial_data = getattr(tq, "financial_data", None)
    calendar_events = getattr(tq, "calendar_events", None)
    key_stats = getattr(tq, "key_stats", None)

    if not isinstance(financial_data, dict):
        financial_data = {}
    if not isinstance(calendar_events, dict):
        calendar_events = {}
    if not isinstance(key_stats, dict):
        key_stats = {}

    rows: List[Dict[str, Any]] = []

    for raw_ticker in tickers_raw:
        if not raw_ticker:
            rows.append(
                {
                    "Ticker": None,
                    "Beta": None,
                    "1YT": None,
                    "Ddate": None,
                    "EarningDate": None,
                    "RepDiv": None,
                }
            )
            continue

        sd = summary_detail.get(raw_ticker, {})
        ks = key_stats.get(raw_ticker, {})
        fin = financial_data.get(raw_ticker, {})
        cal = calendar_events.get(raw_ticker, {})

        # Beta (from summary_detail or key_stats)
        beta = _safe_get(sd, "beta")
        if beta is None:
            beta = _safe_get(ks, "beta")

        # 1YT — target mean price, typically in financial_data
        one_yt = _safe_get(fin, "targetMeanPrice")

        # Ddate — ex-dividend date
        ex_div_ts = _safe_get(sd, "exDividendDate")
        ddate = _epoch_to_date(ex_div_ts)

        # EarningDate — from calendar_events. Structure may vary.
        earning_date: Optional[date] = None
        earn = cal.get("earnings")
        if isinstance(earn, dict):
            ed_val = earn.get("earningsDate")
            # Could be list, dict, or timestamp-like
            if isinstance(ed_val, list) and ed_val:
                # Yahoo often stores list of timestamps or datelike objects
                candidate = ed_val[0]
                if isinstance(candidate, (int, float)):
                    earning_date = _epoch_to_date(candidate)
                elif isinstance(candidate, (datetime, date)):
                    earning_date = candidate.date() if isinstance(candidate, datetime) else candidate
            elif isinstance(ed_val, (int, float)):
                earning_date = _epoch_to_date(ed_val)
            elif isinstance(ed_val, (datetime, date)):
                earning_date = ed_val.date() if isinstance(ed_val, datetime) else ed_val

        # RepDiv — dividend yield as fraction (e.g. 0.034)
        rep_div = _safe_get(sd, "dividendYield")
        if rep_div is None:
            # fallback: trailingAnnualDividendYield is also a fraction
            rep_div = _safe_get(sd, "trailingAnnualDividendYield")

        rows.append(
            {
                "Ticker": raw_ticker,
                "Beta": beta,
                "1YT": one_yt,
                "Ddate": ddate,
                "EarningDate": earning_date,
                "RepDiv": rep_div,
            }
        )

    return pd.DataFrame(rows, columns=["Ticker"] + SLOW_COLS)


# --------------------------------------------------------------------
# Step 4: Write DataSource_Raw.xlsx / RTData_Raw
# --------------------------------------------------------------------

def write_raw_excel(df: pd.DataFrame) -> None:
    """
    Write df into excel/DataSource_Raw.xlsx, sheet RTData_Raw,
    with header, formatting, and one row per df row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET

    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = Font(bold=True)

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Formatting
    header_map = {name: idx + 1 for idx, name in enumerate(df.columns)}

    # Helper to apply number format to a column
    def fmt(col_name: str, num_format: str):
        idx = header_map.get(col_name)
        if not idx:
            return
        col_letter = get_column_letter(idx)
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"].number_format = num_format

    # Prices & numeric fields
    for name in ["Close", "Open", "Last", "Low", "High", "Change"]:
        fmt(name, "0.00")

    fmt("P/E", "0.00")
    fmt("Beta", "0.00")

    # Big numbers with thousand separator
    for name in ["Volume", "VolumeAverage", "1YT"]:
        fmt(name, "#,##0")

    # Percent fields
    for name in ["ChangePct", "RepDiv"]:
        fmt(name, "0.00%")

    # Dates / datetimes
    # RefreshTime = full datetime
    idx_rt = header_map.get("RefreshTime")
    if idx_rt:
        col_letter = get_column_letter(idx_rt)
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"].number_format = "yyyy-mm-dd hh:mm"

    # Ddate, EarningDate = date only
    for name in ["Ddate", "EarningDate"]:
        idx = header_map.get(name)
        if idx:
            col_letter = get_column_letter(idx)
            for r in range(2, ws.max_row + 1):
                ws[f"{col_letter}{r}"].number_format = "yyyy-mm-dd"

    # Auto-fit columns (rough approximation)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for r in range(2, ws.max_row + 1):
            val = ws[f"{col_letter}{r}"].value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = max_len + 2

    RAW_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RAW_XLSX)


# --------------------------------------------------------------------
# Step 5: CLI / Orchestration
# --------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Live session → DataSource_Raw.xlsx")
    parser.add_argument(
        "--slow",
        action="store_true",
        help="Also fetch slow fields (Beta, 1YT, Ddate, EarningDate, RepDiv)",
    )
    args = parser.parse_args()

    mode = "SLOW" if args.slow else "FAST"
    start = time.time()
    print(f"🕒 START at {datetime.now().strftime('%H:%M:%S')} — MODE: {mode}")

    # 1) Read tickers
    t0 = time.time()
    tickers_raw = read_tickers()
    print(f"✔ Loaded {len(tickers_raw)} tickers from TickerList!{TICKER_TABLE_COL}")
    print(f"⏱ Tickers read in {time.time() - t0:.3f} sec")

    # 2) Fast fetch
    t1 = time.time()
    df_fast = fetch_fast(tickers_raw)
    print(f"⏱ Fast fetch in {time.time() - t1:.3f} sec")

    # 3) Slow fetch (optional)
    if args.slow:
        t2 = time.time()
        df_slow = fetch_slow(tickers_raw)

        # Merge on Ticker, preserving row order / length
        df = df_fast.copy()
        # start with NaNs for slow columns
        for c in SLOW_COLS:
            df[c] = None

        # Build mapping from ticker -> slow row
        slow_map = {row["Ticker"]: row for _, row in df_slow.iterrows() if row["Ticker"]}

        for i, ticker in enumerate(df["Ticker"].tolist()):
            if not ticker:
                continue
            srow = slow_map.get(ticker)
            if srow is None:
                continue
            for c in SLOW_COLS:
                df.at[i, c] = srow.get(c, None)

        print(f"⏱ Slow fields fetched in {time.time() - t2:.3f} sec")
    else:
        # Ensure slow columns exist (as empty) in FAST-only mode
        df = df_fast.copy()
        for c in SLOW_COLS:
            if c not in df.columns:
                df[c] = None

    # 4) Column order & write Excel
    df = df[ALL_COLS]  # enforce desired column order

    t3 = time.time()
    write_raw_excel(df)
    print(f"⏱ Excel written in {time.time() - t3:.3f} sec")

    total = time.time() - start
    print(f"✅ Finished. Total time: {total:.3f} sec")
    print(f"📂 Output: {RAW_XLSX}")

    # Auto-open on macOS
    open_excel_file(RAW_XLSX)


if __name__ == "__main__":
    main()