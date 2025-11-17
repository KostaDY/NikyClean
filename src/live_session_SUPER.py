from __future__ import annotations

import argparse
import platform
import re
import subprocess
import time
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from yahooquery import Ticker

# --------------------------------------------------------------------
# Paths & constants
# --------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "excel"
DATASOURCE_XLSX = EXCEL_DIR / "DataSource.xlsx"
RAW_XLSX = EXCEL_DIR / "DataSource_Raw.xlsx"

TICKERLIST_SHEET = "TickerList"
TICKERLIST_COLUMN = "Ticker"

RAW_SHEET = "RTData_Raw"
RAW_TABLE_NAME = "RTData_Raw"

# Final column order in DataSource_Raw
COLUMNS = [
    "Ticker",
    "RefreshTime",      # last trade time (FAST)
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "RepDiv",
]

# Columns filled by FAST mode
FAST_COLS = [
    "Ticker",
    "RefreshTime",
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
]

# Columns filled by SLOW mode
SLOW_COLS = [
    "Ticker",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "RepDiv",
]

HEADERS = {"User-Agent": "Mozilla/5.0"}


# --------------------------------------------------------------------
# Utilities
# --------------------------------------------------------------------

def now_str() -> str:
    return datetime.now().strftime("%H:%M:%S")


def open_excel_file(path: Path) -> None:
    """Open Excel workbook on macOS / Windows if it exists."""
    if not path.exists():
        print(f"⚠ Cannot open — file not found: {path}")
        return

    try:
        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", str(path)])
        elif system == "Windows":
            subprocess.Popen(["start", str(path)], shell=True)
        else:
            # On Linux just print the path
            print(f"📂 Saved at: {path}")
    except Exception as e:
        print(f"⚠ Failed to open Excel file: {e}")


def _extract_time(val: Any) -> Optional[datetime]:
    """
    Try to convert Yahoo 'regularMarketTime' style value to datetime.
    Handles dict(raw/fmt), epoch seconds, ISO strings, datetime.
    """
    if val is None:
        return None

    if isinstance(val, dict):
        val = val.get("raw") or val.get("fmt")

    if isinstance(val, datetime):
        return val

    if isinstance(val, (int, float)):
        try:
            return datetime.fromtimestamp(val)
        except Exception:
            return None

    if isinstance(val, str):
        # Try ISO-ish formats; if it fails, just give up
        try:
            return datetime.fromisoformat(val)
        except Exception:
            return None

    return None


def _parse_date_str(s: Optional[str]) -> Optional[date]:
    """
    Convert strings like 'Nov 10, 2025' or 'Jan 29, 2026 - Feb 1, 2026' to date.
    Returns None if not parseable.
    """
    if not s:
        return None
    s = s.strip()
    if s in ("--", "N/A"):
        return None

    # If a range "Jan 29, 2026 - Feb 1, 2026", take first part
    if " - " in s:
        s = s.split(" - ")[0].strip()

    try:
        return datetime.strptime(s, "%b %d, %Y").date()
    except Exception:
        return None


def _parse_float(s: Optional[str]) -> Optional[float]:
    """Parse a number possibly containing commas; return None if fail."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = s.replace(",", "").strip()
    if s in ("--", "N/A", ""):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _parse_percent_to_fraction(s: Optional[str]) -> Optional[float]:
    """
    Parse strings like '2.83%', '2.83% (0.20)' to 0.0283.
    Returns None if not parseable.
    """
    if not s:
        return None
    if isinstance(s, (int, float)):
        # assume already fraction (0.0283), not 2.83
        return float(s)
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", s)
    if not m:
        return None
    try:
        return float(m.group(1)) / 100.0
    except Exception:
        return None


# --------------------------------------------------------------------
# Step 1: Read tickers from DataSource.xlsx / TickerList!Ticker
# --------------------------------------------------------------------

def read_tickers() -> List[str]:
    """
    Reads TickerList!Ticker from DataSource.xlsx.
    Keeps row order and preserves empty rows as "".
    """
    df = pd.read_excel(DATASOURCE_XLSX, sheet_name=TICKERLIST_SHEET)
    if TICKERLIST_COLUMN not in df.columns:
        raise ValueError(
            f"❌ Column '{TICKERLIST_COLUMN}' not found in sheet '{TICKERLIST_SHEET}'."
        )

    tickers: List[str] = []
    for v in df[TICKERLIST_COLUMN]:
        if pd.isna(v):
            tickers.append("")
        else:
            tickers.append(str(v).strip())

    print(f"✔ Loaded {len(tickers)} tickers from {TICKERLIST_SHEET}!{TICKERLIST_COLUMN}")
    return tickers


# --------------------------------------------------------------------
# Step 2: FAST fetch with yahooquery
# --------------------------------------------------------------------

def fetch_fast(tickers: List[str]) -> pd.DataFrame:
    """
    FAST mode: prices, changes, volumes, and last trade time (RefreshTime).
    Uses yahooquery (price, summary_detail).
    """
    syms = [t for t in tickers if t]
    if not syms:
        return pd.DataFrame({col: [] for col in FAST_COLS})

    tq = Ticker(syms, asynchronous=True)

    raw_price = tq.price or {}
    raw_sd = tq.summary_detail or {}

    def get_dict(d: Any, key: str) -> Dict[str, Any]:
        v = d.get(key)
        return v if isinstance(v, dict) else {}

    price: Dict[str, Dict[str, Any]] = {s: get_dict(raw_price, s) for s in syms}
    sd: Dict[str, Dict[str, Any]] = {s: get_dict(raw_sd, s) for s in syms}

    rows_by_symbol: Dict[str, Dict[str, Any]] = {}

    for s in syms:
        p = price.get(s, {})
        d = sd.get(s, {})

        close = d.get("regularMarketPreviousClose") or p.get("regularMarketPreviousClose")
        open_ = d.get("regularMarketOpen") or p.get("regularMarketOpen")
        last = p.get("regularMarketPrice") or d.get("regularMarketPrice")
        low = d.get("regularMarketDayLow") or p.get("regularMarketDayLow")
        high = d.get("regularMarketDayHigh") or p.get("regularMarketDayHigh")
        pe = d.get("trailingPE") or p.get("trailingPE")

        change = d.get("regularMarketChange") or p.get("regularMarketChange")
        change_pct = d.get("regularMarketChangePercent") or p.get(
            "regularMarketChangePercent"
        )

        if change_pct is not None:
            try:
                # Sometimes it's already fraction; sometimes it's 2.34 (percent)
                if abs(change_pct) > 1:
                    change_pct = change_pct / 100.0
            except Exception:
                pass

        volume = d.get("regularMarketVolume") or p.get("regularMarketVolume")
        volavg = d.get("averageVolume") or p.get("averageVolume")

        # REAL last trade time (no pre/post) — regularMarketTime
        rt_raw = p.get("regularMarketTime") or d.get("regularMarketTime")
        rt_dt = _extract_time(rt_raw)

        rows_by_symbol[s] = {
            "Ticker": s,
            "RefreshTime": rt_dt,
            "Close": close,
            "Open": open_,
            "Last": last,
            "Low": low,
            "High": high,
            "P/E": pe,
            "Change": change,
            "ChangePct": change_pct,
            "Volume": volume,
            "VolumeAverage": volavg,
        }

    # Rebuild in original order, preserving empty rows
    empty_row = {col: None for col in FAST_COLS}
    result_rows: List[Dict[str, Any]] = []

    for t in tickers:
        if t and t in rows_by_symbol:
            result_rows.append(rows_by_symbol[t])
        else:
            r = empty_row.copy()
            r["Ticker"] = t if t else ""
            result_rows.append(r)

    df = pd.DataFrame(result_rows)
    # Ensure all FAST_COLS exist
    for col in FAST_COLS:
        if col not in df.columns:
            df[col] = None

    return df[FAST_COLS]


# --------------------------------------------------------------------
# Step 3: SLOW fetch via HTML scraping
# --------------------------------------------------------------------

def html_fetch_slow_fields(ticker: str) -> Tuple[Any, Any, Any, Any, Any]:
    """
    Scrape: Beta, 1y target price, ex-dividend date, earnings date, dividend yield.
    Returns raw strings (to be parsed later).
    """
    url = f"https://finance.yahoo.com/quote/{ticker}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.raise_for_status()
        html = resp.text
    except Exception:
        return None, None, None, None, None

    soup = BeautifulSoup(html, "lxml")

    def extract(label_pattern: str) -> Optional[str]:
        m = soup.find("span", string=re.compile(label_pattern, re.I))
        if not m:
            return None
        v = m.find_next("span")
        return v.text.strip() if v else None

    beta = extract(r"^Beta")                     # matches Beta, Beta (5Y Monthly), Beta (something)
    tgt = extract(r"1y Target Est")
    exdiv = extract(r"Ex-Dividend Date")
    earnings = extract(r"Earnings Date")
    divy = extract(r"^Dividend Yield")           # matches Dividend Yield, Dividend Yield (Annual)

    return beta, tgt, exdiv, earnings, divy


def html_fetch_slow_fields(ticker: str):
    """
    Rock-solid Yahoo scraper for slow fields:
    Beta, 1y target, ex-dividend, earnings date, dividend yield.
    Uses aria-label based extraction (current Yahoo layout).
    """

    url = f"https://finance.yahoo.com/quote/{ticker}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.raise_for_status()
        html = resp.text
    except Exception:
        return None, None, None, None, None

    soup = BeautifulSoup(html, "lxml")

    # Helper: extract numeric/float or keep string
    def get_attr(prefix: str):
        """Find value using aria-label prefix."""
        td = soup.select_one(f'td[aria-label^="{prefix}"]')
        if not td:
            return None
        return td.text.strip()

    # NEW — correct extraction
    beta  = get_attr("Beta")
    tgt   = get_attr("1y Target Est")
    dy    = get_attr("Dividend Yield")
    exdiv = get_attr("Ex-Dividend Date")
    earn  = get_attr("Earnings Date")

    # Normalization
    beta  = _parse_float(beta)
    tgt   = _parse_float(tgt)

    # Dates
    def parse_date(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%b %d, %Y").date()
        except:
            return None

    exdiv = parse_date(exdiv)
    earn  = parse_date(earn)

    # Dividend Yield is a percentage
    if dy and dy.endswith("%"):
        try:
            dy = float(dy[:-1]) / 100.0
        except:
            dy = None
    else:
        dy = None

    return beta, tgt, exdiv, earn, dy
# --------------------------------------------------------------------
# Step 4: Write Excel (with formatting + table)
# --------------------------------------------------------------------

def write_raw_excel(df: pd.DataFrame) -> None:
    """
    Creates DataSource_Raw.xlsx with a clean RTData_Raw sheet:
    - Header row
    - All rows from df
    - Excel table
    - Basic formatting (numbers, %, dates, thousands separators)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET

    # Header
    for col_idx, name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = c.font.copy(bold=True)

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    end_row = len(df) + 1
    end_col = len(df.columns)

    # Excel table
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ref = f"A1:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=RAW_TABLE_NAME, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Number formats
    header_to_col: Dict[str, int] = {
        ws.cell(row=1, column=col_idx).value: col_idx
        for col_idx in range(1, end_col + 1)
    }

    def apply_format(col_name: str, number_format: str) -> None:
        col_idx = header_to_col.get(col_name)
        if not col_idx:
            return
        for r in range(2, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None:
                cell.number_format = number_format

    # Dates & times
    apply_format("RefreshTime", "dd-mmm-yyyy hh:mm:ss")
    apply_format("Ddate", "dd-mmm-yyyy")
    apply_format("EarningDate", "dd-mmm-yyyy")

    # Prices / floats with 2 decimals
    for name in ["Close", "Open", "Last", "Low", "High", "P/E", "1YT", "Change"]:
        apply_format(name, "0.00")

    # Percentages with 2 decimals
    for name in ["ChangePct", "RepDiv"]:
        apply_format(name, "0.00%")

    # Volumes as integers with thousand separators
    for name in ["Volume", "VolumeAverage"]:
        apply_format(name, "#,##0")

    # Beta as plain number with 2 decimals (where present)
    apply_format("Beta", "0.00")

    wb.save(RAW_XLSX)


# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Live session: DataSource.xlsx → DataSource_Raw.xlsx"
    )
    parser.add_argument(
        "--slow",
        action="store_true",
        help="Also refresh Beta, 1YT, Ddate, EarningDate, RepDiv via HTML scraping",
    )
    args = parser.parse_args()

    mode = "SLOW" if args.slow else "FAST"
    print(f"🕒 START at {now_str()} — MODE: {mode}")

    t0 = time.time()
    tickers = read_tickers()
    t1 = time.time()
    print(f"⏱ Tickers read in {t1 - t0:.3f} sec")

    # FAST
    tf0 = time.time()
    df_fast = fetch_fast(tickers)
    tf1 = time.time()
    print(f"⏱ Fast fetch in {tf1 - tf0:.3f} sec")

    df = df_fast.copy()

    # SLOW (optional)
    if args.slow:
        ts0 = time.time()
        df_slow = fetch_slow(tickers)
        print("\nDEBUG -- SLOW df columns:", df_slow.columns.tolist())
        print("DEBUG -- first rows:\n", df_slow.head())
        ts1 = time.time()
        print(f"⏱ Slow fields fetched in {ts1 - ts0:.3f} sec")

        # Row-by-row alignment — *no grouping by ticker*, preserves duplicates
        for col in SLOW_COLS:
            if col == "Ticker":
                continue
            df[col] = df_slow[col].values

    # Ensure all columns exist
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None

    tw0 = time.time()
    write_raw_excel(df[COLUMNS])
    tw1 = time.time()
    print(f"⏱ Excel written in {tw1 - tw0:.3f} sec")

    total = time.time() - t0
    print(f"✅ Finished. Total time: {total:.3f} sec")
    print(f"📂 Output: {RAW_XLSX}")

    # Auto-open
    open_excel_file(RAW_XLSX)


if __name__ == "__main__":
    main()