#!/usr/bin/env python3
"""
Minimal, stable live Session:

- Reads tickers from DataSource.xlsx / DataRT / RTData / Ticker_Symbol
- Preserves row order and empty rows
- Fetches FAST fields via yahooquery:
    Ticker, RefreshTime, Close, Open, Last, Low, High,
    P/E, Change, ChangePct, Volume, VolumeAverage
- Writes clean DataSource_Raw.xlsx with table RTData_Raw
- SLOW fields (Beta, 1YT, Ddate, EarningDate, RepDiv) are left as None
  so you can fill them via Excel formulas if you wish.

This avoids all crumb/HTML/JSON hassles and focuses on a rock-solid FAST path.
"""

from __future__ import annotations

import argparse
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from yahooquery import Ticker

# --------------------------------------------------------------------
# Paths & constants
# --------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent  # .../NikyClean
EXCEL_DIR = ROOT / "excel"
DATASOURCE_XLSX = EXCEL_DIR / "DataSource.xlsx"
RAW_XLSX = EXCEL_DIR / "DataSource_Raw.xlsx"

# Input location
RT_SHEET = "DataRT"
RT_TABLE_NAME = "RTData"
RT_TICKER_COL_NAME = "Ticker_Symbol"

# Output sheet / table
RAW_SHEET = "RTData_Raw"
RAW_TABLE_NAME = "RTData_Raw"

# Column layout (FAST + placeholder SLOW)
COLUMNS = [
    "Ticker",
    "RefreshTime",
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "RepDiv",
]

FAST_COLS = [
    "Ticker",
    "RefreshTime",
    "Close",
    "Open",
    "Last",
    "Low",
    "High",
    "P/E",
    "Change",
    "ChangePct",
    "Volume",
    "VolumeAverage",
]

SLOW_COLS = [
    "Ticker",
    "Beta",
    "1YT",
    "Ddate",
    "EarningDate",
    "RepDiv",
]


# --------------------------------------------------------------------
# Utilities
# --------------------------------------------------------------------


def now_str() -> str:
    return datetime.now().strftime("%H:%M:%S")


# --------------------------------------------------------------------
# Step 1: Read tickers from DataSource.xlsx / DataRT / RTData
# --------------------------------------------------------------------


def read_tickers() -> List[str]:
    """
    Read ONLY the Ticker_Symbol column from Excel TABLE RTData in sheet DataRT.

    - Preserves row order.
    - Keeps empty rows (returns "" for them).
    """
    from openpyxl import load_workbook

    if not DATASOURCE_XLSX.exists():
        raise FileNotFoundError(f"Input workbook not found: {DATASOURCE_XLSX}")

    wb = load_workbook(DATASOURCE_XLSX, data_only=True)
    if RT_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{RT_SHEET}' not found in {DATASOURCE_XLSX}")

    ws = wb[RT_SHEET]

    # locate Excel table RTData
    table = None
    for t in ws._tables.values():
        if t.name == RT_TABLE_NAME:
            table = t
            break
    if table is None:
        raise ValueError(f"Excel table '{RT_TABLE_NAME}' not found in sheet '{RT_SHEET}'.")

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # header row
    headers = [
        ws.cell(min_row, c).value for c in range(min_col, max_col + 1)
    ]
    if RT_TICKER_COL_NAME not in headers:
        raise ValueError(
            f"Column '{RT_TICKER_COL_NAME}' not found in table '{RT_TABLE_NAME}'. "
            f"Headers: {headers}"
        )

    ticker_col = headers.index(RT_TICKER_COL_NAME) + min_col

    tickers: List[str] = []
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(r, ticker_col).value
        if v is None:
            tickers.append("")  # preserve empty row
        else:
            tickers.append(str(v).strip())

    return tickers


# --------------------------------------------------------------------
# Step 2: FAST fetch via yahooquery
# --------------------------------------------------------------------


def _get_dict(block: Any, key: str) -> Dict[str, Any]:
    if not isinstance(block, dict):
        return {}
    v = block.get(key)
    return v if isinstance(v, dict) else {}


def _first(d: Dict[str, Any], *keys: str) -> Optional[Any]:
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    return None


def fetch_fast(tickers: List[str]) -> pd.DataFrame:
    """
    Fetchs FAST fields for all non-empty tickers via yahooquery:
    Ticker, RefreshTime, Close, Open, Last, Low, High,
    P/E, Change, ChangePct, Volume, VolumeAverage.

    - Uses yahooquery's internal crumb/session logic.
    - No 'retry' or 'backoff_factor' args (they break with curl_cffi).
    """
    syms = [t for t in tickers if t]
    if not syms:
        return pd.DataFrame(columns=FAST_COLS)

    # yahooquery Ticker – keep args minimal to avoid curl_cffi issues
    tq = Ticker(syms, asynchronous=True)

    raw_price = tq.price or {}
    raw_sd = tq.summary_detail or {}

    # Normalize to {symbol: dict}
    price: Dict[str, Dict[str, Any]] = {
        s: _get_dict(raw_price, s) for s in syms
    }
    sd: Dict[str, Dict[str, Any]] = {
        s: _get_dict(raw_sd, s) for s in syms
    }

    rows_by_symbol: Dict[str, Dict[str, Any]] = {}

    for s in syms:
        p = price.get(s, {}) or {}
        d = sd.get(s, {}) or {}

        close = _first(d, "regularMarketPreviousClose", "previousClose") or \
            _first(p, "regularMarketPreviousClose", "previousClose")
        open_ = _first(d, "regularMarketOpen") or _first(p, "regularMarketOpen")
        last = _first(p, "regularMarketPrice", "postMarketPrice", "preMarketPrice") or \
            _first(d, "regularMarketPrice")
        low = _first(d, "regularMarketDayLow") or _first(p, "regularMarketDayLow")
        high = _first(d, "regularMarketDayHigh") or _first(p, "regularMarketDayHigh")
        pe = _first(d, "trailingPE") or _first(p, "trailingPE")

        change = _first(p, "regularMarketChange", "postMarketChange", "preMarketChange") or \
            _first(d, "regularMarketChange")
        change_pct = _first(
            p,
            "regularMarketChangePercent",
            "postMarketChangePercent",
            "preMarketChangePercent",
        ) or _first(d, "regularMarketChangePercent")

        if change_pct is not None:
            try:
                # yahooquery sometimes returns percent as 1.23 instead of 0.0123
                if abs(change_pct) > 1:
                    change_pct = change_pct / 100.0
            except Exception:
                pass

        volume = _first(p, "regularMarketVolume") or _first(d, "regularMarketVolume")
        volavg = _first(d, "averageVolume", "averageDailyVolume10Day") or \
            _first(p, "averageVolume")

        # last trade timestamp: prefer regular session
        rt_ts = _first(
            p,
            "regularMarketTime",
            "postMarketTime",
            "preMarketTime",
        ) or _first(d, "regularMarketTime")

        if isinstance(rt_ts, (int, float)):
            try:
                refresh_time = datetime.fromtimestamp(rt_ts)
            except Exception:
                refresh_time = None
        else:
            refresh_time = None

        rows_by_symbol[s] = {
            "Ticker": s,
            "RefreshTime": refresh_time,
            "Close": close,
            "Open": open_,
            "Last": last,
            "Low": low,
            "High": high,
            "P/E": pe,
            "Change": change,
            "ChangePct": change_pct,
            "Volume": volume,
            "VolumeAverage": volavg,
        }

    # Row set for tickers with no data but non-empty symbol
    empty_fast_template = {col: None for col in FAST_COLS}

    result_rows: List[Dict[str, Any]] = []
    for t in tickers:
        if not t:
            row = empty_fast_template.copy()
            row["Ticker"] = ""
            result_rows.append(row)
        else:
            row = rows_by_symbol.get(t)
            if row is None:
                row = empty_fast_template.copy()
                row["Ticker"] = t
            result_rows.append(row)

    df = pd.DataFrame(result_rows, columns=FAST_COLS)
    return df


# --------------------------------------------------------------------
# Step 3: SLOW fields – stub (left None on purpose)
# --------------------------------------------------------------------


def fetch_slow_stub(tickers: List[str]) -> pd.DataFrame:
    """
    Stub for SLOW fields.

    Given the current external blocking from Yahoo JSON and HTML,
    this returns an all-None DataFrame with the correct shape,
    so you can:
      - keep the column structure,
      - fill Beta / 1YT / Ddate / EarningDate / RepDiv via Excel formulas.

    Each row corresponds 1:1 to 'tickers'.
    """
    data = {
        "Ticker": [],
        "Beta": [],
        "1YT": [],
        "Ddate": [],
        "EarningDate": [],
        "RepDiv": [],
    }
    for t in tickers:
        data["Ticker"].append(t)
        data["Beta"].append(None)
        data["1YT"].append(None)
        data["Ddate"].append(None)
        data["EarningDate"].append(None)
        data["RepDiv"].append(None)

    return pd.DataFrame(data, columns=SLOW_COLS)


# --------------------------------------------------------------------
# Step 4: Write Excel with formatting
# --------------------------------------------------------------------


def write_raw_excel(df: pd.DataFrame) -> None:
    """
    Creates DataSource_Raw.xlsx with a clean RTData_Raw sheet + table.

    - Applies number formats:
        * Floats: 2 decimals
        * Percent: 2 decimals
        * Integers (volumes): 1,000 separator
        * RefreshTime: yyyy-mm-dd hh:mm:ss
    """
    wb = Workbook()
    ws = wb.active
    ws.title = RAW_SHEET

    # Write header
    for col, name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col, value=name)
        # openpyxl deprecation warning is harmless here
        c.font = c.font.copy(bold=True)

    # Write data
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply number formats by column name
    col_index_by_name = {name: idx + 1 for idx, name in enumerate(df.columns)}

    def set_col_format(col_name: str, fmt: str):
        idx = col_index_by_name.get(col_name)
        if not idx:
            return
        for r in range(2, len(df) + 2):
            ws.cell(row=r, column=idx).number_format = fmt

    # RefreshTime as datetime
    set_col_format("RefreshTime", "yyyy-mm-dd hh:mm:ss")

    # Floats: 2 decimals
    for float_col in ["Close", "Open", "Last", "Low", "High", "P/E", "Beta", "1YT"]:
        set_col_format(float_col, "0.00")

    # Percentages: 2 decimals (ChangePct & RepDiv stored as fraction)
    set_col_format("ChangePct", "0.00%")
    set_col_format("RepDiv", "0.00%")

    # Integers with thousand separator
    for int_col in ["Volume", "VolumeAverage"]:
        set_col_format(int_col, "#,##0")

    # Create Excel table
    from openpyxl.worksheet.table import Table, TableStyleInfo

    end_row = len(df) + 1
    end_col = len(df.columns)
    ref = f"A1:{get_column_letter(end_col)}{end_row}"

    table = Table(displayName=RAW_TABLE_NAME, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(RAW_XLSX)


def open_excel_file(path: Path) -> None:
    """
    On macOS, open the generated workbook in Excel using 'open'.
    Safe: only acts if the file exists.
    """
    import subprocess

    if not path.exists():
        print(f"⚠ Cannot open — file not found: {path}")
        return
    try:
        subprocess.Popen(["open", str(path)])
    except Exception as e:
        print(f"⚠ Failed to open Excel file: {e}")


# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Minimal FAST DataSource → DataSource_Raw pipeline"
    )
    parser.add_argument(
        "--slow",
        action="store_true",
        help="(Currently stubbed) also include placeholder columns for Beta, 1YT, Ddate, EarningDate, RepDiv.",
    )
    args = parser.parse_args()

    mode = "SLOW" if args.slow else "FAST"
    print(f"🕒 START at {now_str()} — MODE: {mode}")

    t0 = time.time()
    tickers = read_tickers()
    print(f"✔ Loaded {len(tickers)} tickers from {RT_TABLE_NAME}!{RT_TICKER_COL_NAME}")
    t1 = time.time()
    print(f"⏱ Tickers read in {t1 - t0:.3f} sec")

    # FAST always
    tf0 = time.time()
    df_fast = fetch_fast(tickers)
    tf1 = time.time()
    print(f"⏱ Fast fetch in {tf1 - tf0:.3f} sec")

    # Start from fast frame
    df = df_fast.copy()

    if args.slow:
        # Currently: stub – keeps correct columns & alignment, values None
        ts0 = time.time()
        df_slow = fetch_slow_stub(tickers)
        print("DEBUG -- SLOW df columns:", df_slow.columns.tolist())
        print("DEBUG -- first rows:\n", df_slow.head())
        ts1 = time.time()
        print(f"⏱ Slow stub built in {ts1 - ts0:.3f} sec")

        # merge slow columns by position
        for col in SLOW_COLS:
            if col == "Ticker":
                continue
            df[col] = df_slow[col]

    # Ensure all defined columns exist
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = None

    tw0 = time.time()
    write_raw_excel(df[COLUMNS])
    tw1 = time.time()
    print(f"⏱ Excel written in {tw1 - tw0:.3f} sec")

    total = time.time() - t0
    print(f"✅ Finished. Total time: {total:.3f} sec")
    print(f"📂 Output: {RAW_XLSX}")

    # Auto-open workbook on macOS
    open_excel_file(RAW_XLSX)


if __name__ == "__main__":
    main()