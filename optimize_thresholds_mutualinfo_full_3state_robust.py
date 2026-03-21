#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Shuffle-corrected Mutual Information (full sample) for K=3, ORDER=3.

Digits (3 states):
1 : pct <= -a
2 : -a < pct < +a
3 : pct >= +a

This script:
- Reads tickers from markov.xlsx named range TICKERS (fallback M_Config!A2:A...)
- Downloads 900d daily closes (auto_adjust=True)
- Builds 400-digit strings
- Computes MI_real = I(S_t ; X_{t+1}) for ORDER=3
- Computes MI_bias via shuffling the digit string N_SHUFFLES times
- Reports MI_corr = MI_real - mean(MI_shuffle)
- Reports Z = MI_corr / std(MI_shuffle)
- Never crashes; reports skipped tickers
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, UTC
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"
N_DIGITS = 400
ORDER = 3
K = 3

# Use your best K=3 threshold from optimization
A = 1.0  # percent

N_SHUFFLES = 50  # increase to 100+ for smoother estimates

# ============================================================
# TICKERS
# ============================================================

def read_tickers() -> list[str]:
    wb = load_workbook(XLSX, data_only=True)
    tickers: list[str] = []

    dn = None
    try:
        dn = wb.defined_names.get(TICKERS_NAME)
    except Exception:
        dn = None

    if dn is not None:
        try:
            sheet_name, cell_range = next(dn.destinations)
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s:
                        tickers.append(s)
        except Exception as e:
            print("Named range read error:", e)

    if not tickers:
        if CFG_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{CFG_SHEET}' not found in {XLSX}")
        print("Falling back to column A in M_Config")
        ws = wb[CFG_SHEET]
        r = 2
        while True:
            v = ws[f"A{r}"].value
            if v is None or (isinstance(v, str) and not v.strip()):
                break
            tickers.append(str(v).strip())
            r += 1

    # de-dupe preserve order
    seen = set()
    out = []
    for t in tickers:
        if t and t not in seen:
            out.append(t)
            seen.add(t)
    return out

# ============================================================
# DATA
# ============================================================

def download_closes(ticker: str) -> pd.Series | None:
    try:
        df = yf.download(
            ticker,
            period=PERIOD,
            interval="1d",
            auto_adjust=True,
            progress=False,
            threads=True
        )
        if df is None or df.empty:
            return None
        s = df["Close"].dropna()
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:, 0].dropna()
        if s.empty:
            return None
        return s
    except Exception:
        return None

# ============================================================
# QUANTIZATION (K=3)
# ============================================================

def quantize_3(pct: float) -> int:
    if pct <= -A:
        return 1
    if pct >= A:
        return 3
    return 2

def build_digits(closes: pd.Series) -> str | None:
    closes = closes.dropna()
    if len(closes) < N_DIGITS + 1:
        return None
    closes = closes.tail(N_DIGITS + 1)
    pct = closes.pct_change().iloc[1:] * 100.0
    if len(pct) != N_DIGITS:
        return None
    digits_chrono = [str(quantize_3(float(x))) for x in pct.tolist()]  # oldest->youngest
    return "".join(digits_chrono[::-1])  # youngest-first

# ============================================================
# INFORMATION THEORY
# ============================================================

def entropy_from_counts(counts: np.ndarray) -> float:
    tot = float(counts.sum())
    if tot <= 0:
        return 0.0
    p = counts / tot
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())

def mutual_info_fullsample(digits_youngest_first: str) -> tuple[float, float]:
    """
    Computes MI_real = I(S_t ; X_{t+1}) in bits (plug-in estimate).
    Also returns H(X) for context.
    """
    chrono = digits_youngest_first[::-1]  # oldest->youngest
    n = len(chrono)
    if n <= ORDER:
        return 0.0, 0.0

    state_to_i: dict[str, int] = {}
    joint_rows: list[np.ndarray] = []  # each row length K
    next_counts = np.zeros(K, dtype=float)

    def get_state_index(s: str) -> int:
        if s in state_to_i:
            return state_to_i[s]
        state_to_i[s] = len(joint_rows)
        joint_rows.append(np.zeros(K, dtype=float))
        return state_to_i[s]

    for t in range(ORDER, n):
        s = chrono[t-ORDER:t]
        x = int(chrono[t]) - 1
        i = get_state_index(s)
        joint_rows[i][x] += 1.0
        next_counts[x] += 1.0

    Hx = entropy_from_counts(next_counts)

    total = float(next_counts.sum())
    HxS = 0.0
    for row in joint_rows:
        rs = float(row.sum())
        if rs <= 0:
            continue
        HxS += (rs / total) * entropy_from_counts(row)

    MI = Hx - HxS
    return float(MI), float(Hx)

# ============================================================
# MAIN
# ============================================================

def main():
    tickers = read_tickers()
    print(f"Tickers listed: {len(tickers)}")

    used = 0
    skipped = 0

    port_real = 0.0
    port_bias = 0.0
    port_corr = 0.0
    port_z = 0.0

    rng = np.random.default_rng()

    for t in tickers:
        closes = download_closes(t)
        if closes is None:
            print(f"{t:8s}  SKIP  download/empty")
            skipped += 1
            continue

        digits = build_digits(closes)
        if digits is None or len(digits) != N_DIGITS:
            print(f"{t:8s}  SKIP  insufficient history for {N_DIGITS} digits")
            skipped += 1
            continue

        MI_real, Hx = mutual_info_fullsample(digits)

        # Shuffle baseline
        arr = np.array(list(digits), dtype="<U1")
        sh = np.empty(N_SHUFFLES, dtype=float)
        for i in range(N_SHUFFLES):
            rng.shuffle(arr)
            MI_s, _ = mutual_info_fullsample("".join(arr))
            sh[i] = MI_s

        bias = float(sh.mean())
        std = float(sh.std(ddof=0))
        corr = MI_real - bias
        z = corr / std if std > 0 else 0.0

        print(
            f"{t:8s}  "
            f"MI_real={MI_real:.4f}  "
            f"Bias={bias:.4f}  "
            f"MI_corr={corr:.4f}  "
            f"Z={z:.2f}  "
            f"Hx={Hx:.3f}"
        )

        port_real += MI_real
        port_bias += bias
        port_corr += corr
        port_z += z
        used += 1

    print("\n================ PORTFOLIO ================")
    print(f"Tickers used:     {used}")
    print(f"Tickers skipped:  {skipped}")

    if used == 0:
        print("No valid tickers.")
        return

    print(f"Avg MI_real  : {port_real/used:.4f} bits")
    print(f"Avg Bias     : {port_bias/used:.4f} bits")
    print(f"Avg MI_corr  : {port_corr/used:.4f} bits")
    print(f"Avg Z-score  : {port_z/used:.2f}")
    print(f"GeneratedUTC : {datetime.now(UTC).isoformat()}")

if __name__ == "__main__":
    main()