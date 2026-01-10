import os
import time
import pandas as pd
import pytz
import traceback
import yfinance as yf
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess

# ============================================================
# SETTINGS
# ============================================================
WORKBOOK = "ASXData.xlsx"
INPUT_SHEET = "Tickers"
OUTPUT_SHEET = "LatestData"

HISTORY_PERIOD = "7d"
HISTORY_INTERVAL = "1d"
AVG_VOLUME_DAYS = 7

INTRADAY_PERIOD = "1d"
INTRADAY_INTERVAL = "1m"

SKIP_FULL_INFO_TICKERS = {"LOCK.L"}

YAHOO_DELAY_SEC = 0.3   # 300 ms abuse protection

# ============================================================
# EXCEL CONTROL (macOS)
# ============================================================
def close_excel_workbook(path):
    try:
        subprocess.run(
            [
                "osascript",
                "-e",
                f'''
                tell application "Microsoft Excel"
                    if exists workbook "{os.path.basename(path)}" then
                        close workbook "{os.path.basename(path)}" saving no
                    end if
                end tell
                '''
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
    except Exception:
        pass

def open_excel_workbook(path):
    try:
        subprocess.run(
            ["open", path],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
    except Exception:
        pass

# ============================================================
# VERIFY FILE
# ============================================================
if not os.path.exists(WORKBOOK):
    raise FileNotFoundError(f"Workbook '{WORKBOOK}' not found")

close_excel_workbook(WORKBOOK)

# ============================================================
# READ TICKERS (HARD NaN FILTER)
# ============================================================
tickers_df = pd.read_excel(WORKBOOK, sheet_name=INPUT_SHEET)
col_name = "Ticker" if "Ticker" in tickers_df.columns else "Tickers"

tickers = (
    tickers_df[col_name]
    .dropna()
    .astype(str)
    .str.strip()
    .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA})
    .dropna()
    .tolist()
)

if not tickers:
    raise ValueError("No tickers found")

# ============================================================
# FETCH DATA (ONE ROW PER TICKER)
# ============================================================
rows = []

for ticker in tickers:
    time.sleep(YAHOO_DELAY_SEC)

    row = {
        "Ticker": ticker,
        "Status": "OK",
        "MissingFields": ""
    }

    try:
        tk = yf.Ticker(ticker)

        info = tk.fast_info if hasattr(tk, "fast_info") else {}
        full = {}

        if ticker not in SKIP_FULL_INFO_TICKERS:
            try:
                full = tk.info or {}
            except Exception:
                full = {}

        # --- Market prices ---
        row["Open"] = (
            info.get("open")
            or full.get("regularMarketOpen")
            or full.get("open")
            or ""
        )

        row["MarketPrice"] = info.get("lastPrice", "")
        row["MarketDayHigh"] = info.get("dayHigh", "")
        row["MarketDayLow"] = info.get("dayLow", "")
        row["MarketPreviousClose"] = info.get("previousClose", "")
        row["Volume"] = info.get("volume", "")

        # IMPORTANT: MarketTime is intraday ONLY
        row["MarketTime"] = ""

        row["AverageVolume"] = ""

        # --- Valuation / fundamentals ---
        row["marketCap"] = full.get("marketCap", "")
        row["fiftyTwoWeekHigh"] = full.get("fiftyTwoWeekHigh", "")
        row["fiftyTwoWeekLow"] = full.get("fiftyTwoWeekLow", "")
        row["currency"] = full.get("currency", "")

        # --- Analyst data ---
        row["targetMeanPrice"] = full.get("targetMeanPrice", "")
        row["recommendationMean"] = full.get("recommendationMean", "")

        # --- Dividends ---
        row["dividendRate"] = full.get("dividendRate", "")
        row["dividendYield"] = full.get("dividendYield", "")
        row["exDividendDate"] = full.get("exDividendDate", "")

        # --- Earnings ---
        row["earningsDate"] = ""
        cal = tk.calendar

        if isinstance(cal, pd.DataFrame) and not cal.empty:
            if "Earnings Date" in cal.index:
                dates = cal.loc["Earnings Date"].tolist()
            elif "Earnings Date" in cal.columns:
                dates = cal["Earnings Date"].tolist()
            else:
                dates = None

            if dates:
                first = dates[0]
                if isinstance(first, (list, tuple)) and first:
                    first = first[0]
                row["earningsDate"] = first

        elif isinstance(cal, dict) and "Earnings Date" in cal:
            d = cal.get("Earnings Date")
            if isinstance(d, (list, tuple)) and d:
                row["earningsDate"] = d[0]
            else:
                row["earningsDate"] = d

        # --- Daily history (NO MarketTime here) ---
        try:
            hist = tk.history(period=HISTORY_PERIOD, interval=HISTORY_INTERVAL)
            if isinstance(hist, pd.DataFrame) and not hist.empty:
                if not row["Open"] and "Open" in hist:
                    opens = hist["Open"].dropna()
                    if not opens.empty:
                        row["Open"] = opens.iloc[-1]

                if "Volume" in hist:
                    vol = hist["Volume"].dropna()
                    if not vol.empty:
                        row["Volume"] = row["Volume"] or vol.iloc[-1]
                        row["AverageVolume"] = vol.tail(AVG_VOLUME_DAYS).mean()
        except Exception:
            pass

        if not row["AverageVolume"]:
            row["AverageVolume"] = (
                full.get("averageVolume")
                or full.get("averageVolume10days")
                or ""
            )

        # --- INTRADAY MarketTime ONLY (naive datetime) ---
        try:
            intraday = tk.history(
                period=INTRADAY_PERIOD,
                interval=INTRADAY_INTERVAL
            )
            if isinstance(intraday, pd.DataFrame) and not intraday.empty:
                ts = intraday.index[-1]
                if hasattr(ts, "tzinfo") and ts.tzinfo is not None:
                    ts = ts.tz_localize(None)
                row["MarketTime"] = ts
        except Exception:
            pass

        # --- Diagnostics ---
        missing = [k for k, v in row.items() if v in ("", None)]
        if len(missing) > 3:
            row["Status"] = "PARTIAL"
            row["MissingFields"] = ",".join(missing)

    except Exception as e:
        row["Status"] = "NO_DATA"
        row["MissingFields"] = str(e)

    rows.append(row)

# ============================================================
# DATAFRAME
# ============================================================
df = pd.DataFrame(rows)

# ============================================================
# DATE NORMALIZATION (NO MarketTime TOUCH)
# ============================================================
def normalize_date(x):
    try:
        ts = pd.to_datetime(x, utc=True, errors="coerce")
        return "" if pd.isna(ts) else ts.strftime("%d-%b-%Y")
    except Exception:
        return ""

for col in ("exDividendDate", "earningsDate"):
    if col in df.columns:
        df[col] = df[col].apply(normalize_date)

df = df.fillna("")
df.columns = df.columns.astype(str)

# ============================================================
# EXCEL TABLE HELPERS
# ============================================================
def add_or_replace_table(ws, name):
    try:
        ws.tables.pop(name, None)
    except Exception:
        pass

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )
    ws.add_table(table)

# ============================================================
# WRITE TO EXCEL
# ============================================================
with pd.ExcelWriter(
    WORKBOOK,
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    df.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
    ws = writer.sheets[OUTPUT_SHEET]
    add_or_replace_table(ws, "RTdata")

open_excel_workbook(WORKBOOK)