#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations
from pathlib import Path
from datetime import datetime, UTC
import math
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

# ============================================================
# SETTINGS
# ============================================================

XLSX = Path("markov.xlsx")
CFG_SHEET = "M_Config"
TICKERS_NAME = "TICKERS"

PERIOD = "900d"      # enough to get 401 closes reliably
N_DIGITS = 400

K = 5               # 5 states
ORDER = 3           # set 2 or 3 (MI tends to rise with order, but beware sparsity)

# Symmetric threshold grid: (-a, -c, +c, +a), a > c > 0
A_GRID = np.arange(1.0, 4.1, 0.25)
C_GRID = np.arange(0.25, 2.1, 0.25)

# Optional: minimum entropy constraint to avoid degenerate distributions
# If H(X) < MIN_HX, the digitization is too imbalanced → skip.
MIN_HX = 1.0  # bits (tweak or set to 0.0 to disable)

# ============================================================
# EXCEL TICKERS
# ============================================================

def read_tickers_from_excel(
    xlsx_path: Path,
    cfg_sheet: str,
    named_range: str,
    fallback_col: str = "A",
    fallback_start_row: int = 2,
    fallback_max_rows: int = 2000,
) -> list[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    tickers: list[str] = []

    dn = None
    try:
        dn = wb.defined_names.get(named_range)
    except Exception:
        dn = None

    if dn is not None:
        try:
            dests = list(dn.destinations)
            if dests:
                sheet_name, cell_range = dests[0]
                ws = wb[sheet_name]
                for row in ws[cell_range]:
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.strip():
                            tickers.append(v.strip())
        except Exception:
            tickers = []

    if not tickers:
        if cfg_sheet not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{cfg_sheet}' not found in {xlsx_path}")
        ws = wb[cfg_sheet]
        col_idx = ord(fallback_col.upper()) - ord("A") + 1
        for r in range(fallback_start_row, fallback_start_row + fallback_max_rows):
            v = ws.cell(r, col_idx).value
            if v is None or (isinstance(v, str) and not v.strip()):
                break
            tickers.append(str(v).strip())

    seen = set()
    out = []
    for t in tickers:
        if t and t not in seen:
            out.append(t)
            seen.add(t)
    if not out:
        raise RuntimeError("No tickers found.")
    return out

# ============================================================
# DATA + DIGITS
# ============================================================

def download_closes(ticker: str, period: str) -> pd.Series:
    df = yf.download(ticker, period=period, interval="1d",
                     auto_adjust=True, progress=False, threads=True)
    if df is None or df.empty:
        raise RuntimeError(f"{ticker}: empty download")
    s = df["Close"].dropna()
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0].dropna()
    if s.empty:
        raise RuntimeError(f"{ticker}: no Close")
    return s

def quantize_5_fixed(pct: float, b1: float, b2: float, b3: float, b4: float) -> int:
    # returns integer state 1..5
    if pct <= b1: return 1
    if pct <= b2: return 2
    if pct <  b3: return 3
    if pct <  b4: return 4
    return 5

def build_digits(closes: pd.Series, thresholds: tuple[float,float,float,float], n_digits: int) -> str:
    need = n_digits + 1
    closes = closes.dropna()
    if len(closes) < need:
        raise RuntimeError(f"Need {need} closes, got {len(closes)}")
    closes = closes.tail(need)

    pct = closes.pct_change().iloc[1:] * 100.0  # length n_digits
    b1,b2,b3,b4 = thresholds
    digits_chrono = [str(quantize_5_fixed(float(x), b1,b2,b3,b4)) for x in pct.tolist()]
    return "".join(digits_chrono[::-1])  # youngest-first

# ============================================================
# INFORMATION THEORY
# ============================================================

def entropy_from_counts(counts: np.ndarray) -> float:
    tot = counts.sum()
    if tot <= 0:
        return 0.0
    p = counts / tot
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())

def mutual_information_fullsample(digits_youngest_first: str, order: int, K: int):
    """
    Computes I(S_t ; X_{t+1}) in bits, full-sample, where:
      S_t = last 'order' digits (history state)
      X_{t+1} = next digit (1..K)

    Returns:
      MI, Hx, Hx_given_s, n_obs
    """
    chrono = digits_youngest_first[::-1]  # oldest->youngest
    n = len(chrono)

    if n <= order:
        return 0.0, 0.0, 0.0, 0

    # Map history state strings to row index
    # Build joint counts for (state, next_digit)
    state_to_i = {}
    joint = []  # list of np arrays length K
    next_counts = np.zeros(K, dtype=float)

    def get_state_index(s: str) -> int:
        if s in state_to_i:
            return state_to_i[s]
        state_to_i[s] = len(joint)
        joint.append(np.zeros(K, dtype=float))
        return state_to_i[s]

    obs = 0
    for t in range(order, n):
        s = chrono[t-order:t]
        x = int(chrono[t]) - 1  # 0..K-1
        i = get_state_index(s)
        joint[i][x] += 1.0
        next_counts[x] += 1.0
        obs += 1

    # H(X)
    Hx = entropy_from_counts(next_counts)

    # H(X|S) = sum_s p(s) H(p(x|s))
    total = float(obs)
    HxS = 0.0
    for row in joint:
        row_sum = row.sum()
        if row_sum <= 0:
            continue
        HxS += (row_sum / total) * entropy_from_counts(row)

    MI = Hx - HxS
    return float(MI), float(Hx), float(HxS), obs

# ============================================================
# PORTFOLIO OBJECTIVE
# ============================================================

def portfolio_objective_MI(thresholds, tickers, closes_map):
    MI_sum = 0.0
    Hx_sum = 0.0
    HxS_sum = 0.0
    used = 0

    for t in tickers:
        try:
            digits = build_digits(closes_map[t], thresholds, N_DIGITS)
            MI, Hx, HxS, obs = mutual_information_fullsample(digits, ORDER, K)
            # Optional constraint to avoid degenerate digitizations
            if Hx < MIN_HX:
                continue
            MI_sum += MI
            Hx_sum += Hx
            HxS_sum += HxS
            used += 1
        except Exception:
            continue

    if used == 0:
        return -1e9, 0.0, 0.0, 0.0, 0

    return MI_sum, (Hx_sum/used), (HxS_sum/used), (MI_sum/used), used

# ============================================================
# MAIN: GRID SEARCH
# ============================================================

def main():
    tickers = read_tickers_from_excel(XLSX, CFG_SHEET, TICKERS_NAME)
    print(f"Tickers listed: {len(tickers)}")

    closes_map = {}
    for t in tickers:
        try:
            closes_map[t] = download_closes(t, PERIOD)
        except Exception as e:
            print(f"Skip {t}: {e}")

    good = [t for t in tickers if t in closes_map]
    print(f"Tickers with data: {len(good)}")

    best = (-1e9, None, None)

    for a in A_GRID:
        for c in C_GRID:
            if c >= a:
                continue
            thr = (-float(a), -float(c), float(c), float(a))

            MI_sum, Hx_avg, HxS_avg, MI_avg, used = portfolio_objective_MI(thr, good, closes_map)
            if MI_sum > best[0]:
                best = (MI_sum, thr, (Hx_avg, HxS_avg, MI_avg, used))
                print(
                    f"NEW BEST MI_sum={MI_sum:.6f}  thr={thr}  "
                    f"used={used}  Hx_avg={Hx_avg:.3f}  Hx|S_avg={HxS_avg:.3f}  MI_avg={MI_avg:.3f}"
                )

    print("\nDONE")
    if best[1] is None:
        print("No solution met the MIN_HX constraint; try lowering MIN_HX.")
        return

    MI_sum, thr, stats = best
    Hx_avg, HxS_avg, MI_avg, used = stats
    print(f"Best thresholds: {thr}")
    print(f"MI_sum (portfolio): {MI_sum:.6f} bits")
    print(f"Avg H(X):          {Hx_avg:.3f} bits (max for K=5 is {math.log(5,2):.3f})")
    print(f"Avg H(X|S):        {HxS_avg:.3f} bits")
    print(f"Avg MI per ticker: {MI_avg:.3f} bits")
    print(f"Tickers used:      {used}")
    print(f"GeneratedUTC:      {datetime.now(UTC).isoformat()}")

if __name__ == "__main__":
    main()