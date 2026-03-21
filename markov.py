#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import math
import subprocess
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ============================================================
# SETTINGS
# ============================================================

TICKER = "AMZN"
DAYS = 200

OUT_STOCK = Path("stock_history.csv")
OUT_DIGITS = Path("digit_string.txt")
OUT_MARKOV_CSV = Path("markov.csv")
OUT_MARKOV_XLSX = Path("markov.xlsx")
SHEET = "M_Result"

# ============================================================
# QUANTIZATION (3 STATES)
# ============================================================

def quantize_3(pct: float) -> str:
    if math.isnan(pct):
        return ""
    if pct <= -1.0:
        return "1"
    if pct >= 1.0:
        return "3"
    return "2"

# ============================================================
# FETCH DATA & BUILD DIGIT STRING
# ============================================================

def fetch_history():
    df = yf.download(
        TICKER,
        period="400d",
        interval="1d",
        progress=False
    )

    df = df.tail(DAYS).copy()
    df.reset_index(inplace=True)

    df["PctChange"] = df["Close"].pct_change() * 100.0
    df["Digit"] = df["PctChange"].apply(quantize_3)

    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")
    df.to_csv(OUT_STOCK, index=False)

    digits = df.loc[df["Digit"] != "", "Digit"].tolist()
    if len(digits) != 199:
        raise RuntimeError("Expected 199 digit values")

    # youngest first (for future incremental updates)
    s = "".join(digits[::-1])
    OUT_DIGITS.write_text(s)

    return s

# ============================================================
# MARKOV CORE
# ============================================================

def markov_chain(digits_youngest_first: str, order: int):
    chrono = digits_youngest_first[::-1]  # oldest → youngest

    transitions = []
    for i in range(len(chrono) - order):
        a = chrono[i:i + order]
        b = chrono[i + 1:i + order + 1]
        transitions.append((a, b))

    states = sorted(set(s for t in transitions for s in t))
    idx = {s: i for i, s in enumerate(states)}

    n = len(states)
    C = np.zeros((n, n))

    for a, b in transitions:
        C[idx[a], idx[b]] += 1

    P = np.zeros_like(C)
    for i in range(n):
        if C[i].sum() > 0:
            P[i] = C[i] / C[i].sum()
        else:
            P[i, i] = 1.0

    # Stationary distribution
    pi = np.ones(n) / n
    for _ in range(5000):
        pi_new = pi @ P
        if np.linalg.norm(pi_new - pi, 1) < 1e-12:
            break
        pi = pi_new
    pi /= pi.sum()

    # Prediction
    current = chrono[-order:]
    probs = {"1": 0.0, "2": 0.0, "3": 0.0}

    if current in idx:
        row = P[idx[current]]
        for j, s in enumerate(states):
            probs[s[-1]] += row[j]
    else:
        for j, s in enumerate(states):
            probs[s[-1]] += pi[j]

    return states, P, pi, current, probs

# ============================================================
# SAVE RESULTS
# ============================================================

def save_results(blocks):
    # ---------------- CSV ----------------
    rows = [f"GeneratedUTC,{datetime.utcnow().isoformat()}Z"]

    for b in blocks:
        states = b["states"]
        P = b["P"]

        rows += [
            "",
            f"[ORDER_{b['order']}]",
            f"CurrentState,{b['current']}",
            f"Prediction,{max(b['probs'], key=b['probs'].get)}",
            "NextDigitProbs,1,2,3",
            "," + ",".join(f"{b['probs'][d]:.6f}" for d in "123"),
            "",
            "StationaryVector",
            "State,Pi"
        ]

        for s, p in zip(states, b["pi"]):
            rows.append(f"{s},{p:.10f}")

        rows += [
            "",
            "TransitionMatrix",
            "From\\To," + ",".join(states)
        ]

        for i, s in enumerate(states):
            row = ",".join(f"{P[i, j]:.10f}" for j in range(len(states)))
            rows.append(f"{s},{row}")

    OUT_MARKOV_CSV.write_text("\n".join(rows))

    # ---------------- EXCEL ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    bold = Font(bold=True)

    ws["A1"] = "Ticker"
    ws["B1"] = TICKER
    ws["A2"] = "GeneratedUTC"
    ws["B2"] = datetime.utcnow().isoformat() + "Z"
    ws["A1"].font = ws["A2"].font = bold

    r = 4
    for b in blocks:
        states = b["states"]
        P = b["P"]

        ws.cell(r, 1, f"ORDER {b['order']}").font = Font(bold=True, size=14)
        r += 2

        ws.cell(r, 1, "StationaryVector").font = bold
        r += 1
        ws.cell(r, 1, "State").font = bold
        ws.cell(r, 2, "Pi").font = bold
        r += 1

        for s, p in zip(states, b["pi"]):
            ws.cell(r, 1, s)
            ws.cell(r, 2, p)
            r += 1

        r += 1
        ws.cell(r, 1, "TransitionMatrix").font = bold
        r += 1

        ws.cell(r, 1, "From\\To").font = bold
        for j, s_to in enumerate(states, start=2):
            ws.cell(r, j, s_to).font = bold
            ws.cell(r, j).alignment = Alignment(text_rotation=90)
        r += 1

        for i, s_from in enumerate(states):
            ws.cell(r, 1, s_from).font = bold
            for j in range(len(states)):
                ws.cell(r, j + 2, P[i, j])
            r += 1

        r += 3

    wb.save(OUT_MARKOV_XLSX)
    subprocess.run(["open", OUT_MARKOV_XLSX], check=False)

# ============================================================
# MAIN
# ============================================================

def main():
    digits = fetch_history()

    blocks = []
    for order in (2, 3):
        states, P, pi, current, probs = markov_chain(digits, order)
        blocks.append(dict(
            order=order,
            states=states,
            P=P,
            pi=pi,
            current=current,
            probs=probs
        ))

    save_results(blocks)
    print("Done.")

if __name__ == "__main__":
    main()