#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
from openpyxl import load_workbook

TRADES_XLSX = "/Users/kostayanev/NikyClean/TRADES.xlsx"
CSV_PATH    = "/Users/kostayanev/NikyClean/Transactions.csv"
SHEET_NAME  = "Transactions"
TABLE_NAME  = "TabTrans"

def export_tabtrans_to_csv():
    wb = load_workbook(TRADES_XLSX, data_only=True)
    ws = wb[SHEET_NAME]
    table = ws.tables[TABLE_NAME]

    ref = table.ref
    rows = ws[ref]

    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow([cell.value for cell in row])

    print(f"✔ Exported TabTrans → {CSV_PATH}")

if __name__ == "__main__":
    export_tabtrans_to_csv()