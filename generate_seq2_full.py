#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE  = "MarkovStationary_Random5.xlsm"
INPUT_SHEET = "Sheet2"
INPUT_CELL  = "D1"
OUTPUT_FILE = "Markov_seq2_full.xlsx"

digits = [str(d) for d in range(10)]

# ============================================================
# 1. LOAD THE 200-DIGIT TEXT
# ============================================================

wb = load_workbook(INPUT_FILE, data_only=True)
ws = wb[INPUT_SHEET]

raw = ws[INPUT_CELL].value
if raw is None:
    raise ValueError(f"ERROR: Cell {INPUT_SHEET}!{INPUT_CELL} is empty.")

txt = str(raw).strip()
if not txt.isdigit():
    raise ValueError(f"Cell {INPUT_SHEET}!{INPUT_CELL} must contain ONLY digits.")

n = len(txt)
print(f"Loaded {n}-digit number.")

# Reverse the string to simplify right→left alignment
rev = txt[::-1]

# ============================================================
# 2. FUNCTION: compute all required data for given m
# ============================================================

def compute_for_m(m):

    Npairs = n - m
    if Npairs <= 0:
        raise ValueError(f"ERROR: m={m} too large for n={n}")

    # Step A — Raw pairs
    states = [rev[i:i+m] for i in range(Npairs)]
    uppers = [rev[i+m] for i in range(Npairs)]

    df_pairs = pd.DataFrame({"state": states, "upper": uppers})

    # Step B — Unique states
    unique_states = df_pairs["state"].unique()

    rows_conditional = []

    # Build conditional distributions p(i|j) for each unique state
    for st in unique_states:
        sub = df_pairs[df_pairs["state"] == st]["upper"]
        total = len(sub)

        k = {d: (sub == d).sum() for d in digits}
        p = {d: (k[d] / total) if total > 0 else 0 for d in digits}

        row = {"state": st}
        row.update(p)
        rows_conditional.append(row)

    df_conditional = pd.DataFrame(rows_conditional)
    df_conditional = df_conditional[["state"] + digits]

    # Step C — Final P(i) by averaging conditional rows
    Pm = {d: df_conditional[d].mean() for d in digits}
    final_row = [Pm[d] for d in digits]

    return df_pairs, df_conditional, final_row


# ============================================================
# 3. MAIN EXECUTION — compute for m=1..5 and assemble output
# ============================================================

final_rows = []

# Using pandas ExcelWriter to create multiple sheets
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # Sheet 1: final combined table
    for m in range(1, 6):
        print(f"Processing m={m} ...")
        df_pairs, df_cond, row = compute_for_m(m)
        final_rows.append(row)

        # Write per-m detailed sheets
        df_cond.to_excel(writer, sheet_name=f"UniqueStates_m{m}", index=False)
        df_pairs.to_excel(writer, sheet_name=f"StateUpperPairs_m{m}", index=False)

    # Write the final combined 5×10 table
    df_final = pd.DataFrame(final_rows, columns=digits)
    df_final.index = [f"m={m}" for m in range(1, 6)]
    df_final.to_excel(writer, sheet_name="ConditionalProbabilities")

print(f"SUCCESS: Created {OUTPUT_FILE} with full analysis.")