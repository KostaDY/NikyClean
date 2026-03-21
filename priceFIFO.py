import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ============================================================
# SETTINGS
# ============================================================

INPUT_WB  = "Solid_REG_ALL.xlsm"
NAMED_RANGE = "Register"

OUTPUT_WB = "priceFIFO.xlsx"
OUTPUT_WS = "PriceFIFO"

# ============================================================
# LOAD NAMED RANGE (READ-ONLY)
# ============================================================

wb = load_workbook(INPUT_WB, data_only=True, read_only=True)

if NAMED_RANGE not in wb.defined_names:
    raise ValueError(f'Named range "{NAMED_RANGE}" not found')

defn = wb.defined_names[NAMED_RANGE]
sheet_name, ref = next(defn.destinations)
ws = wb[sheet_name]

rows = []
for row in ws[ref]:
    rows.append([cell.value for cell in row])

# ============================================================
# PARSE TICKER GROUPS
# ============================================================

groups = {}           # ticker -> list of (date, number, price)
current_ticker = None

for r in rows:
    if len(r) < 3:
        continue

    date, number, price = r[0], r[1], r[2]

    # Ticker header row
    if isinstance(date, str) and isinstance(number, str) and isinstance(price, str):
        current_ticker = date.strip()
        groups[current_ticker] = []
        continue

    # Data row
    if current_ticker is not None and date is not None:
        groups[current_ticker].append((
            pd.to_datetime(date),
            int(number),
            float(price)
        ))

# ============================================================
# PRICE-FIFO FUNCTION (GLOBAL, PRICE-CENTRIC)
# ============================================================

def price_fifo(trades):
    """
    trades: list of (date, number, price)
    returns: DataFrame with remaining purchase lots
    """

    df = pd.DataFrame(trades, columns=["Date", "Number", "Price"])

    purchases = df[df["Price"] < 0].copy()
    sales     = df[df["Price"] > 0].copy()

    if purchases.empty:
        return pd.DataFrame(columns=["Date", "Number", "Price"])

    # Sort
    purchases["AbsPrice"] = -purchases["Price"]
    purchases.sort_values(["AbsPrice", "Date"], inplace=True)

    sales.sort_values("Price", inplace=True)

    purchases["Remaining"] = purchases["Number"]
    sales["Remaining"]     = sales["Number"]

    p_idx = s_idx = 0

    while p_idx < len(purchases) and s_idx < len(sales):
        p_rem = purchases.iloc[p_idx]["Remaining"]
        s_rem = sales.iloc[s_idx]["Remaining"]

        take = min(p_rem, s_rem)

        purchases.iat[p_idx, purchases.columns.get_loc("Remaining")] -= take
        sales.iat[s_idx, sales.columns.get_loc("Remaining")] -= take

        if purchases.iloc[p_idx]["Remaining"] == 0:
            p_idx += 1
        if sales.iloc[s_idx]["Remaining"] == 0:
            s_idx += 1

    out = purchases[purchases["Remaining"] > 0][
        ["Date", "Remaining", "Price"]
    ].rename(columns={"Remaining": "Number"})

    return out

# ============================================================
# BUILD SINGLE OUTPUT SHEET (WITH TICKER SEPARATORS)
# ============================================================

output_rows = []

for ticker, trades in groups.items():
    if not trades:
        continue

    # Ticker separator row
    output_rows.append([ticker, None, None])

    fifo_df = price_fifo(trades)
    for _, r in fifo_df.iterrows():
        output_rows.append([
            r["Date"],
            int(r["Number"]),
            float(r["Price"])
        ])

# ============================================================
# WRITE OUTPUT WORKBOOK
# ============================================================

out_df = pd.DataFrame(output_rows, columns=["Date", "Number", "Price"])

with pd.ExcelWriter(OUTPUT_WB, engine="openpyxl", mode="w") as writer:
    out_df.to_excel(writer, sheet_name=OUTPUT_WS, index=False)

print(f"✅ Written: {Path(OUTPUT_WB).resolve()}")